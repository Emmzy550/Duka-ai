from __future__ import annotations

import html
import json
import os
import re
import textwrap
from pathlib import Path
from typing import Any

import datetime
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import streamlit.components.v1 as components

from agents import _print_provider_banner
from agents.followup_agent import (
    AGENT_LABELS,
    answer_followup_question,
    route_question,
    stream_followup_for_agent,
)

_print_provider_banner()
from agents.orchestrator import generate_business_report
from agents.visualization_agent import build_chart_data, build_followup_visual
from config import get_settings
from tools.document_parser import DOCUMENT_TYPES, parse_uploaded_file, summarize_document_records
from tools.financial_calculator import format_currency
from tools.financial_engine import (
    SCENARIOS as FE_SCENARIOS,
    calculate_forecast,
    compute_expense_breakdown,
    compute_metrics,
    get_baseline,
    get_loan_label,
)
from tools.text_parser import parse_business_input
from tools.transaction_analyzer import analyze_transactions, load_transactions_frame


ROOT_DIR = Path(__file__).resolve().parent
DATA_DIR = ROOT_DIR / "data"
SAMPLE_CASES_PATH = DATA_DIR / "sample_business_cases.json"
SAMPLE_TRANSACTIONS_PATH = DATA_DIR / "sample_transactions.csv"
SAMPLE_TEMPLATES = {
    "Income Statement CSV": DATA_DIR / "sample_income_statement.csv",
    "Cash Flow Record CSV": DATA_DIR / "sample_cash_flow_record.csv",
    "Sales Record CSV": DATA_DIR / "sample_sales_record.csv",
    "Expense Record CSV": DATA_DIR / "sample_expense_record.csv",
    "Mobile Money Statement CSV": DATA_DIR / "sample_mobile_money_statement.csv",
    "General Business Workbook": DATA_DIR / "sample_business_records.xlsx",
}

THEME = {
    "bg": "#0F172A",
    "bg_alt": "#111827",
    "card": "#F8FAFC",
    "white": "#FFFFFF",
    "text": "#0F172A",
    "muted": "#64748B",
    "border": "#E5E7EB",
    "emerald": "#10B981",
    "blue": "#2563EB",
    "gold": "#F59E0B",
    "red": "#EF4444",
    "orange": "#F97316",
}

DEMO_PROVIDERS = [
    ("MTN Mobile Money", "#FFCB05", "#111827", "MTN"),
    ("Airtel Money", "#EF4444", "#FFFFFF", "Airtel"),
    ("Zamtel Money", "#10B981", "#FFFFFF", "Zamtel"),
    ("Bank Account", "#2563EB", "#FFFFFF", "Bank"),
]

PROVIDER_BRANDING = {
    "MTN Mobile Money": {"abbr": "MTN", "bg": "#FFCB05", "fg": "#111827"},
    "Airtel Money": {"abbr": "Airtel", "bg": "#EF4444", "fg": "#FFFFFF"},
    "Zamtel Money": {"abbr": "Zamtel", "bg": "#10B981", "fg": "#FFFFFF"},
    "Bank Account": {"abbr": "Bank", "bg": "#2563EB", "fg": "#FFFFFF"},
    "Upload Statement": {"abbr": "DOC", "bg": "#64748B", "fg": "#FFFFFF"},
    "Manual Entry": {"abbr": "TXT", "bg": "#7C3AED", "fg": "#FFFFFF"},
}

PROVIDER_DETAILS = {
    "MTN Mobile Money": {
        "kind": "Mobile money wallet",
        "signal": "Airtime, customer payments, merchant flows",
        "demo": "47 demo transactions",
    },
    "Airtel Money": {
        "kind": "Mobile money wallet",
        "signal": "Agent cash-ins, transfers, daily receipts",
        "demo": "47 demo transactions",
    },
    "Zamtel Money": {
        "kind": "Mobile money wallet",
        "signal": "Wallet activity, recurring payments",
        "demo": "47 demo transactions",
    },
    "Bank Account": {
        "kind": "Bank statement",
        "signal": "Deposits, supplier payments, debt service",
        "demo": "Statement demo",
    },
}

FOLLOWUP_SUGGESTIONS = [
    "Where is my money going?",
    "Am I ready for a loan?",
    "What should I reduce?",
    "How do I grow revenue?",
    "Is this a good season?",
    "Explain my cash flow",
]

EXAMPLE_PROMPT_LABELS = {
    "Lusaka Grocery Shop": "🏪 Lusaka Grocery",
    "Salon Slow Month": "💇 Salon Slow Month",
    "Farmer Restocking Decision": "🌾 Farmer Decision",
}

FOLLOWUP_SUGGESTION_PILLS = [
    ("💸 Where's my money going?", "Where is my money going?"),
    ("🏦 Am I ready for a loan?", "Am I ready for a loan?"),
    ("✂️ What should I reduce?", "What should I reduce?"),
    ("📈 How do I grow revenue?", "How do I grow revenue?"),
    ("🌦️ Is this a good season?", "Is this a good season?"),
    ("🧾 Explain my cash flow", "Explain my cash flow"),
]

NAV_SECTIONS = {
    "MAIN": ["💬 Chat Advisor", "📊 Dashboard"],
    "TOOLS": ["📈 Cash Flow Forecast", "🔮 Scenario Planner", "🏦 Loan Calculator", "🔍 Expense Analyzer"],
    "REPORTS": ["📄 Generate Report", "🌍 Market Intel"],
    "SETTINGS": ["⚙️ Settings"],
}

NAV_LABEL_TO_PAGE = {
    "💬 Chat Advisor": "Chat Advisor",
    "📊 Dashboard": "Dashboard",
    "📈 Cash Flow Forecast": "Cash Flow Forecast",
    "🔮 Scenario Planner": "Scenario Planner",
    "🏦 Loan Calculator": "Loan Calculator",
    "🔍 Expense Analyzer": "Expense Analyzer",
    "📄 Generate Report": "Generate Report",
    "🌍 Market Intel": "Market Intel",
    "⚙️ Settings": "Settings",
}

NUMBER_SOURCE_OPTIONS = [
    ("📄 Upload file", "upload"),
    ("🔗 Connect account", "connect"),
    ("✏️ Enter manually", "manual"),
]

AGENT_STATUS_MESSAGES = {
    "cashflow": ("📊", "Cash Flow Agent", "Analyzing revenue, expenses, and profit margins..."),
    "advisor": ("💡", "Business Advisor Agent", "Generating specific recommendations for your business..."),
    "loan": ("🏦", "Loan Readiness Agent", "Calculating your credit score and borrowing capacity..."),
    "market": ("📈", "Market Intelligence Agent", "Checking Zambian market trends and opportunities..."),
    "summary": ("🔄", "Executive Summary Agent", "Writing your personalized financial summary..."),
}


def _normalize_assistant_response(text: str) -> str:
    """Tidy up agent replies before display/storage.

    - Strip trailing 'Next step' lines that just rephrase a question the
      agent already asked (so we don't get redundant Q + Next step pairs).
    - Force the 'Next step:' line onto its own paragraph with a blank line
      above it, so markdown renders it as a separate paragraph.
    - Wrap a clean Next-step line in a styled callout div so it pops
      visually (green left bar) instead of bleeding into the body text.
    """
    if not text:
        return text

    body = text.rstrip()

    # Find the LAST occurrence of "Next step:" — tolerate optional bold markers
    # and tolerate the model putting it directly after a sentence with no space.
    m = re.search(
        r"(?:\n+|(?<=[\.\?\!\)])\s*|^)\s*(?:\*\*\s*)?Next\s*step\s*:\s*(?:\*\*)?\s*(.*?)\s*$",
        body,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if not m:
        return text

    next_step_text = m.group(1).strip()
    pre = body[:m.start()].rstrip()

    # If the body already ends with a question (the agent is asking the user
    # for info), drop the redundant Next step entirely.
    if pre.endswith("?"):
        return pre

    # If the next step itself is just a question rephrasing the same thing, drop it.
    if next_step_text.endswith("?"):
        return pre

    # Otherwise render as a styled callout block.
    safe_text = next_step_text.replace("<", "&lt;").replace(">", "&gt;")
    callout = f'<div class="next-step-callout"><strong>Next step:</strong>{safe_text}</div>'
    return f"{pre}\n\n{callout}"


def _resolve_agents(question: str, messages: list[dict]) -> list[str]:
    """LLM-route the question, fall back to keyword routing if LLM unavailable.

    Handles 'continuation' by returning the last agent that responded.
    """
    recent: list[dict[str, str]] = []
    for i, msg in enumerate(messages):
        if msg["role"] == "user" and i + 1 < len(messages):
            nxt = messages[i + 1]
            if nxt["role"] == "assistant" and nxt.get("type") != "initial_analysis":
                recent.append({"question": msg["content"], "answer": nxt.get("content", "")})
    recent = recent[-3:]

    routed = route_question(question, recent)

    if routed == ["continuation"]:
        last = st.session_state.get("last_agent", "advisor")
        return [last]

    return routed


def load_sample_cases() -> list[dict[str, str]]:
    try:
        return json.loads(SAMPLE_CASES_PATH.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def load_template_bytes(path: Path) -> bytes:
    try:
        return path.read_bytes()
    except FileNotFoundError:
        return b""


def empty_parsed_data(raw_text: str = "") -> dict[str, Any]:
    return {
        "raw_text": raw_text,
        "revenue": 0.0,
        "expenses": 0.0,
        "expenses_breakdown": {
            "stock": 0.0,
            "rent": 0.0,
            "transport": 0.0,
            "other_expenses": 0.0,
        },
        "debt": 0.0,
        "loan_amount": 0.0,
        "business_type": None,
        "location": None,
        "categories_detected": [],
    }


def get_loan_band(score: int, profit_margin: float = 0.0) -> tuple[str, str, str]:
    """Single consistent loan label — always margin-based via financial_engine."""
    margin = profit_margin if profit_margin > 0 else 0.0
    label, color, _ = get_loan_label(margin)
    explanations = {
        "Loan ready": "Borrowing looks manageable if it serves a clear business purpose.",
        "Loan possible with caution": "Small, targeted borrowing may be feasible — keep repayments below 20% of monthly profit.",
        "Limited borrowing capacity": "Profit margin is too thin to safely absorb loan repayments right now — improve margins first.",
    }
    return (label, color, explanations[label])


def get_provider_badge_html(provider: str, large: bool = False) -> str:
    branding = PROVIDER_BRANDING.get(provider, {"abbr": provider[:4], "bg": "#334155", "fg": "#FFFFFF"})
    size_class = "provider-badge-large" if large else ""
    return (
        f'<div class="provider-badge {size_class}" '
        f'style="background:{branding["bg"]}; color:{branding["fg"]};">{branding["abbr"]}</div>'
    )


def get_connected_demo(provider: str | None) -> dict[str, Any] | None:
    if not provider:
        return None
    frame = load_transactions_frame(str(SAMPLE_TRANSACTIONS_PATH))
    filtered = frame[frame["Provider"] == provider].copy()
    if filtered.empty:
        return None
    summary = analyze_transactions(filtered)
    summary["provider_selected"] = provider
    return summary


def build_business_context_text(profile: dict[str, str], manual_notes: str) -> str:
    parts = []
    if profile.get("business_type"):
        parts.append(f"Business type: {profile['business_type']}.")
    if profile.get("location"):
        parts.append(f"Location: {profile['location']}.")
    if profile.get("products_services"):
        parts.append(f"Main products or services: {profile['products_services']}.")
    if profile.get("main_question"):
        parts.append(f"Main question: {profile['main_question']}.")
    if manual_notes.strip():
        parts.append(manual_notes.strip())
    return " ".join(parts).strip()


def combine_analysis_inputs(
    business_profile: dict[str, str],
    manual_notes: str,
    document_data: dict[str, Any] | None,
    connected_data: dict[str, Any] | None,
    manual_figures: dict[str, float | int] | None = None,
) -> tuple[str, dict[str, Any], list[str], list[str], list[str]]:
    business_context = build_business_context_text(business_profile, manual_notes)
    parsed_text = parse_business_input(business_context) if business_context else empty_parsed_data()
    combined = dict(parsed_text)
    combined["business_type"] = business_profile.get("business_type") or combined.get("business_type")
    combined["location"] = business_profile.get("location") or combined.get("location")

    if document_data:
        document_metadata = document_data.get("metadata", {})
        combined["business_type"] = combined.get("business_type") or document_metadata.get("business_type")
        combined["location"] = combined.get("location") or document_metadata.get("location")

    basis_notes: list[str] = []
    source_labels: list[str] = []
    data_sources: list[str] = []

    if any(value.strip() for value in business_profile.values()):
        basis_notes.append("Analysis used the business profile you provided.")
        source_labels.append("Business context")
        data_sources.append("Business context")

    if manual_notes.strip():
        basis_notes.append("Additional context was included from your written description.")
        source_labels.append("Manual text input")
        data_sources.append("Manual text input")

    figures_provided = any(float(manual_figures.get(key, 0) or 0) > 0 for key in ["revenue", "expenses", "debt", "staff"]) if manual_figures else False
    if figures_provided:
        basis_notes.append("Analysis used the manual figures you entered.")
        source_labels.append("Manual figures")
        data_sources.append("Manual figures")

    if connected_data:
        combined["revenue"] = round(float(connected_data["total_revenue"]), 2)
        combined["expenses"] = round(float(connected_data["total_expenses"]), 2)
        if combined.get("debt", 0.0) <= 0 and connected_data.get("debt_payments", 0.0) > 0:
            combined["debt"] = round(float(connected_data["debt_payments"]), 2)
        basis_notes.append("Analysis used simulated connected transaction data.")
        source_labels.append("Connected financial data demo")
        data_sources.append(f"Simulated connection: {connected_data['provider_selected']}")

    if document_data:
        if not connected_data and document_data.get("revenue", 0) > 0:
            combined["revenue"] = round(float(document_data["revenue"]), 2)
        if not connected_data and document_data.get("expenses", 0) > 0:
            combined["expenses"] = round(float(document_data["expenses"]), 2)
        if combined.get("debt", 0.0) <= 0 and document_data.get("debt", 0) > 0:
            combined["debt"] = round(float(document_data["debt"]), 2)
        # Carry through expense breakdown extracted by the LLM parser
        doc_breakdown = document_data.get("expenses_breakdown") or {}
        if doc_breakdown and not any(v > 0 for v in combined.get("expenses_breakdown", {}).values()):
            combined["expenses_breakdown"] = doc_breakdown
        basis_notes.append(f"Analysis used an uploaded {document_data.get('document_type', 'document').lower()}.")
        source_labels.append(f"Uploaded document: {document_data.get('document_type', 'Other')}")
        data_sources.append(f"Uploaded document: {document_data.get('document_type', 'Other')}")

    if manual_figures:
        manual_revenue = round(float(manual_figures.get("revenue", 0.0) or 0.0), 2)
        manual_expenses = round(float(manual_figures.get("expenses", 0.0) or 0.0), 2)
        manual_debt = round(float(manual_figures.get("debt", 0.0) or 0.0), 2)
        if manual_revenue > 0 and not connected_data and not (document_data and document_data.get("revenue", 0) > 0):
            combined["revenue"] = manual_revenue
        if manual_expenses > 0 and not connected_data and not (document_data and document_data.get("expenses", 0) > 0):
            combined["expenses"] = manual_expenses
        if manual_debt > 0 and combined.get("debt", 0.0) <= 0:
            combined["debt"] = manual_debt

    if not basis_notes:
        basis_notes.append("Analysis used the written business information that was provided.")
        source_labels.append("Manual text input")
        data_sources.append("Manual text input")

    combined["raw_text"] = business_context
    combined["categories_detected"] = list(dict.fromkeys(combined.get("categories_detected", [])))

    context_parts: list[str] = []
    if any(value.strip() for value in business_profile.values()):
        context_parts.append("Business profile: " + "; ".join(f"{key.replace('_', ' ')}: {value}" for key, value in business_profile.items() if value.strip()))
    if manual_notes.strip():
        context_parts.append("Manual business notes: " + manual_notes.strip())
    if figures_provided and manual_figures:
        figure_parts = []
        if float(manual_figures.get("revenue", 0) or 0) > 0:
            figure_parts.append(f"Revenue: {format_currency(float(manual_figures['revenue']))}")
        if float(manual_figures.get("expenses", 0) or 0) > 0:
            figure_parts.append(f"Expenses: {format_currency(float(manual_figures['expenses']))}")
        if float(manual_figures.get("debt", 0) or 0) > 0:
            figure_parts.append(f"Debt: {format_currency(float(manual_figures['debt']))}")
        if int(manual_figures.get("staff", 0) or 0) > 0:
            figure_parts.append(f"Staff: {int(manual_figures['staff'])}")
        if figure_parts:
            context_parts.append("Manual figures: " + "; ".join(figure_parts))
    if document_data:
        context_parts.append("Uploaded document summary: " + summarize_document_records(document_data))
    if connected_data:
        context_parts.append("Connected transaction summary: " + connected_data.get("summary", "Transaction patterns were analyzed."))

    analysis_context = "\n\n".join(part for part in context_parts if part).strip()
    if not analysis_context:
        analysis_context = "Business information was provided for analysis."

    data_sources.append("Market Intelligence")
    return (
        analysis_context,
        combined,
        basis_notes,
        list(dict.fromkeys(source_labels)),
        list(dict.fromkeys(data_sources)),
    )


# ─── Render helpers ────────────────────────────────────────────────────────────

def render_section_intro(title: str, subtitle: str) -> None:
    st.markdown(
        f"""
        <div class="section-intro">
            <div class="section-intro-title">{title}</div>
            <div class="section-intro-subtitle">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _loading_card(
    title: str,
    subtitle: str = "",
    indicator: str = "ring",   # "ring" | "dots"
    badge: str = "Duka AI",
) -> str:
    """Return themed HTML for a branded loading card (render with unsafe_allow_html=True).

    Use inside an st.empty() placeholder that you clear once the operation completes.
    indicator="ring"  → spinning teal ring
    indicator="dots"  → three bouncing dots
    """
    if indicator == "dots":
        ind_html = '<div class="duka-dots"><span></span><span></span><span></span></div>'
    else:
        ind_html = '<div class="duka-ring"></div>'
    sub_html = f'<div class="duka-loader-sub">{subtitle}</div>' if subtitle else ""
    badge_html = f'<div class="duka-loader-badge">{badge}</div>' if badge else ""
    return (
        f'<div class="duka-loader">'
        f'{ind_html}'
        f'<div class="duka-loader-body">'
        f'<div class="duka-loader-title">{title}</div>'
        f'{sub_html}'
        f'</div>'
        f'{badge_html}'
        f'</div>'
    )


def _skeleton_lines(widths: list[str] | None = None) -> str:
    """Return HTML skeleton shimmer lines for content-is-loading placeholders."""
    classes = widths or ["wide", "med", "short"]
    lines = "".join(f'<div class="duka-skeleton duka-skeleton-line {w}"></div>' for w in classes)
    return f'<div style="padding:8px 0;">{lines}</div>'


def _profit_label_value(profit: float, margin: float | None = None) -> tuple[str, str, str]:
    """Return (label, formatted_value, accent_color) for a profit/loss figure.

    Positive profit → "Profit", teal/blue display.
    Negative profit → "Net Loss", red display with positive K amount.
    """
    if profit < 0:
        label = "Net Loss"
        value = f"K{abs(profit):,.2f}"
        color = "#E24B4A"
    else:
        label = "Profit"
        value = format_currency(profit)
        color = THEME["blue"]
    if margin is not None:
        value = f"{value}  ({margin:.1f}% margin)"
    return label, value, color


def render_metric_card(label: str, value: str, note: str, accent: str, icon: str) -> None:
    st.markdown(
        f"""
        <div class="metric-card" style="border-top: 4px solid {accent};">
            <div class="metric-row">
                <span class="metric-icon">{icon}</span>
                <span class="metric-label">{label}</span>
            </div>
            <div class="metric-value">{value}</div>
            <div class="metric-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_metric_grid(report: dict[str, Any]) -> None:
    cashflow = report["cashflow"]
    business_health = report["business_health"]
    profit = cashflow["profit"]
    p_label, p_value, p_color = _profit_label_value(profit)
    p_note = "What remains after expenses." if profit >= 0 else "Expenses exceed revenue — this business is at a loss."
    items = [
        ("Revenue", format_currency(cashflow["revenue"]), "Money coming into the business.", THEME["emerald"], "K"),
        ("Expenses", format_currency(cashflow["expenses"]), "Costs and outgoing spending.", THEME["orange"], "#"),
        (p_label, p_value, p_note, p_color, "+" if profit >= 0 else "⚠"),
        ("Profit Margin", f'{cashflow["profit_margin"]:.1f}%', "Profit as a share of revenue.", THEME["gold"], "%"),
        ("Business Health", f'{business_health["score"]}/100', f'{business_health["status"]} · Overall fitness, not loan score', THEME["emerald"], "*"),
    ]
    for column, item in zip(st.columns(5, gap="medium"), items):
        with column:
            render_metric_card(*item)


def render_prompt_card(case: dict[str, str], key: str) -> None:
    preview = textwrap.shorten(case["input"], width=110, placeholder="...")
    st.markdown(
        f"""
        <div class="prompt-card">
            <div class="prompt-title">{case["title"]}</div>
            <div class="prompt-preview">{preview}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    _spacer, try_col = st.columns([3, 1])
    with try_col:
        if st.button("→ Try this", key=key):
            st.session_state.manual_notes = case["input"]
            st.rerun()


def populate_example_prompt(case: dict[str, str]) -> None:
    parsed = parse_business_input(case["input"])
    prompt_label = EXAMPLE_PROMPT_LABELS.get(case["title"], case["title"])
    st.session_state.manual_notes = case["input"]
    st.session_state.business_type = parsed.get("business_type") or st.session_state.business_type
    st.session_state.location = parsed.get("location") or st.session_state.location

    if not st.session_state.products_services and parsed.get("categories_detected"):
        st.session_state.products_services = ", ".join(parsed["categories_detected"][:3])

    if "?" in case["input"]:
        last_question = case["input"].rsplit("?", 1)[0].split(".")[-1].strip()
        if last_question:
            st.session_state.main_question = f"{last_question}?"

    st.session_state.selected_example_prompt = prompt_label
    st.session_state.selected_example_prompt_text = case["input"]
    st.session_state.scroll_target = "duka-ai-analyze-anchor"


def render_example_prompt_pills(sample_cases: list[dict[str, str]]) -> None:
    _CARD_META: dict[str, dict] = {
        "Lusaka Grocery Shop": {
            "icon": "🏪", "icon_bg": "rgba(29,158,117,0.2)",
            "desc": "Weekly sales analysis, cash flow check, loan readiness for a Lusaka trader.",
            "featured": True,
        },
        "Salon Slow Month": {
            "icon": "💇", "icon_bg": "rgba(55,138,221,0.2)",
            "desc": "Ndola salon navigating a slow January with high product costs.",
            "featured": False,
        },
        "Farmer Restocking Decision": {
            "icon": "🌾", "icon_bg": "rgba(245,158,11,0.2)",
            "desc": "Seasonal restocking decisions for a Zambian smallholder farmer.",
            "featured": False,
        },
    }

    prompt_columns = st.columns(max(1, min(3, len(sample_cases) or 1)), gap="small")
    for index, case in enumerate(sample_cases[:3]):
        meta = _CARD_META.get(case["title"], {
            "icon": "📊", "icon_bg": "rgba(100,100,100,0.2)",
            "desc": textwrap.shorten(case["input"], width=90, placeholder="..."),
            "featured": False,
        })
        with prompt_columns[index]:
            if meta["featured"]:
                st.markdown(
                    f"""
                    <div class="sample-card-featured">
                        <span class="recommended-badge">&#11088; Recommended</span>
                        <div class="card-icon" style="background:{meta['icon_bg']};">{meta['icon']}</div>
                        <div style="font-weight:600;color:#E2E8F0;font-size:14px;
                                    margin-bottom:4px;">{case['title']}</div>
                        <div style="color:#6B7280;font-size:12px;line-height:1.5;">
                            {meta['desc']}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                if st.button("Try this →", key=f"example_pill_{index}",
                             use_container_width=True, type="primary"):
                    populate_example_prompt(case)
                    st.rerun()
            else:
                st.markdown(
                    f"""
                    <div class="sample-card-normal">
                        <div class="card-icon" style="background:{meta['icon_bg']};">{meta['icon']}</div>
                        <div style="font-weight:600;color:#E2E8F0;font-size:14px;
                                    margin-bottom:4px;">{case['title']}</div>
                        <div style="color:#6B7280;font-size:12px;line-height:1.5;">
                            {meta['desc']}</div>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                if st.button("Try this →", key=f"example_pill_{index}",
                             use_container_width=True):
                    populate_example_prompt(case)
                    st.rerun()


def render_chat_metrics_bar(cashflow: dict[str, Any]) -> None:
    _profit = cashflow["profit"]
    _p_label, _p_val, _ = _profit_label_value(_profit)
    _p_arrow = " ↑" if _profit > 0 else " ↓" if _profit < 0 else ""
    metric_items = [
        ("Revenue", format_currency(cashflow["revenue"])),
        ("Expenses", format_currency(cashflow["expenses"])),
        (_p_label, f"{_p_val}{_p_arrow}".strip()),
        ("Margin", f'{cashflow.get("profit_margin", 0.0):.1f}%'),
    ]
    metric_columns = st.columns(4, gap="small")
    for column, (label, value) in zip(metric_columns, metric_items):
        with column:
            st.metric(label, value)


def inject_ui_behavior(
    sample_cases: list[dict[str, str]],
    selected_example_prompt: str,
    selected_example_text: str,
    selected_numbers_mode: str,
    scroll_requested: bool,
) -> None:
    example_map = {EXAMPLE_PROMPT_LABELS.get(case["title"], case["title"]): case["input"] for case in sample_cases[:3]}
    suggestion_labels = [label for label, _ in FOLLOWUP_SUGGESTION_PILLS]
    number_labels = [label for label, _ in NUMBER_SOURCE_OPTIONS]
    components.html(
        f"""
        <script>
        const doc = window.parent.document;
        const exampleMap = {json.dumps(example_map)};
        const suggestionLabels = {json.dumps(suggestion_labels)};
        const numberLabels = {json.dumps(number_labels)};
        const selectedExamplePrompt = {json.dumps(selected_example_prompt)};
        const selectedExampleText = {json.dumps(selected_example_text or "Hover over a prompt to preview it here.")};
        const selectedNumbersMode = {json.dumps(selected_numbers_mode)};
        const scrollRequested = {json.dumps(scroll_requested)};

        const findButton = (label) => Array.from(doc.querySelectorAll("button")).find(
            (button) => button.innerText.trim() === label
        );

        const descriptionNode = doc.getElementById("duka-ai-example-description");
        const setDescription = (text) => {{
            if (descriptionNode) {{
                descriptionNode.textContent = text;
            }}
        }};

        Object.entries(exampleMap).forEach(([label, description]) => {{
            const button = findButton(label);
            if (!button) return;
            button.classList.add("duka-ai-example-pill-btn");
            if (label === selectedExamplePrompt) {{
                button.classList.add("is-selected");
            }}
            button.onmouseenter = () => setDescription(description);
            button.onfocus = () => setDescription(description);
            button.onmouseleave = () => setDescription(selectedExampleText);
            button.onblur = () => setDescription(selectedExampleText);
        }});

        suggestionLabels.forEach((label) => {{
            const button = findButton(label);
            if (button) {{
                button.classList.add("duka-ai-suggestion-pill-btn");
            }}
        }});

        numberLabels.forEach((label) => {{
            const button = findButton(label);
            if (!button) return;
            button.classList.add("duka-ai-number-pill-btn");
            const mode = label.includes("Upload") ? "upload" : label.includes("Connect") ? "connect" : "manual";
            if (mode === selectedNumbersMode) {{
                button.classList.add("is-selected");
            }}
        }});

        const analyzeLabel = "⚡ Run Full Business Analysis";
        let analyzeButton = findButton(analyzeLabel);
        if (!analyzeButton) {{
            analyzeButton = findButton("⚡ Analyze My Business →");
        }}
        if (analyzeButton) {{
            analyzeButton.classList.add("duka-ai-analyze-btn");
        }}
        if (scrollRequested) {{
            window.setTimeout(() => {{
                var anchorEl = doc.getElementById("duka-ai-analyze-anchor");
                if (anchorEl) {{
                    anchorEl.scrollIntoView({{ behavior: "smooth", block: "start" }});
                }} else if (analyzeButton) {{
                    analyzeButton.scrollIntoView({{ behavior: "smooth", block: "center" }});
                }}
            }}, 260);
        }}
        </script>
        """,
        height=0,
        width=0,
    )


def render_clickable_provider_cards() -> None:
    """Clickable provider cards with brand colors, selection state, and simulated connection form."""
    selected: set[str] = st.session_state.get("selected_providers", set())
    provider_names = [provider for provider, *_ in DEMO_PROVIDERS]
    preferred_provider = st.session_state.get("connected_provider") or st.session_state.get("provider_picker") or provider_names[0]
    selected_index = provider_names.index(preferred_provider) if preferred_provider in provider_names else 0
    provider_choice = st.selectbox(
        "Choose financial source",
        provider_names,
        index=selected_index,
        key="provider_picker",
        help="Select the mobile money wallet or bank account to preview and connect.",
    )
    chosen_provider, chosen_bg, chosen_fg, chosen_abbr = next(
        item for item in DEMO_PROVIDERS if item[0] == provider_choice
    )
    chosen_details = PROVIDER_DETAILS.get(chosen_provider, {})
    picker_left, picker_right = st.columns([1.7, 0.8], gap="large")
    with picker_left:
        st.markdown(
            f"""
            <div class="provider-dropdown-preview" style="--provider:{chosen_bg}; border-left-color:{chosen_bg};">
                <div class="provider-connected-mark" style="background:{chosen_bg}; color:{chosen_fg};">{chosen_abbr}</div>
                <div>
                    <div class="provider-connected-title">{chosen_provider}</div>
                    <div class="provider-connected-copy">{chosen_details.get("kind", "Financial source")} · {chosen_details.get("signal", "Transaction patterns and cash flow signals")}</div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with picker_right:
        already_connected = chosen_provider in selected
        picker_action = "Disconnect selected" if already_connected else "Connect selected"
        if st.button(picker_action, key="provider_dropdown_action", use_container_width=True, type="primary" if not already_connected else "secondary"):
            new_selected: set[str] = set(selected)
            if already_connected:
                new_selected.discard(chosen_provider)
                if st.session_state.get("connected_provider") == chosen_provider:
                    remaining = new_selected - {chosen_provider}
                    st.session_state.connected_provider = next(iter(remaining), None)
            else:
                new_selected.add(chosen_provider)
                st.session_state.connected_provider = chosen_provider
            st.session_state.selected_providers = new_selected
            st.rerun()

    # Show connection forms for selected providers
    for provider, bg_color, fg_color, abbr in DEMO_PROVIDERS:
        if provider in selected:
            if provider == "Bank Account":
                form_hint = "Upload bank statement CSV — using simulated data for demo"
            else:
                form_hint = f"Phone number registered with {provider} — using simulated data for demo"
            st.markdown(
                f"✅ Connected — 47 transactions loaded from **{provider}** (simulated demo data)\n\n"
                f"""
                <div class="provider-connected-summary" style="border-left-color:{bg_color};">
                    <div class="provider-connected-mark" style="background:{bg_color}; color:{fg_color};">{abbr}</div>
                    <div>
                        <div class="provider-connected-title">{provider} connected</div>
                        <div class="provider-connected-copy">47 simulated transactions loaded. {form_hint}.</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def render_file_preview(document_data: dict[str, Any]) -> None:
    confidence = document_data.get("confidence", "low")
    used_llm = document_data.get("used_llm", False)
    metadata = document_data.get("metadata") if isinstance(document_data.get("metadata"), dict) else {}
    used_structured_sheet_parser = bool(metadata.get("sheets_used"))

    first_warning = (document_data.get("warnings") or [""])[0]
    is_rate_limited = "token limit" in first_warning.lower() or "rate limit" in first_warning.lower() or "429" in first_warning

    if used_llm and confidence == "high":
        st.success(f"✅ LLM successfully extracted financial data from **{document_data.get('document_type', 'document')}** — high confidence.")
    elif used_llm and confidence == "medium":
        st.warning(f"⚠️ LLM partial extraction from **{document_data.get('document_type', 'document')}** — verify figures below.")
    elif used_llm and confidence == "low":
        st.error("❌ LLM could not extract reliable data. Please check the file or enter figures manually.")
    elif used_structured_sheet_parser and confidence == "high":
        st.success(
            f"✅ Structured extraction loaded **{document_data.get('document_type', 'document')}** — high confidence."
        )
    elif is_rate_limited:
        st.error(f"⏳ {first_warning}")
    else:
        st.warning("⚠️ LLM parsing skipped — used keyword-based fallback. Numbers may be inaccurate for multi-sheet files.")

    profit_value = float(document_data.get("profit", 0.0) or 0.0)
    revenue_value = float(document_data.get("revenue", 0.0) or 0.0)
    if revenue_value > 0 or profit_value != 0:
        if profit_value < 0:
            st.error(f"⚠️ This business is making a LOSS of K{abs(profit_value):,.0f}")
        elif profit_value > 0:
            st.success(f"✅ This business is profitable: K{profit_value:,.0f} profit")

    warnings_html = "".join(
        f'<div class="preview-warning">{w}</div>' for w in document_data.get("warnings", [])[:4]
        if "LLM unavailable" not in w and "low confidence" not in w.lower()
    )
    volume = document_data.get("rows_analyzed") or document_data.get("pages_analyzed") or document_data.get("characters_analyzed") or 0
    volume_label = "Rows" if document_data.get("rows_analyzed") else "Pages" if document_data.get("pages_analyzed") else "Characters"
    period = document_data.get("period_description") or document_data.get("date_range") or "Unknown"

    st.markdown(
        f"""
        <div class="preview-card">
            <div class="preview-title">AI Document Analysis</div>
            <div class="preview-grid">
                <div><span>Document Type</span><strong>{document_data.get("document_type", "Other")}</strong></div>
                <div><span>File Name</span><strong>{document_data.get("file_name", "Uploaded file")}</strong></div>
                <div><span>Period</span><strong>{period}</strong></div>
                <div><span>{volume_label} Analyzed</span><strong>{volume}</strong></div>
                <div><span>Revenue</span><strong>{format_currency(document_data.get("revenue", 0.0))}</strong></div>
                <div><span>Expenses</span><strong>{format_currency(document_data.get("expenses", 0.0))}</strong></div>
                <div><span>Profit</span><strong>{format_currency(document_data.get("profit", 0.0))}</strong></div>
                <div><span>Profit Margin</span><strong>{document_data.get("profit_margin", 0.0):.1f}%</strong></div>
                <div><span>Debt</span><strong>{format_currency(document_data.get("debt", 0.0))}</strong></div>
            </div>
            <div class="preview-summary">{document_data.get("summary", "")}</div>
            {warnings_html}
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Monthly breakdown chart
    monthly = document_data.get("monthly_breakdown")
    if monthly and isinstance(monthly, list) and len(monthly) > 1:
        try:
            df_monthly = pd.DataFrame(monthly)
            if "month" in df_monthly.columns:
                df_monthly = df_monthly.set_index("month")
            numeric_cols = [c for c in ["revenue", "expenses", "profit"] if c in df_monthly.columns]
            if numeric_cols:
                st.markdown(
                    '<div class="preview-title" style="margin-top:1rem;color:#0F172A;">Monthly Breakdown</div>',
                    unsafe_allow_html=True,
                )
                colors = {
                    "revenue": THEME["emerald"],
                    "expenses": THEME["orange"],
                    "profit": THEME["blue"],
                }
                fig = go.Figure()
                for col in numeric_cols:
                    fig.add_trace(go.Bar(
                        name=col.title(),
                        x=df_monthly.index.tolist(),
                        y=df_monthly[col].tolist(),
                        marker_color=colors.get(col, THEME["blue"]),
                        text=[format_currency(v) for v in df_monthly[col]],
                        textposition="outside",
                    ))
                fig.update_layout(
                    height=280,
                    barmode="group",
                    margin=dict(l=8, r=8, t=8, b=8),
                    paper_bgcolor="rgba(0,0,0,0)",
                    plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color=THEME["text"], family="Segoe UI, Arial, sans-serif"),
                    xaxis=dict(showgrid=False, title=""),
                    yaxis=dict(gridcolor="rgba(100,116,139,0.18)", zeroline=False),
                    legend=dict(orientation="h", y=1.08),
                    showlegend=True,
                )
                st.plotly_chart(fig, use_container_width=True)
        except Exception:
            pass

    # Key observations
    observations = document_data.get("key_observations", [])
    if observations:
        obs_html = "".join(f'<div class="agent-detail">{obs}</div>' for obs in observations[:4])
        st.markdown(
            f'<div class="preview-card" style="margin-top:0.5rem;">'
            f'<div class="preview-title">Key Observations</div>{obs_html}</div>',
            unsafe_allow_html=True,
        )


def render_connected_preview(connected_data: dict[str, Any]) -> None:
    st.markdown(
        f"""
        <div class="preview-card">
            <div class="preview-title">Connected Financial Data Demo</div>
            <div class="provider-selected">
                {get_provider_badge_html(connected_data["provider_selected"], large=True)}
                <div>
                    <div class="provider-selected-label">Selected provider</div>
                    <div class="provider-selected-name">{connected_data["provider_selected"]}</div>
                </div>
            </div>
            <div class="preview-summary">
                Demo connection enabled. Duka AI is using sample transaction data to simulate secure financial analysis.
            </div>
            <div class="preview-grid">
                <div><span>Transactions</span><strong>{connected_data["rows_analyzed"]}</strong></div>
                <div><span>Total Revenue</span><strong>{format_currency(connected_data["total_revenue"])}</strong></div>
                <div><span>Total Expenses</span><strong>{format_currency(connected_data["total_expenses"])}</strong></div>
                <div><span>Net Cash Flow</span><strong>{format_currency(connected_data["net_cash_flow"])}</strong></div>
                <div><span>Date Range</span><strong>{connected_data["date_range"]}</strong></div>
                <div><span>Biggest Expense</span><strong>{connected_data["biggest_expense_category"]}</strong></div>
                <div><span>Average Daily Sales</span><strong>{format_currency(connected_data["average_daily_sales"])}</strong></div>
                <div><span>Stability Score</span><strong>{connected_data["transaction_stability_score"]}/100</strong></div>
            </div>
            <div class="preview-warning">Your data is used only for analysis in this demo.</div>
            <div class="preview-warning">For the hackathon MVP, this is simulated data. Real integrations would require secure user consent, encryption, and provider approval.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_data_sources_card(report: dict[str, Any]) -> None:
    chips = "".join(
        f'<div class="source-line"><span class="source-check">+</span><span>{source}</span></div>'
        for source in report.get("data_sources", [])
    )
    agent_statuses = "".join(
        f'<div class="source-line"><span class="source-check">{("✓" if used_llm else "F")}</span>'
        f'<span>{name}: {("⚡ Live AI" if used_llm else "Standard mode")}</span></div>'
        for name, used_llm in report.get("agent_statuses", {}).items()
    )
    st.markdown(
        f"""
        <div class="summary-panel">
            <div>
                <div class="summary-kicker">Data Sources Used</div>
                <div class="summary-headline">Duka AI combined your business context, records, transaction patterns, and market context to produce this report.</div>
            </div>
            <div class="source-list">{chips}</div>
        </div>
        <div class="summary-panel" style="margin-top:-0.35rem;">
            <div>
                <div class="summary-kicker">Agent Execution Status</div>
                <div class="summary-headline">Which specialist agents contributed to this analysis.</div>
            </div>
            <div class="source-list">{agent_statuses}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_summary_panel(report: dict[str, Any]) -> None:
    cashflow = report["cashflow"]
    loan = report["loan"]
    business_health = report["business_health"]
    band_label, band_color, _ = get_loan_band(loan["loan_readiness_score"], cashflow.get("profit_margin", 0.0))
    source_html = "".join(f'<div class="source-chip">{label}</div>' for label in report.get("source_labels", []))
    st.markdown(
        f"""
        <div class="summary-panel">
            <div class="summary-copy">
                <div class="summary-kicker">Business Financial Health Summary</div>
                <div class="summary-headline">{cashflow["summary"]}</div>
                <div class="source-chip-row">{source_html}</div>
            </div>
            <div class="summary-statuses">
                <div class="status-card">
                    <span>Business Health</span>
                    <strong>{business_health["status"]}</strong>
                </div>
                <div class="status-card">
                    <span>Cash Flow</span>
                    <strong>{cashflow["cash_flow_status"]}</strong>
                </div>
                <div class="status-card" style="border-color:{band_color};">
                    <span>Borrowing View</span>
                    <strong style="color:{band_color};">{band_label}</strong>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_business_health_card(report: dict[str, Any]) -> None:
    health = report["business_health"]
    color_map = {
        "Healthy": THEME["emerald"],
        "Stable but needs attention": THEME["gold"],
        "At risk": THEME["orange"],
        "Critical": THEME["red"],
    }
    status_color = color_map.get(health["status"], THEME["blue"])
    st.markdown(
        f"""
        <div class="loan-card">
            <div class="card-title">Business Health Score</div>
            <div class="loan-score-row">
                <div class="loan-score">{health["score"]}/100</div>
                <div class="loan-badge" style="color:{status_color}; border-color:{status_color};">{health["status"]}</div>
            </div>
            <div class="progress-shell">
                <div class="progress-fill" style="width:{health["score"]}%; background:{status_color};"></div>
            </div>
            <div class="loan-explainer">Business health is an overall financial fitness score — profitability, margin strength, and expense control. This is <em>separate</em> from the Loan Readiness score below, which specifically measures borrowing safety.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_loan_readiness_card(report: dict[str, Any]) -> None:
    loan = report["loan"]
    cashflow = report.get("cashflow", {})
    label, color, explanation = get_loan_band(loan["loan_readiness_score"], cashflow.get("profit_margin", 0.0))
    risk_warning = loan["warnings"][0] if loan.get("warnings") else "No immediate borrowing warning was detected from the current numbers."
    st.markdown(
        f"""
        <div class="loan-card">
            <div class="card-title">Loan Readiness Score — Borrowing & Debt Review</div>
            <div class="loan-score-row">
                <div class="loan-score">{loan["loan_readiness_score"]}/100</div>
                <div class="loan-badge" style="color:{color}; border-color:{color};">{label}</div>
            </div>
            <div class="progress-shell">
                <div class="progress-fill" style="width:{loan["loan_readiness_score"]}%; background:{color};"></div>
            </div>
            <div class="loan-explainer">{explanation}</div>
            <div class="loan-meta">Safe borrowing amount: <strong>{format_currency(loan["suggested_loan_amount"])}</strong> &nbsp;·&nbsp; calculated as {loan.get("loan_multiplier", 1.5)}× your monthly profit — a conservative SME lending rule of thumb</div>
            <div class="loan-risk-note"><strong>Risk warning:</strong> {risk_warning}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_financial_chart(report: dict[str, Any]) -> None:
    cashflow = report["cashflow"]
    metrics = ["Revenue", "Expenses", "Profit"]
    amounts = [cashflow["revenue"], cashflow["expenses"], cashflow["profit"]]
    colors = [THEME["emerald"], THEME["orange"], THEME["blue"]]
    figure = go.Figure(
        data=[
            go.Bar(
                x=metrics,
                y=amounts,
                marker_color=colors,
                text=[format_currency(value) for value in amounts],
                textposition="outside",
                hovertemplate="%{x}<br>%{text}<extra></extra>",
            )
        ]
    )
    figure.update_layout(
        height=320,
        margin=dict(l=8, r=8, t=8, b=8),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color=THEME["text"], family="Segoe UI, Arial, sans-serif"),
        xaxis=dict(showgrid=False, title=""),
        yaxis=dict(title="Amount (Kwacha)", gridcolor="rgba(100, 116, 139, 0.18)", zeroline=False),
        showlegend=False,
    )
    st.plotly_chart(figure, use_container_width=True)


def render_risk_card(text: str) -> None:
    st.markdown(
        f"""
        <div class="risk-card">
            <div class="risk-icon">!</div>
            <div class="risk-text">{text}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_action_card(index: int, action: str) -> None:
    st.markdown(
        f"""
        <div class="action-card">
            <div class="action-number">{index}</div>
            <div class="action-text">{action}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_financial_advice_section(report: dict[str, Any]) -> None:
    advisor = report["advisor"]
    st.markdown(
        f"""
        <div class="agent-card">
            <div class="card-title">Financial Advice</div>
            <div class="market-subhead">What this means</div>
            <div class="agent-detail">{advisor["what_this_means"]}</div>
            <div class="market-subhead">What you are doing well</div>
            <div class="agent-detail">{advisor["doing_well"]}</div>
            <div class="market-subhead">What needs attention</div>
            <div class="agent-detail">{advisor["needs_attention"]}</div>
            <div class="market-subhead">Best next move</div>
            <div class="agent-detail">{advisor["best_next_move"]}</div>
            <div class="market-subhead">This week's action</div>
            <div class="agent-detail">{advisor["this_weeks_action"]}</div>
            <div class="market-subhead">Growth recommendation</div>
            <div class="agent-detail">{advisor["growth_recommendation"]}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    with st.expander("Detailed reasoning: Financial Advice"):
        st.write(advisor["reasoning"])


def render_cashflow_section(report: dict[str, Any]) -> None:
    cashflow = report["cashflow"]
    st.markdown(
        """
        <div class="agent-card">
            <div class="card-title">Cash Flow & Profit Analysis</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    render_financial_chart(report)
    details = [
        f"Cash flow status: {cashflow['cash_flow_status']}",
        f"Expense ratio: {cashflow['expense_ratio']:.1f}%",
        *cashflow["warnings"],
    ]
    for detail in details:
        st.markdown(f'<div class="agent-detail">{detail}</div>', unsafe_allow_html=True)
    with st.expander("Detailed reasoning: Cash Flow & Profit Analysis"):
        st.write(cashflow["reasoning"])


def render_market_intelligence_section(report: dict[str, Any]) -> None:
    market = report["market"]
    web_used = market.get("web_search_used", False)
    sources  = market.get("sources", [])
    queries  = market.get("queries_used", [])

    # Source badge
    source_badge = (
        f'<span style="font-size:11px;color:#6EE7B7;background:rgba(16,185,129,0.12);'
        f'padding:2px 8px;border-radius:999px;">🌐 Live web search · {len(sources)} sources</span>'
        if web_used else
        '<span style="font-size:11px;color:#94A3B8;background:rgba(148,163,184,0.1);'
        'padding:2px 8px;border-radius:999px;">📚 AI knowledge base</span>'
    )

    risk_cards    = "".join(f'<div class="market-pill">{r}</div>'                    for r in market.get("risk_factors", []))
    monitor_cards = "".join(f'<div class="market-pill market-pill-blue">{m}</div>'   for m in market.get("monitor_next", []))

    st.markdown(
        f"""
        <div class="agent-card">
            <div class="card-title" style="display:flex;justify-content:space-between;align-items:center;">
                <span>Market Intelligence</span>{source_badge}
            </div>
            <div class="agent-summary">{market.get("market_summary", "")}</div>
            <div class="market-subhead">Key market risks</div>
            <div class="market-pill-row">{risk_cards}</div>
            <div class="market-subhead">Opportunity</div>
            <div class="agent-detail">{market.get("opportunity", "")}</div>
            <div class="market-subhead">Recommended move</div>
            <div class="agent-detail">{market.get("recommendation", "")}</div>
            <div class="market-subhead">What to monitor</div>
            <div class="market-pill-row">{monitor_cards}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if web_used and sources:
        with st.expander(f"📚 Sources ({len(sources)} web results used)", expanded=False):
            seen: set[str] = set()
            for src in sources:
                url = src.get("url", "")
                title = src.get("title", url)
                if url and url not in seen:
                    seen.add(url)
                    st.markdown(f"- [{title}]({url})")

    if web_used and queries:
        with st.expander("🔍 Search queries used", expanded=False):
            for q in queries:
                st.caption(f"• {q}")

    if market.get("reasoning"):
        with st.expander("Reasoning: Market Intelligence", expanded=False):
            st.write(market["reasoning"])


def render_transaction_card(transaction_summary: dict[str, Any]) -> None:
    st.markdown(
        f"""
        <div class="mini-card">
            <div class="card-title">Transaction Analysis</div>
            <div class="agent-detail">Average daily sales: {format_currency(transaction_summary["average_daily_sales"])}</div>
            <div class="agent-detail">Biggest expense category: {transaction_summary["biggest_expense_category"]}</div>
            <div class="agent-detail">Recurring expense patterns: {", ".join(transaction_summary["recurring_expense_patterns"]) or "None detected"}</div>
            <div class="agent-detail">Cash flow risk level: {transaction_summary["cash_flow_risk"]}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_followup_visual(visual: dict[str, Any]) -> None:
    rows = "".join(
        f'<div class="visual-row"><span>{item["label"]}</span><strong>{item["value"]}</strong></div>'
        for item in visual.get("items", [])
    )
    st.markdown(
        f"""
        <div class="visual-card">
            <div class="card-title">{visual["title"]}</div>
            {rows}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_followup_chart(chart: dict[str, Any]) -> None:
    chart_type = chart.get("chart_type", "bar")
    title = chart.get("title", "")

    base_layout = dict(
        title=dict(text=title, font=dict(size=14, color=THEME["text"])),
        paper_bgcolor="white",
        plot_bgcolor="white",
        margin=dict(t=48, b=30, l=50, r=20),
        font=dict(family="Inter, sans-serif", color=THEME["text"]),
    )

    if chart_type == "pie":
        fig = go.Figure(go.Pie(
            labels=chart["labels"],
            values=chart["values"],
            hole=0.38,
            marker_colors=["#10B981", "#EF4444", "#2563EB", "#F59E0B", "#7C3AED", "#EC4899"],
            textinfo="label+percent",
            hovertemplate="%{label}: K%{value:,.2f}<extra></extra>",
        ))
        fig.update_layout(**base_layout)

    elif chart_type == "bar":
        raw_colors = chart.get("colors", ["#2563EB"] * len(chart["values"]))
        colors = ["#EF4444" if v < 0 else c for v, c in zip(chart["values"], raw_colors)]
        fig = go.Figure(go.Bar(
            x=chart["labels"],
            y=chart["values"],
            marker_color=colors,
            text=[format_currency(v) for v in chart["values"]],
            textposition="outside",
            hovertemplate="%{x}: K%{y:,.2f}<extra></extra>",
        ))
        fig.update_layout(
            yaxis=dict(title="Amount (Kwacha)", gridcolor="rgba(100,116,139,0.15)", zeroline=True),
            **base_layout,
        )

    elif chart_type == "line":
        fig = go.Figure()
        months = chart.get("months", [])
        for series in chart.get("series", []):
            fig.add_trace(go.Scatter(
                x=months,
                y=series["values"],
                name=series["name"],
                line=dict(color=series["color"], width=2),
                mode="lines+markers",
                hovertemplate=f"%{{x}}: K%{{y:,.2f}}<extra>{series['name']}</extra>",
            ))
        fig.update_layout(
            yaxis=dict(title="Amount (Kwacha)", gridcolor="rgba(100,116,139,0.15)", zeroline=False),
            legend=dict(orientation="h", y=1.08),
            **base_layout,
        )

    else:
        return

    st.plotly_chart(fig, use_container_width=True)


def render_download_buttons() -> None:
    columns = st.columns(3, gap="small")
    template_items = list(SAMPLE_TEMPLATES.items())
    for index, (label, path) in enumerate(template_items):
        with columns[index % 3]:
            mime = "text/csv" if path.suffix == ".csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            st.download_button(
                f"Download {label}",
                data=load_template_bytes(path),
                file_name=path.name,
                mime=mime,
                use_container_width=True,
                key=f"download_{path.stem}",
            )


def split_summary_text(summary: str) -> tuple[str, str]:
    cleaned = " ".join(summary.split()).strip()
    if not cleaned:
        return ("Your analysis is ready.", "")
    sentences = [part.strip() for part in re.split(r"(?<=[.!?])\s+", cleaned) if part.strip()]
    if len(sentences) <= 1:
        return (sentences[0], "")
    return (sentences[0], " ".join(sentences[1:3]))


def render_initial_analysis_in_chat(report: dict[str, Any]) -> None:
    """Clean conversational initial analysis — minimal Claude.ai style."""
    advisor = report["advisor"]
    cashflow = report["cashflow"]
    loan = report["loan"]
    doc_analysis = report.get("document_analysis") or {}
    headline, support = split_summary_text(report["final_summary"])

    # One-line inline metrics strip
    _cf_profit = cashflow["profit"]
    _pl_label, _pl_value, _ = _profit_label_value(_cf_profit)
    metrics_line = (
        f"Revenue {format_currency(cashflow['revenue'])} &nbsp;·&nbsp; "
        f"Expenses {format_currency(cashflow['expenses'])} &nbsp;·&nbsp; "
        f"{_pl_label} {_pl_value} &nbsp;·&nbsp; "
        f"Margin {cashflow.get('profit_margin', 0.0):.1f}%"
    )

    # Natural flowing text — no ALL CAPS headers
    st.markdown(
        f"""
        <div class="analysis-hero">
            <div class="analysis-kicker">Analysis complete</div>
            <div class="analysis-headline">{html.escape(headline)}</div>
            <div class="analysis-support">{html.escape(support or advisor.get("needs_attention", ""))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown('<div class="chat-block-gap"></div>', unsafe_allow_html=True)
    render_chat_metrics_bar(cashflow)
    st.markdown('<div class="chat-block-gap"></div>', unsafe_allow_html=True)
    needs = advisor.get("needs_attention", "").rstrip(".")
    move = advisor.get("best_next_move", "").rstrip(".")
    if needs and move:
        st.markdown(
            f"""
            <div class="analysis-guidance">
                <strong>What to focus on now:</strong> {html.escape(needs)}.
                <br />
                <strong>Best next move:</strong> {html.escape(move)}.
            </div>
            """,
            unsafe_allow_html=True,
        )

    # Full dashboard in expander — keep for detail seekers
    with st.expander("📊 See Full Dashboard — Charts, Loan Score, Market Intelligence", expanded=False):
        render_summary_panel(report)
        render_data_sources_card(report)
        st.markdown("<div style='height:0.5rem;'></div>", unsafe_allow_html=True)

        dash_left, dash_right = st.columns([1.15, 0.85], gap="large")
        with dash_left:
            render_cashflow_section(report)
            render_financial_advice_section(report)
        with dash_right:
            render_business_health_card(report)
            st.markdown("<div style='height:0.8rem;'></div>", unsafe_allow_html=True)
            if report.get("transaction_summary"):
                render_transaction_card(report["transaction_summary"])
                st.markdown("<div style='height:0.8rem;'></div>", unsafe_allow_html=True)
            st.markdown(
                '<div class="mini-card"><div class="card-title">All Recommended Actions</div></div>',
                unsafe_allow_html=True,
            )
            for i, action in enumerate(report["final_recommended_actions"], start=1):
                render_action_card(i, action)

        st.markdown("<div style='height:0.85rem;'></div>", unsafe_allow_html=True)
        lower_left, lower_right = st.columns(2, gap="large")
        with lower_left:
            render_market_intelligence_section(report)
        with lower_right:
            render_loan_readiness_card(report)
            for detail in [
                f"Risk level: {loan['risk_level']}",
                f"Reason: {loan['reason']}",
                *loan["how_to_improve"],
            ]:
                st.markdown(f'<div class="agent-detail">{detail}</div>', unsafe_allow_html=True)
            with st.expander("Detailed reasoning: Borrowing & Debt Review"):
                st.write(loan["reasoning"])

    # Conversational follow-up invitation
    doc_type = doc_analysis.get("document_type", "")
    period = doc_analysis.get("period_description") or doc_analysis.get("date_range") or ""
    if doc_type:
        doc_desc = f"your {doc_type.lower()}" + (f" ({period})" if period else "")
    else:
        doc_desc = "your business"
    st.markdown(
        f'<div class="chat-invite">'
        f"I've analyzed {doc_desc}. "
        f"What would you like to explore first — your expenses, loan readiness, or growth opportunities?"
        f"</div>",
        unsafe_allow_html=True,
    )


def init_session_state() -> None:
    defaults = {
        "business_type": "",
        "location": "",
        "products_services": "",
        "main_question": "",
        "manual_notes": "",
        "manual_revenue": 0.0,
        "manual_expenses": 0.0,
        "manual_debt": 0.0,
        "manual_staff": 0,
        "document_type": "Income Statement",
        "numbers_entry_mode": "upload",
        "selected_example_prompt": "",
        "selected_example_prompt_text": "",
        "scroll_target": None,
        "connected_provider": None,
        "selected_providers": set(),
        "provider_picker": "MTN Mobile Money",
        "analysis_report": None,
        "baseline": None,
        "metrics": None,
        "messages": [],
        "last_agent": "advisor",
        "active_page": "Chat Advisor",
        "show_welcome": True,
        "form_prefilled": False,
        "scroll_to_numbers": False,
        "scroll_to_results": False,
        "scroll_to_chat_conversation": False,
        "pending_demo_analysis_run": False,
        "scroll_to_demo_analyze": False,
        "hide_example_prompts": False,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)


class _LocalUploadShim:
    """Mimic Streamlit's UploadedFile API so bundled samples can flow through
    `parse_uploaded_file` without an actual file_uploader event."""

    def __init__(self, path: Path) -> None:
        self._bytes = path.read_bytes()
        self.name = path.name
        self.size = len(self._bytes)

    def getvalue(self) -> bytes:
        return self._bytes


def _load_demo_analysis() -> bool:
    """Load a bundled sample document and prep all session state required by
    the analysis pipeline. Sets `auto_run_demo` so the main page kicks off
    Run Full Business Analysis automatically on the next render.

    Returns True on success, False if the sample file is unavailable.
    """
    demo_path = SAMPLE_TEMPLATES.get("Income Statement CSV")
    if demo_path is None or not Path(demo_path).exists():
        return False

    shim = _LocalUploadShim(Path(demo_path))
    try:
        parsed = parse_uploaded_file(shim, "Income Statement")
    except Exception:
        return False

    for _k in (
        "baseline", "metrics", "parsed_document",
        "manual_revenue", "manual_expenses", "manual_profit",
        "document_revenue", "document_expenses",
        "generated_report", "analysis_report",
    ):
        st.session_state.pop(_k, None)
    for _k in [k for k in st.session_state if k.startswith("forecast_summary_")]:
        del st.session_state[_k]
    st.session_state["forecast_chat"] = []

    st.session_state["business_type"] = "Grocery shop"
    st.session_state["location"] = "Lusaka"
    st.session_state["products_services"] = "Groceries, mealie meal, drinks"
    st.session_state["main_question"] = "Should I restock or save cash?"
    st.session_state["manual_notes"] = (
        "I run a small grocery shop in Lusaka. This week I made K4,500 in sales. "
        "I spent K2,700 on stock, K500 on rent, K250 on transport, and K300 on other expenses. "
        "I owe K1,000 to my supplier."
    )
    st.session_state["manual_debt"] = 1000.0

    st.session_state["parsed_document"] = parsed
    doc_rev = float(parsed.get("revenue") or 0)
    doc_exp = float(parsed.get("expenses") or 0)
    if doc_rev > 0 or doc_exp > 0:
        st.session_state["manual_revenue"] = doc_rev
        st.session_state["manual_expenses"] = doc_exp
        baseline = {
            "monthly_revenue":  doc_rev,
            "monthly_expenses": doc_exp,
            "monthly_profit":   doc_rev - doc_exp,
            "months_of_data":   int(parsed.get("months_of_data") or 1),
            "source": (
                f"Demo sample: {parsed.get('document_type', 'document')} "
                f"({Path(demo_path).name})"
            ),
            "expenses_breakdown": parsed.get("expenses_breakdown", {}),
        }
        st.session_state["baseline"] = baseline
        st.session_state["metrics"] = compute_metrics(baseline)

    st.session_state["_doc_fingerprint"] = f"{shim.name}:{shim.size}"
    st.session_state["numbers_entry_mode"] = "upload"
    st.session_state["hide_example_prompts"] = True
    st.session_state["show_welcome"] = False
    st.session_state["form_prefilled"] = True
    st.session_state["scroll_to_numbers"] = False
    st.session_state["auto_run_demo"] = True
    st.session_state["demo_mode_active"] = True
    st.session_state["pending_demo_analysis_run"] = True
    st.session_state["scroll_to_demo_analyze"] = True
    st.session_state["scroll_target"] = "duka-ai-analyze-anchor"
    return True


def render_welcome_screen(_sample_cases: list[dict]) -> None:
    st.markdown(
        """
        <div class="welcome-hero">
            <div class="welcome-brand">Duka <span style="color:#1D9E75;">AI</span></div>
            <div class="welcome-tagline">Smart finance for every small business</div>
            <div class="welcome-desc">
                Understand your cash flow, loan readiness, and growth opportunities — in seconds.
                Built for African small businesses.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div style="height:1.25rem;"></div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3, gap="medium")

    with col1:
        st.markdown(
            """
            <div class="welcome-option-card">
                <div class="woc-badge">⭐ Recommended</div>
                <div class="woc-icon">⚡</div>
                <div class="woc-title">Try Demo Analysis</div>
                <div class="woc-desc">
                    One click loads a real sample <strong>income statement</strong>, pre-fills the
                    business profile, and runs the full multi-agent analysis automatically — perfect
                    for first-time visitors.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Try Demo Analysis →", key="welcome_sample", use_container_width=True, type="primary"):
            with st.status("⚡ Starting your demo…", expanded=True) as _demo_status:
                st.write("📂 Loading bundled sample income statement…")
                ok = _load_demo_analysis()
                if ok:
                    st.write("📊 Extracting revenue, expenses, and profit from the sample…")
                    st.write("🏪 Applying **Lusaka Grocery Shop** business profile…")
                    _demo_status.update(label="✅ Demo ready — opening workspace…", state="complete")
                else:
                    st.write("⚠️ Sample CSV not found — falling back to manual figures.")
                    _demo_status.update(label="Using manual demo figures…", state="complete")
            if ok:
                st.rerun()
            else:
                st.session_state.business_type = "Grocery shop"
                st.session_state.location = "Lusaka"
                st.session_state.products_services = "Groceries, mealie meal, drinks"
                st.session_state.main_question = "Should I restock or save cash?"
                st.session_state.manual_notes = (
                    "I run a small grocery shop in Lusaka. This week I made K4,500 in sales. "
                    "I spent K2,700 on stock, K500 on rent, K250 on transport, and K300 on other expenses. "
                    "I owe K1,000 to my supplier."
                )
                st.session_state.manual_revenue = 4500.0
                st.session_state.manual_expenses = 3750.0
                st.session_state.manual_debt = 1000.0
                st.session_state.numbers_entry_mode = "manual"
                st.session_state.hide_example_prompts = True
                st.session_state.show_welcome = False
                st.session_state.form_prefilled = True
                st.session_state.scroll_to_numbers = True
                st.session_state.auto_run_demo = True
                st.session_state.pending_demo_analysis_run = True
                st.session_state.scroll_to_demo_analyze = True
                st.session_state.scroll_target = "duka-ai-analyze-anchor"
                st.session_state.demo_mode_active = True
                st.rerun()

    with col2:
        st.markdown(
            """
            <div class="welcome-option-card">
                <div class="woc-icon">📄</div>
                <div class="woc-title">Upload Documents</div>
                <div class="woc-desc">
                    Upload a CSV, Excel, or PDF — income statement, bank statement, or
                    sales record for AI extraction.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Upload Documents →", key="welcome_upload", use_container_width=True):
            # Navigate to form, scroll to upload section — user uploads then clicks Analyze
            st.session_state.show_welcome = False
            st.session_state.numbers_entry_mode = "upload"
            st.session_state.scroll_to_numbers = True
            st.rerun()

    with col3:
        st.markdown(
            """
            <div class="welcome-option-card">
                <div class="woc-icon">✏️</div>
                <div class="woc-title">Start from Scratch</div>
                <div class="woc-desc">
                    Describe your business and enter figures manually — no documents needed,
                    just what you know.
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        if st.button("Start from Scratch →", key="welcome_scratch", use_container_width=True):
            # Navigate to the empty form — user fills in their details then clicks Analyze
            st.session_state.show_welcome = False
            st.session_state.numbers_entry_mode = "manual"
            st.session_state.form_prefilled = False
            st.rerun()

def render_sidebar_navigation(settings: Any) -> str:
    with st.sidebar:
        st.markdown(
            """
            <div class="sidebar-brand">
                <div class="sidebar-brand-row">
                    <div class="sidebar-brand-mark" aria-hidden="true">🌍</div>
                    <div class="sidebar-brand-copy">
                        <div class="sidebar-brand-title">Duka AI</div>
                        <div class="sidebar-brand-subtitle">SME finance workspace</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        for section, items in NAV_SECTIONS.items():
            st.markdown(f'<span class="sidebar-section-label">{section.title()}</span>', unsafe_allow_html=True)
            for label in items:
                page = NAV_LABEL_TO_PAGE[label]
                is_active = page == st.session_state.active_page
                if st.button(
                    label,
                    key=f"nav_{page.lower().replace(' ', '_')}",
                    use_container_width=True,
                    type="primary" if is_active else "secondary",
                ):
                    st.session_state.active_page = page
                    st.rerun()
        st.markdown(
            '<div class="sidebar-footer">Start in Chat Advisor, then use tools and reports from this sidebar.</div>',
            unsafe_allow_html=True,
        )
    return st.session_state.active_page


def require_report() -> dict[str, Any] | None:
    report = st.session_state.get("analysis_report")
    if report:
        return report
    st.markdown(
        """
        <div class="empty-state">
            <div class="empty-state-kicker">Analysis required</div>
            <div class="empty-state-title">Run Chat Advisor first</div>
            <div class="empty-state-copy">This tool needs your latest revenue, expenses, profit, and business context before it can calculate anything useful.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button("Open Chat Advisor", type="primary"):
        st.session_state.active_page = "Chat Advisor"
        st.rerun()
    return None


def render_page_header(title: str, subtitle: str, badge: str | None = None) -> None:
    badge_html = f'<div class="page-badge">{html.escape(badge)}</div>' if badge else ""
    st.markdown(
        f"""
        <div class="page-header">
            <div>
                <div class="page-title">{html.escape(title)}</div>
                <div class="page-subtitle">{html.escape(subtitle)}</div>
            </div>
            {badge_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_tool_tip(title: str, copy: str) -> None:
    st.markdown(
        f"""
        <div class="tool-tip">
            <strong>{html.escape(title)}</strong>
            <span>{html.escape(copy)}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_dashboard_page() -> None:
    render_page_header(
        "Dashboard",
        "Full business snapshot: financial health, cash flow, loan readiness, market context, and next actions.",
        "Full report",
    )
    report = require_report()
    if not report:
        return
    render_summary_panel(report)
    render_data_sources_card(report)
    left, right = st.columns([1.15, 0.85], gap="large")
    with left:
        render_cashflow_section(report)
        render_financial_advice_section(report)
    with right:
        render_business_health_card(report)
        if report.get("transaction_summary"):
            render_transaction_card(report["transaction_summary"])
        st.markdown('<div class="mini-card"><div class="card-title">All Recommended Actions</div></div>', unsafe_allow_html=True)
        for i, action in enumerate(report["final_recommended_actions"], start=1):
            render_action_card(i, action)
    lower_left, lower_right = st.columns(2, gap="large")
    with lower_left:
        render_market_intelligence_section(report)
    with lower_right:
        render_loan_readiness_card(report)


def _generate_forecast_ai_summary(baseline: dict, forecast: dict, scenario_key: str) -> str:
    """Call LLM to narrate the forecast. Falls back to template if LLM unavailable."""
    from agents import request_llm
    s = forecast["summary"]
    sc = forecast["scenario"]
    system = (
        "You are Duka AI's Cash Flow Analyst. In 2-3 sentences summarise what this "
        "6-month forecast means for the business owner. Be direct and practical. Use K (Kwacha). "
        "No bullet points."
    )
    user = (
        f"Scenario: {sc['label']} — {sc['description']}\n"
        f"Starting: revenue K{baseline['monthly_revenue']:,.0f}, "
        f"expenses K{baseline['monthly_expenses']:,.0f}, profit K{baseline['monthly_profit']:,.0f}\n"
        f"Month 6: revenue K{s['final_revenue']:,.0f}, expenses K{s['final_expenses']:,.0f}, "
        f"profit K{s['final_profit']:,.0f}\n"
        f"Profitable months: {s['profitable_months']}/{s['total_months']}\n"
        + (f"First difficult month: {s['first_loss_month']}\n" if s["first_loss_month"] else "")
        + "Summarise in 2-3 plain sentences what this means for the owner."
    )
    result = request_llm(system, user, max_tokens=120)
    if result:
        return result.strip()
    if s["all_profitable"]:
        return (
            f"Under the {sc['label']} scenario all 6 months stay profitable, "
            f"ending at K{s['final_profit']:,.0f} profit and a {s['final_margin']:.1f}% margin. "
            f"Revenue grows from K{baseline['monthly_revenue']:,.0f} to K{s['final_revenue']:,.0f}."
        )
    return (
        f"The {sc['label']} scenario shows {s['profitable_months']} out of {s['total_months']} months profitable. "
        f"The first difficult month is {s['first_loss_month']}. "
        "Consider reducing the expense growth rate or boosting revenue to stay positive."
    )


_CHART_TRIGGERS_FORECAST = ("graph", "chart", "plot", "visualize", "visualise", "show me a chart", "show graph", "line chart")
# Quick prompts / natural phrases that imply a chart without saying "chart"
_FORECAST_CHART_PHRASES = (
    "riskiest",
    "expense breakdown",
    "profit trend",
    "revenue vs",
    "show profit",
    "trendline",
    "trend line",
)
_CHART_AFFIRM_FORECAST = ("yes", "yeah", "yep", "sure", "ok", "okay", "please", "go ahead", "show me", "show it")


def _is_forecast_chart_request(q: str, prev_assistant_msg: str) -> bool:
    """True if the user is asking for / agreeing to a chart on the forecast page."""
    q = (q or "").lower().strip()
    if not q:
        return False
    if any(kw in q for kw in _CHART_TRIGGERS_FORECAST):
        return True
    if any(p in q for p in _FORECAST_CHART_PHRASES):
        return True
    # Short affirmation right after the agent offered a chart
    if len(q.split()) <= 6 and any(
        q == w or q.startswith(w + " ") or q.startswith(w + ",") or q.startswith(w + ".")
        for w in _CHART_AFFIRM_FORECAST
    ):
        if "chart" in (prev_assistant_msg or "").lower() or "graph" in (prev_assistant_msg or "").lower() or "plot" in (prev_assistant_msg or "").lower():
            return True
    return False


def _margin_band_colors(margins: list[float]) -> list[str]:
    """SME-tuned bands: >20% healthy (green), 10–20% watch (amber),
    0–10% thin (red), <0% loss (deep red). Aligns with health/loan scoring."""
    out: list[str] = []
    for v in margins:
        if v <= 0:
            out.append("#B91C1C")
        elif v < 10:
            out.append("#EF4444")
        elif v < 20:
            out.append("#F59E0B")
        else:
            out.append("#10B981")
    return out


def _build_forecast_page_chart(
    user_q: str,
    baseline: dict,
    scenario_key: str,
    forecast: dict,
    chat_history: list,
) -> dict[str, Any] | None:
    """Build a Plotly-ready chart dict (same shape as AI <CHART> JSON) when keywords match.

    Honours the horizon when intent detects months (e.g. 'in 10 months').
    """
    from tools.financial_engine import detect_forecast_intent

    prev_assistant = ""
    for m in reversed(chat_history[:-1] if chat_history else []):
        if m.get("role") == "assistant":
            prev_assistant = m.get("content", "")
            break

    if not _is_forecast_chart_request(user_q, prev_assistant):
        return None

    intent = detect_forecast_intent(user_q)
    if not intent:
        for m in reversed(chat_history[:-1] if chat_history else []):
            if m.get("role") == "user":
                intent = detect_forecast_intent(m.get("content", ""))
                if intent:
                    break

    months_n = intent["months"] if intent else len(forecast.get("months") or []) or 6
    fc = calculate_forecast(baseline, scenario_key, months=months_n)
    rows = fc["months"]
    labels = [r["month_name"] for r in rows]
    ql = (user_q or "").lower()
    scen_label = fc["scenario"]["label"]

    if "riskiest" in ql:
        margins = [round(float(r["margin"]), 1) for r in rows]
        return {
            "type": "bar",
            "title": f"Riskiest months — profit margin % ({scen_label})",
            "data": {
                "labels": labels,
                "datasets": [
                    {
                        "name": "Profit Margin (%)",
                        "values": margins,
                        "colors": _margin_band_colors(margins),
                    }
                ],
            },
        }

    if "expense" in ql and "breakdown" in ql:
        return {
            "type": "bar",
            "title": f"Monthly expenses ({scen_label})",
            "data": {
                "labels": labels,
                "datasets": [
                    {
                        "name": "Monthly Expenses",
                        "values": [int(r["expenses"]) for r in rows],
                        "color": "#E24B4A",
                    }
                ],
            },
        }

    if "profit" in ql and "trend" in ql:
        return {
            "type": "line",
            "title": f"Profit trend ({scen_label})",
            "data": {
                "labels": labels,
                "datasets": [
                    {
                        "name": "Profit",
                        "values": [int(r["profit"]) for r in rows],
                        "color": "#378ADD",
                    }
                ],
            },
        }

    if ("revenue" in ql and "expense" in ql) or "revenue vs" in ql:
        return {
            "type": "line",
            "title": f"Revenue vs expenses ({scen_label})",
            "data": {
                "labels": labels,
                "datasets": [
                    {"name": "Revenue", "values": [int(r["revenue"]) for r in rows], "color": "#1D9E75"},
                    {"name": "Expenses", "values": [int(r["expenses"]) for r in rows], "color": "#E24B4A"},
                ],
            },
        }

    return {
        "type": "line",
        "title": f"Forecast — {scen_label} ({months_n} months)",
        "data": {
            "labels": labels,
            "datasets": [
                {"name": "Revenue", "values": [int(r["revenue"]) for r in rows], "color": "#10B981"},
                {"name": "Expenses", "values": [int(r["expenses"]) for r in rows], "color": "#EF4444"},
                {"name": "Profit", "values": [int(r["profit"]) for r in rows], "color": "#2563EB"},
            ],
        },
    }


_FORECAST_GREETING_PATTERNS = (
    r"^\s*(hi|hey|hello|hola|yo|sup|howdy|good\s*(morning|afternoon|evening))\b",
    r"\bhow\s+are\s+you\b",
    r"\bwho\s+are\s+you\b",
    r"\bwhat\s+are\s+you\b",
    r"\bwhat\s+can\s+you\s+do\b",
    r"\bhelp\b\s*$",
)


def _classify_forecast_user_input(text: str, months: list[dict]) -> str:
    """Classify the user's input before we ever call the LLM.

    Returns one of: 'gibberish', 'greeting', 'identity', 'thanks', 'normal'.
    """
    raw = (text or "").strip()
    if not raw:
        return "gibberish"

    low = raw.lower()
    if any(re.search(p, low) for p in _FORECAST_GREETING_PATTERNS):
        if any(w in low for w in ("who", "what")):
            return "identity"
        return "greeting"
    if low in ("thanks", "thank you", "ty", "ta", "cheers"):
        return "thanks"

    if _is_low_vowel_noise(raw, 4, 0.20):
        return "gibberish"

    letters = "".join(c for c in raw if c.isalpha())
    has_digit = any(c.isdigit() for c in raw)
    has_finance_term = any(
        w in low
        for w in (
            "revenue", "expense", "expenses", "profit", "margin", "month", "chart",
            "graph", "plot", "trend", "forecast", "scenario", "loss", "earn", "make",
            "growth", "risk", "kwacha", "k", "show", "compare", "vs", "versus", "next",
            "cash", "money", "income",
        )
    )
    has_month = any(
        m["month_name"].split()[0].lower() in low for m in months
    ) or any(
        x in low for x in (
            "jan", "feb", "mar", "apr", "may", "jun",
            "jul", "aug", "sep", "oct", "nov", "dec",
        )
    )
    if (
        len(letters) <= 8
        and not has_digit
        and not has_finance_term
        and not has_month
        and len(low.split()) <= 2
    ):
        return "gibberish"
    return "normal"


def _forecast_intro_reply(months: list[dict], sc: dict) -> str:
    first, last = months[0], months[-1]
    return (
        "I’m **Duka AI’s Cash Flow Analyst**. I read your **6‑month forecast** and answer "
        "questions in plain language using the numbers from this scenario.\n\n"
        f"Right now you’re looking at the **{sc['label']}** scenario "
        f"({first['month_name']} → {last['month_name']}). Try a quick prompt above, or ask:\n\n"
        "• *“What’s my profit in 3 months?”*  \n"
        "• *“Show profit trend.”*  \n"
        "• *“Why are expenses rising?”*"
    )


def _forecast_greeting_reply(months: list[dict], sc: dict) -> str:
    return (
        "Hi! I’m **Duka AI’s Cash Flow Analyst** — here to help with your forecast. "
        f"You’re on the **{sc['label']}** scenario ({months[0]['month_name']}–{months[-1]['month_name']}). "
        "Ask me about a month, profit, expenses, or say *“revenue vs expenses”* for a chart."
    )


def _forecast_gibberish_reply(months: list[dict]) -> str:
    return (
        "I didn’t catch that. Try a short question about **this forecast** "
        f"({months[0]['month_name']}–{months[-1]['month_name']}) — for example: "
        "*“profit in 3 months”*, *“expense breakdown”*, or *“riskiest months”*."
    )


def _handle_forecast_question(
    question: str,
    baseline: dict,
    forecast: dict,
    scenario_key: str,
    chat_history: list,
) -> str:
    """Answer a forecast question using ONLY computed forecast numbers.

    If the user asks about a horizon longer than the page's default forecast
    (e.g. "in 10 months" when the page shows 6), recompute a fresh forecast
    that covers that horizon so the answer is grounded in real numbers.
    """
    from agents import request_llm_chat
    from tools.financial_engine import detect_forecast_intent

    s = forecast["summary"]
    sc = forecast["scenario"]
    months_now = forecast["months"]

    # ── Pre-LLM classification: short-circuit for noise / greetings / identity
    kind = _classify_forecast_user_input(question, months_now)
    if kind == "gibberish":
        return _forecast_gibberish_reply(months_now)
    if kind == "identity":
        return _forecast_intro_reply(months_now, sc)
    if kind == "greeting":
        return _forecast_greeting_reply(months_now, sc)
    if kind == "thanks":
        return "You’re welcome! Ask me anything else about this forecast."

    intent = detect_forecast_intent(question)
    requested_months = intent["months"] if intent else 0
    page_months = len(forecast["months"])

    # If the user wants a horizon beyond what the page rendered, recompute.
    if requested_months > page_months:
        forecast = calculate_forecast(baseline, scenario_key, months=requested_months)

    s = forecast["summary"]
    sc = forecast["scenario"]
    months = forecast["months"]

    target_block = ""
    if requested_months and 1 <= requested_months <= len(months):
        target = months[requested_months - 1]
        target_block = (
            f"\nDIRECT ANSWER FOR MONTH {requested_months} ({target['month_name']}):\n"
            f"  Revenue:  K{target['revenue']:,.0f}\n"
            f"  Expenses: K{target['expenses']:,.0f}\n"
            f"  Profit:   K{target['profit']:,.0f}  (margin {target['margin']:.1f}%)\n"
            "Use these exact figures when answering. Do NOT say data is unavailable.\n"
        )

    cumulative_block = ""
    if requested_months and requested_months > 1 and len(months) >= requested_months:
        segment = months[:requested_months]
        tot_r = sum(m["revenue"] for m in segment)
        tot_e = sum(m["expenses"] for m in segment)
        tot_p = sum(m["profit"] for m in segment)
        cumulative_block = (
            f"\nCUMULATIVE FOR FIRST {requested_months} MONTHS (sum of month-by-month rows):\n"
            f"  Total revenue:  K{tot_r:,.0f}\n"
            f"  Total expenses: K{tot_e:,.0f}\n"
            f"  Total profit:   K{tot_p:,.0f}\n"
            "If the user asks how much they make/earn or total profit over that many months, "
            "lead with these cumulative totals (not only the last month).\n"
        )

    context = (
        "FORECAST CONTEXT (use ONLY these numbers - never invent figures):\n"
        f"Scenario: {sc['label']} - {sc['description']}\n"
        f"Starting: revenue K{baseline['monthly_revenue']:,.0f}, "
        f"expenses K{baseline['monthly_expenses']:,.0f}, profit K{baseline['monthly_profit']:,.0f}\n"
        f"Revenue growth: {sc['rev_growth']*100:+.1f}% per month | "
        f"Expense growth: {sc['exp_growth']*100:+.1f}% per month\n"
        f"Profitable months: {s['profitable_months']}/{s['total_months']}\n"
        + (f"First loss month: {s['first_loss_month']}\n" if s["first_loss_month"] else "")
        + "Month-by-month:\n"
        + "\n".join(
            f"  Month {m['month_num']} ({m['month_name']}): rev K{m['revenue']:,.0f}, "
            f"exp K{m['expenses']:,.0f}, profit K{m['profit']:,.0f} ({m['margin']:.1f}%)"
            for m in months
        )
        + target_block
        + cumulative_block
    )

    # Build chart-ready arrays the AI can copy verbatim into JSON
    chart_labels = [m["month_name"] for m in months]
    chart_rev    = [int(m["revenue"])  for m in months]
    chart_exp    = [int(m["expenses"]) for m in months]
    chart_profit = [int(m["profit"])   for m in months]
    chart_margins = [round(m["margin"], 1) for m in months]
    margin_colors = [
        "#10B981" if v > 25 else ("#F59E0B" if v >= 15 else "#EF4444")
        for v in chart_margins
    ]

    chart_data_block = (
        "\n\nCHART-READY DATA — copy these EXACT values into <CHART> JSON, do NOT alter them:\n"
        f"Labels:         {json.dumps(chart_labels)}\n"
        f"Revenue:        {json.dumps(chart_rev)}\n"
        f"Expenses:       {json.dumps(chart_exp)}\n"
        f"Profit:         {json.dumps(chart_profit)}\n"
        f"Margins (%):    {json.dumps(chart_margins)}\n"
        f"Margin colors:  {json.dumps(margin_colors)}\n"
    )

    chart_instructions = (
        "\n\nCHART INSTRUCTIONS:\n"
        "When the user asks for a chart, graph, trend, visualization, or breakdown, "
        "FIRST write a 1–2 sentence plain-language explanation of what the chart will show "
        "(use the actual numbers, e.g. 'Revenue grows from K50,553 to K61,505 while expenses rise from K75,874 to K77,790'). "
        "THEN end your response with a <CHART> block using this exact format. "
        "NEVER reply with only the JSON — there must always be readable prose first.\n"
        "<CHART>\n"
        "{\n"
        '  "type": "line",\n'
        '  "title": "...",\n'
        '  "data": {\n'
        '    "labels": [...],\n'
        '    "datasets": [{"name": "...", "values": [...], "color": "..."}]\n'
        '  }\n'
        "}\n"
        "</CHART>\n\n"
        "If you show BOTH monthly profit in Kwacha AND margin % on one line chart, use two datasets: "
        "one named e.g. 'Profit' and one named e.g. 'Profit margin (%)' or 'Margin (%)' so % stays on its own scale.\n"
        "What to produce per trigger (use CHART-READY DATA above):\n"
        "• 'show profit trend'            → type=line,  1 dataset:  Profit #378ADD\n"
        "• 'revenue vs expenses'          → type=line,  2 datasets: Revenue #1D9E75, Expenses #E24B4A\n"
        "• 'riskiest months' / 'risk'     → type=bar,   1 dataset:  name='Profit Margin %', values=Margins, "
        "use 'colors' (array) = Margin colors — NOT a single 'color'\n"
        "• 'expense breakdown'            → type=bar,   1 dataset:  name='Monthly Expenses', values=Expenses, color=#E24B4A\n"
        "• 'show me a chart' / 'all'      → type=line,  3 datasets: Revenue #1D9E75, Expenses #E24B4A, Profit #378ADD\n\n"
        "IMPORTANT: Use ONLY the CHART-READY DATA values — never invent numbers.\n"
        "IMPORTANT: For riskiest months use per-bar colors as 'colors': [...] or 'color': [...] (array).\n"
        "Do NOT include a <CHART> block if the user is not asking for a chart or visualization."
    )

    system = (
        "You are **Duka AI's Cash Flow Analyst**, a financial assistant inside the Duka AI app. "
        "If the user asks who/what you are, say so plainly — never call yourself a generic computer program. "
        "Use ONLY the numbers in FORECAST CONTEXT. Every Kwacha figure you mention must match "
        "a value listed there (or a cumulative sum of those rows). Never invent month-specific amounts. "
        "SCOPE: Only answer using this scenario's month-by-month rows. If the question is unrelated "
        "(general trivia, other businesses, topics with no numbers here), reply in ONE short sentence that you "
        "only have this forecast and suggest one example question — do not fabricate data, do not give "
        "long generic business advice. "
        "Do not output random characters, keyboard mash, or filler before your answer. "
        "If a CUMULATIVE FOR FIRST N MONTHS block is present and the user asked about that span, "
        "lead with those totals. If only DIRECT ANSWER FOR MONTH X applies, use that month. "
        "When the forecast rows fully answer the horizon, do not refuse — use those numbers. "
        "Be concise (2-4 sentences). For non-chart questions you may offer: 'Want me to plot this on a chart?' "
        "OUTPUT STYLE: Plain sentences and Markdown only. Do NOT use LaTeX (no \\times, \\text{...}, \\( \\), "
        "\\[ \\]). For maths use readable lines like \"K50,553 × 1.04 = K52,575\"."
        f"\n\n{context}"
        + chart_data_block
        + chart_instructions
    )
    messages = [
        {"role": "system", "content": system},
        *[{"role": m["role"], "content": m["content"]} for m in chat_history[-6:]],
        {"role": "user", "content": question},
    ]
    result = request_llm_chat(messages, temperature=0.15, max_tokens=600)
    if not result:
        return "I couldn't generate a response right now. Please try again."

    chart_tag = ""
    m_chart = re.search(r"<CHART>.*?</CHART>", result, flags=re.DOTALL | re.IGNORECASE)
    if m_chart:
        chart_tag = m_chart.group(0).strip()
    prose = re.sub(r"<CHART>.*?</CHART>", "", result, flags=re.DOTALL | re.IGNORECASE).strip()
    _bare, prose_no_json = _try_extract_bare_chart_json(prose)
    prose_check = prose_no_json if prose_no_json else prose

    if prose_check.strip():
        if _is_low_vowel_noise(prose_check, 15, 0.13):
            fixed = _safe_deterministic_forecast_answer(question, months, sc, baseline)
            return f"{fixed}\n\n{chart_tag}" if chart_tag else fixed
        if not _forecast_prose_is_grounded(prose_check, months, baseline):
            prose_fixed = _repair_or_replace_forecast_prose(
                prose_check, question, months, sc, baseline
            )
            if chart_tag:
                return f"{prose_fixed}\n\n{chart_tag}"
            return prose_fixed

    return result


def _forecast_allowed_kwacha_ints(months: list[dict], baseline: dict) -> set[int]:
    """Every integer Kwacha amount that may appear in a grounded forecast reply."""
    s: set[int] = set()
    for m in months:
        for key in ("revenue", "expenses", "profit"):
            s.add(int(round(float(m[key]))))
    for key in ("monthly_revenue", "monthly_expenses", "monthly_profit"):
        v = baseline.get(key)
        if v is not None:
            s.add(int(round(float(v))))
    n = len(months)
    for i in range(n):
        seg = months[: i + 1]
        s.add(int(round(sum(float(x["revenue"]) for x in seg))))
        s.add(int(round(sum(float(x["expenses"]) for x in seg))))
        s.add(int(round(sum(float(x["profit"]) for x in seg))))
    return s


def _extract_kwacha_ints_from_prose(text: str) -> list[int]:
    out: list[int] = []
    for m in re.finditer(r"K\s*([\d,]+)", text or ""):
        try:
            out.append(int(m.group(1).replace(",", "")))
        except ValueError:
            continue
    return out


def _forecast_prose_is_grounded(prose: str, months: list[dict], baseline: dict) -> bool:
    """True if every K-amount in prose appears in the allowed forecast math."""
    if not (prose or "").strip():
        return True
    allowed = _forecast_allowed_kwacha_ints(months, baseline)
    amounts = _extract_kwacha_ints_from_prose(prose)
    return all(a in allowed for a in amounts)


def _is_low_vowel_noise(text: str, min_letters: int = 12, vowel_ratio_max: float = 0.14) -> bool:
    """Very low vowel density usually means noise / accidental keyboard mash."""
    t = "".join(c for c in (text or "") if c.isalpha())
    if len(t) < min_letters:
        return False
    vowels = sum(1 for c in t if c.lower() in "aeiou")
    return (vowels / len(t)) < vowel_ratio_max


def _is_gibberish_user_question(text: str) -> bool:
    return _is_low_vowel_noise(text or "", 12, 0.14)


def _safe_deterministic_forecast_answer(
    question: str,
    months: list[dict],
    sc: dict[str, Any],
    baseline: dict,
) -> str:
    """Facts-only reply when the model hallucinated amounts."""
    ql = (question or "").lower()
    first, last = months[0], months[-1]
    hi_exp_m = max(months, key=lambda m: float(m["expenses"]))
    hi_rev_m = max(months, key=lambda m: float(m["revenue"]))
    if any(
        w in ql
        for w in (
            "revenue",
            "expense",
            "vs",
            "versus",
            "compare",
            "highest",
            "most",
            "biggest",
            "largest",
        )
    ):
        return (
            f"In the **{sc['label']}** scenario: revenue moves from **K{first['revenue']:,.0f}** "
            f"({first['month_name']}) to **K{last['revenue']:,.0f}** ({last['month_name']}); "
            f"expenses from **K{first['expenses']:,.0f}** to **K{last['expenses']:,.0f}**. "
            f"Highest expense month in this forecast: **{hi_exp_m['month_name']}** at **K{hi_exp_m['expenses']:,.0f}**. "
            f"Peak revenue month: **{hi_rev_m['month_name']}** (**K{hi_rev_m['revenue']:,.0f}**)."
        )
    return (
        f"**{sc['label']}** forecast ({first['month_name']} → {last['month_name']}): "
        f"revenue **K{first['revenue']:,.0f}**→**K{last['revenue']:,.0f}**, "
        f"expenses **K{first['expenses']:,.0f}**→**K{last['expenses']:,.0f}**, "
        f"profit **K{first['profit']:,.0f}**→**K{last['profit']:,.0f}**. "
        "Ask about a specific month or say “revenue vs expenses” for a chart."
    )


def _repair_or_replace_forecast_prose(
    prose: str,
    question: str,
    months: list[dict],
    sc: dict[str, Any],
    baseline: dict,
) -> str:
    """Drop leading junk paragraphs, or replace with deterministic facts."""
    if _forecast_prose_is_grounded(prose, months, baseline):
        return prose.strip()
    chunks = [c.strip() for c in re.split(r"\n\s*\n+", prose) if c.strip()]
    for i in range(len(chunks)):
        tail = "\n\n".join(chunks[i:])
        if _forecast_prose_is_grounded(tail, months, baseline):
            return tail
    return _safe_deterministic_forecast_answer(question, months, sc, baseline)


def sanitize_forecast_answer_text(raw: str) -> str:
    """Turn LaTeX-style fragments into readable text for Markdown (Streamlit-safe)."""
    if not raw:
        return raw
    s = raw.replace("\\times", "×").replace("\\cdot", "×")
    for _ in range(8):
        n = re.sub(r"\\text\{([^{}]*)\}", r"\1", s)
        if n == s:
            break
        s = n
    s = re.sub(r"\\\(|\\\)|\\\[|\\\]", "", s)
    return re.sub(r"\n{3,}", "\n\n", s).strip()


def _looks_like_forecast_chart_json(d: Any) -> bool:
    if not isinstance(d, dict):
        return False
    if d.get("type") not in ("line", "bar"):
        return False
    data = d.get("data")
    if not isinstance(data, dict):
        return False
    labels, sets = data.get("labels"), data.get("datasets")
    return isinstance(labels, list) and isinstance(sets, list) and len(labels) > 0 and len(sets) > 0


def _extract_first_brace_json_object(s: str, start: int = 0) -> tuple[str | None, int]:
    """Return (json_blob, index_after_blob) for first balanced `{...}` from start, string-aware."""
    i = s.find("{", start)
    if i < 0:
        return None, start
    depth = 0
    in_str = False
    esc = False
    for j in range(i, len(s)):
        c = s[j]
        if in_str:
            if esc:
                esc = False
            elif c == "\\":
                esc = True
            elif c == '"':
                in_str = False
            continue
        if c == '"':
            in_str = True
        elif c == "{":
            depth += 1
        elif c == "}":
            depth -= 1
            if depth == 0:
                return s[i : j + 1], j + 1
    return None, start


def _try_extract_bare_chart_json(text: str) -> tuple[dict | None, str]:
    """When the model emits chart JSON without <CHART> tags, strip it and return chart_data."""
    raw = text.strip()
    if not raw:
        return None, text

    # Markdown fenced block
    fence = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", raw, re.DOTALL)
    if fence:
        try:
            d = json.loads(fence.group(1))
            if _looks_like_forecast_chart_json(d):
                clean = raw.replace(fence.group(0), "").strip()
                return d, re.sub(r"\n{3,}", "\n\n", clean)
        except Exception:
            pass

    # Whole message is just JSON
    if raw.startswith("{") and raw.endswith("}"):
        try:
            d = json.loads(raw)
            if _looks_like_forecast_chart_json(d):
                return d, ""
        except Exception:
            pass

    # Scan for embedded chart object (common when prose + JSON)
    search_from = 0
    while True:
        blob, end = _extract_first_brace_json_object(raw, search_from)
        if not blob:
            break
        try:
            d = json.loads(blob)
            if _looks_like_forecast_chart_json(d):
                pos = raw.find(blob)
                clean = (raw[:pos] + raw[pos + len(blob) :]).strip()
                return d, re.sub(r"\n{3,}", "\n\n", clean)
        except Exception:
            pass
        search_from = raw.find("{", search_from + 1)
        if search_from < 0:
            break

    return None, text


def parse_and_render_chart(response_text: str) -> tuple[str, dict | None]:
    """Extract a <CHART>…</CHART> block from an AI response.

    Returns (clean_text, chart_data_dict). chart_data_dict is None when no
    valid block is found.
    """
    chart_pattern = r"<CHART>(.*?)</CHART>"
    match = re.search(chart_pattern, response_text, re.DOTALL | re.IGNORECASE)
    if match:
        clean_text = re.sub(chart_pattern, "", response_text, flags=re.DOTALL | re.IGNORECASE).strip()
        try:
            chart_data = json.loads(match.group(1).strip())
            return clean_text, chart_data
        except Exception:
            return clean_text, None

    bare, stripped = _try_extract_bare_chart_json(response_text)
    return stripped, bare


def _format_kwacha(v: float) -> str:
    return f"K{v:,.0f}"


def _format_pct(v: float) -> str:
    return f"{v:+.1f}%" if v else "0.0%"


def _caption_for_forecast_chart(chart_data: dict, scenario_label: str | None = None) -> str:
    """Generate a multi-sentence breakdown from a chart JSON.

    Used both when the model returns no prose and when its prose is too thin
    (so the user always gets a numbered breakdown alongside the chart).
    """
    try:
        title = (chart_data.get("title") or "").strip()
        data = chart_data.get("data") or {}
        labels = data.get("labels") or []
        datasets = data.get("datasets") or []
        if not labels or not datasets:
            return ""

        first, last = labels[0], labels[-1]
        scen_suffix = f" under the **{scenario_label}** scenario" if scenario_label else ""

        def _delta_pct(v0: float, vN: float) -> str:
            if v0 == 0:
                return ""
            pct = (vN - v0) / abs(v0) * 100
            return f" ({pct:+.1f}%)"

        # Single-series shortcuts ─────────────────────────────────────────────
        if len(datasets) == 1:
            d = datasets[0]
            name = (d.get("name") or "").strip() or title or "Series"
            vals = d.get("values") or []
            if not vals:
                return ""
            is_pct = "%" in name or "margin" in name.lower()
            v0, vN = vals[0], vals[-1]
            if is_pct:
                trend = "improves" if vN > v0 else ("declines" if vN < v0 else "stays flat")
                avg = sum(vals) / len(vals)
                return (
                    f"**{name}** {trend} from **{v0:.1f}%** in {first} to **{vN:.1f}%** in "
                    f"{last}{scen_suffix} — average of **{avg:.1f}%** across the {len(vals)} months.\n\n"
                    f"- **Highest:** {max(vals):.1f}%   ·   **Lowest:** {min(vals):.1f}%\n"
                    f"- **Net change:** {_format_pct(vN - v0)}\n"
                )
            trend = "rises" if vN > v0 else ("falls" if vN < v0 else "stays flat")
            total = sum(vals)
            return (
                f"**{name}** {trend} from **{_format_kwacha(v0)}** in {first} to "
                f"**{_format_kwacha(vN)}** in {last}{scen_suffix}{_delta_pct(v0, vN)}.\n\n"
                f"- **Total across the period:** {_format_kwacha(total)}\n"
                f"- **Average per month:** {_format_kwacha(total / len(vals))}\n"
            )

        # Two-series — typical revenue vs expenses, profit vs margin ─────────
        if len(datasets) == 2:
            a, b = datasets[0], datasets[1]
            name_a = (a.get("name") or "Series A").strip()
            name_b = (b.get("name") or "Series B").strip()
            va, vb = a.get("values") or [], b.get("values") or []
            if not va or not vb:
                return ""
            a_pct = "%" in name_a or "margin" in name_a.lower()
            b_pct = "%" in name_b or "margin" in name_b.lower()
            fmt_a = (lambda v: f"{v:.1f}%") if a_pct else _format_kwacha
            fmt_b = (lambda v: f"{v:.1f}%") if b_pct else _format_kwacha

            lines = [
                f"Here's how **{name_a}** and **{name_b}** move from {first} to "
                f"{last}{scen_suffix}:",
                "",
                f"- **{name_a}:** {fmt_a(va[0])} → {fmt_a(va[-1])}"
                + (f" {_delta_pct(va[0], va[-1])}" if not a_pct else f" ({(va[-1] - va[0]):+.1f} pts)"),
                f"- **{name_b}:** {fmt_b(vb[0])} → {fmt_b(vb[-1])}"
                + (f" {_delta_pct(vb[0], vb[-1])}" if not b_pct else f" ({(vb[-1] - vb[0]):+.1f} pts)"),
            ]
            # If looks like Revenue vs Expenses, add net (gap)
            if not a_pct and not b_pct and len(va) == len(vb):
                gap_first = va[0] - vb[0]
                gap_last = va[-1] - vb[-1]
                lines.append(
                    f"- **Gap ({name_a} − {name_b}):** "
                    f"{_format_kwacha(gap_first)} → {_format_kwacha(gap_last)}"
                )
            return "\n".join(lines)

        # Three-series fallback (e.g. Revenue + Expenses + Profit) ───────────
        try:
            names = [(d.get("name") or "?").strip() for d in datasets]
            firsts = [d.get("values", [None])[0] for d in datasets]
            lasts = [d.get("values", [None])[-1] for d in datasets]
            lines = [
                f"Here's the breakdown from {first} to {last}{scen_suffix}:",
                "",
            ]
            for n, v0, vN in zip(names, firsts, lasts):
                if v0 is None or vN is None:
                    continue
                is_pct = "%" in n or "margin" in n.lower()
                fmt = (lambda v: f"{v:.1f}%") if is_pct else _format_kwacha
                lines.append(f"- **{n}:** {fmt(v0)} → {fmt(vN)}")
            return "\n".join(lines)
        except Exception:
            return f"{title or 'Forecast view'}: {', '.join(names)} from {first} to {last}{scen_suffix}."
    except Exception:
        return ""


def _forecast_series_is_percentage_axis(name: str) -> bool:
    """Names that denote margin / % — must use a separate axis from Kwacha amounts."""
    if not name:
        return False
    if "%" in name:
        return True
    low = name.lower()
    return "margin" in low


def build_plotly_chart(chart_data: dict) -> go.Figure:
    """Build a dark-themed Plotly chart from AI-generated chart JSON data."""
    fig = go.Figure()
    chart_type = chart_data.get("type", "line")
    labels = chart_data["data"]["labels"]
    datasets = chart_data["data"]["datasets"]

    pct_flags = [_forecast_series_is_percentage_axis(d.get("name", "")) for d in datasets]
    dual_axis = bool(datasets) and any(pct_flags) and not all(pct_flags)

    for dataset in datasets:
        name = dataset["name"]
        values = dataset["values"]
        per_bar_colors = dataset.get("colors")
        raw_color = dataset.get("color", "#1D9E75")
        if per_bar_colors is None and isinstance(raw_color, list):
            per_bar_colors = raw_color
            single_color = "#1D9E75"
        else:
            single_color = raw_color if isinstance(raw_color, str) else "#1D9E75"

        is_pct_series = _forecast_series_is_percentage_axis(name)
        yaxis_ref = "y2" if dual_axis and is_pct_series else "y"

        if chart_type == "line":
            hover = (
                f"%{{x}}: %{{y:.1f}}%<extra>{name}</extra>"
                if is_pct_series
                else f"%{{x}}: K%{{y:,.0f}}<extra>{name}</extra>"
            )
            fig.add_trace(go.Scatter(
                x=labels,
                y=values,
                name=name,
                yaxis=yaxis_ref,
                line=dict(color=single_color, width=3),
                mode="lines+markers",
                marker=dict(size=9, color=single_color,
                            line=dict(width=1.5, color="rgba(255,255,255,0.35)")),
                hovertemplate=hover,
            ))
        elif chart_type == "bar":
            is_margin = is_pct_series
            hover = (
                f"%{{x}}: %{{y:.1f}}%<extra>{name}</extra>"
                if is_margin
                else f"%{{x}}: K%{{y:,.0f}}<extra>{name}</extra>"
            )
            n_bars = len(values)
            if per_bar_colors:
                if len(per_bar_colors) < n_bars:
                    last = per_bar_colors[-1] if per_bar_colors else single_color
                    bar_colors = list(per_bar_colors) + [last] * (n_bars - len(per_bar_colors))
                else:
                    bar_colors = list(per_bar_colors)[:n_bars]
            else:
                bar_colors = [single_color] * n_bars
            fig.add_trace(go.Bar(
                x=labels,
                y=values,
                name=name,
                yaxis=yaxis_ref,
                marker=dict(color=bar_colors, line=dict(width=0)),
                hovertemplate=hover,
                text=[f"{v:.1f}%" if is_margin else f"K{v:,.0f}" for v in values],
                textposition="outside",
                textfont=dict(size=11, color="#E2E8F0"),
            ))

    only_pct_single_axis = bool(datasets) and all(pct_flags)

    fig.update_layout(
        title=dict(
            text=chart_data.get("title", "Financial Chart"),
            font=dict(size=17, color="#F1F5F9", family="Segoe UI, Arial, sans-serif"),
            x=0.5,
            xanchor="center",
        ),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(13,31,60,0.88)",
        font=dict(color="#E2E8F0", size=14, family="Segoe UI, Arial, sans-serif"),
        legend=dict(
            bgcolor="rgba(15,23,42,0.78)",
            bordercolor="rgba(148,163,184,0.22)",
            borderwidth=1,
            font=dict(size=13),
            orientation="h",
            y=1.09,
            x=0.5,
            xanchor="center",
        ),
        xaxis=dict(
            gridcolor="rgba(148,163,184,0.1)",
            color="#E2E8F0",
            tickfont=dict(size=13),
            showline=True,
            linecolor="rgba(148,163,184,0.25)",
        ),
        margin=dict(l=56, r=56 if dual_axis else 28, t=76, b=44),
        height=int(chart_data.get("height", 455)),
        hovermode="x unified",
        bargap=0.28,
    )

    if dual_axis:
        fig.update_layout(
            yaxis=dict(
                title=dict(text="Amount (Kwacha)", font=dict(size=13)),
                gridcolor="rgba(148,163,184,0.12)",
                color="#93C5FD",
                tickprefix="K",
                ticksuffix="",
                zeroline=True,
                zerolinecolor="rgba(148,163,184,0.35)",
                zerolinewidth=1,
                tickfont=dict(size=13),
            ),
            yaxis2=dict(
                title=dict(text="Margin / %", font=dict(size=13)),
                overlaying="y",
                side="right",
                gridcolor="rgba(148,163,184,0.06)",
                color="#FCA5A5",
                tickprefix="",
                ticksuffix="%",
                zeroline=False,
                showgrid=False,
                tickfont=dict(size=13),
            ),
        )
    else:
        fig.update_layout(
            yaxis=dict(
                title=dict(
                    text="Margin (%)" if only_pct_single_axis else "Amount (Kwacha)",
                    font=dict(size=13),
                ),
                gridcolor="rgba(148,163,184,0.12)",
                color="#CBD5E1",
                tickprefix="" if only_pct_single_axis else "K",
                ticksuffix="%" if only_pct_single_axis else "",
                zeroline=True,
                zerolinecolor="rgba(148,163,184,0.3)",
                zerolinewidth=1,
                tickfont=dict(size=13),
            ),
        )

    return fig


def render_cash_flow_forecast_page() -> None:
    render_page_header(
        "Cash Flow Forecast",
        "Project the next 6 months using your current revenue and expenses.",
        "6 months",
    )

    # ── Resolve baseline (single source of truth) ──────────────────────────────
    baseline = st.session_state.get("baseline") or get_baseline(st.session_state)
    st.markdown(
        '<div id="duka-ai-forecast-page-marker" aria-hidden="true" style="position:absolute;width:0;height:0;overflow:hidden"></div>',
        unsafe_allow_html=True,
    )
    if not baseline:
        st.markdown(
            f'<div style="background:rgba(245,158,11,0.1);border-left:3px solid #F59E0B;'
            f'padding:12px 16px;border-radius:6px;color:#FCD34D;margin-bottom:16px;">'
            f'<strong>⚠ No financial data found.</strong><br>'
            f'Go to Chat Advisor and run an analysis first, or enter your numbers in the Business Context form.'
            f'</div>',
            unsafe_allow_html=True,
        )
        if st.button("→ Go to Chat Advisor"):
            st.session_state.active_page = "Chat Advisor"
            st.rerun()
        return

    # Data source banner
    months_note = (
        f" &nbsp;·&nbsp; Average of {baseline['months_of_data']} months of actual data"
        if baseline["months_of_data"] > 1 else ""
    )
    st.markdown(
        f'<div style="background:rgba(37,99,235,0.08);border:1px solid rgba(37,99,235,0.2);'
        f'border-radius:8px;padding:10px 16px;font-size:13px;color:#93C5FD;margin-bottom:16px;">'
        f'<strong>📊 Forecast based on:</strong> {baseline["source"]}'
        f' &nbsp;·&nbsp; Starting monthly revenue: <strong>K{baseline["monthly_revenue"]:,.2f}</strong>'
        f' &nbsp;·&nbsp; Expenses: <strong>K{baseline["monthly_expenses"]:,.2f}</strong>'
        f'{months_note}</div>',
        unsafe_allow_html=True,
    )

    # ── Scenario selector ──────────────────────────────────────────────────────
    if "forecast_scenario" not in st.session_state:
        st.session_state["forecast_scenario"] = "base_case"

    s_col1, s_col2, s_col3 = st.columns(3, gap="small")
    for col, key in zip([s_col1, s_col2, s_col3], ["pessimistic", "base_case", "optimistic"]):
        sc = FE_SCENARIOS[key]
        active = st.session_state["forecast_scenario"] == key
        btn_label = f"{sc['icon']} **{sc['label']}**\n{sc['description']}"
        if col.button(btn_label, key=f"sc_{key}", use_container_width=True,
                      type="primary" if active else "secondary"):
            st.session_state["forecast_scenario"] = key
            for k in list(st.session_state.keys()):
                if k.startswith("forecast_summary_"):
                    del st.session_state[k]
            st.rerun()

    scenario_key = st.session_state["forecast_scenario"]
    forecast = calculate_forecast(baseline, scenario_key, months=6)
    s = forecast["summary"]
    rows = forecast["months"]

    # ── Single-column dashboard (full width). Conversation lives below in a
    # centered, narrower column so chat doesn't sprawl across the whole page.
    st.markdown('<div class="duka-forecast-dashboard">', unsafe_allow_html=True)

    # KPI strip — full width
    k1, k2, k3, k4 = st.columns(4, gap="small")
    profit_now = baseline["monthly_profit"]
    rev_now = baseline["monthly_revenue"]
    exp_now = baseline["monthly_expenses"]
    k1.metric(
        "Month 6 Revenue", f"K{s['final_revenue']:,.0f}",
        f"{(s['final_revenue']-rev_now)/max(rev_now,1)*100:+.0f}% vs now",
    )
    k2.metric(
        "Month 6 Expenses", f"K{s['final_expenses']:,.0f}",
        f"{(s['final_expenses']-exp_now)/max(exp_now,1)*100:+.0f}% vs now",
    )
    profit_delta = s["final_profit"] - profit_now
    k3.metric(
        "Month 6 Profit", f"K{s['final_profit']:,.0f}",
        f"K{profit_delta:+,.0f} vs now",
        delta_color="normal" if profit_delta >= 0 else "inverse",
    )
    k4.metric(
        "Profitable Months", f"{s['profitable_months']}/6",
        "All positive" if s["all_profitable"] else f"Risk from {s['first_loss_month']}",
        delta_color="off" if s["all_profitable"] else "inverse",
    )

    # ── Chart + AI summary side-by-side. Both end at roughly the same height
    # so there is no orphaned empty space when conversation grows below.
    col_chart, col_ai = st.columns([1.75, 1], gap="large")

    with col_chart:

        # Area + Bar chart
        month_labels = [r["month_name"] for r in rows]
        rev_vals    = [r["revenue"]  for r in rows]
        exp_vals    = [r["expenses"] for r in rows]
        profit_vals = [r["profit"]   for r in rows]
        bar_colors  = [THEME["emerald"] if p >= 0 else THEME["red"] for p in profit_vals]

        fig = go.Figure()
        fig.add_trace(go.Scatter(
            x=month_labels, y=rev_vals, name="Revenue",
            mode="lines+markers", line=dict(color=THEME["emerald"], width=2.5),
            fill="tozeroy", fillcolor="rgba(16,185,129,0.07)", marker=dict(size=6),
        ))
        fig.add_trace(go.Scatter(
            x=month_labels, y=exp_vals, name="Expenses",
            mode="lines+markers", line=dict(color=THEME["orange"], width=2.5, dash="dot"),
            fill="tozeroy", fillcolor="rgba(249,115,22,0.05)", marker=dict(size=6),
        ))
        fig.add_trace(go.Bar(
            x=month_labels, y=profit_vals, name="Profit",
            marker_color=bar_colors, opacity=0.75, yaxis="y2",
        ))
        fig.add_hline(
            y=0, line_dash="dash", line_color="rgba(148,163,184,0.4)",
            annotation_text="Break-even", annotation_position="bottom right",
            annotation_font=dict(color="rgba(148,163,184,0.6)", size=10),
        )
        fig.update_layout(
            height=480,
            margin=dict(l=8, r=12, t=110, b=12),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(15,23,42,0.35)",
            font=dict(color="#E2E8F0", size=14, family="Segoe UI, Arial, sans-serif"),
            title=dict(
                text="<b>6-month outlook</b> — revenue, expenses & profit",
                font=dict(size=17, color="#F8FAFC"),
                x=0.005,
                xanchor="left",
                y=0.97,
                yanchor="top",
            ),
            xaxis=dict(showgrid=False, tickfont=dict(size=13), color="#94A3B8"),
            yaxis=dict(
                title=dict(text="Revenue & expenses (K)", font=dict(size=13)),
                gridcolor="rgba(148,163,184,0.12)",
                zeroline=False,
                tickfont=dict(size=13),
                tickprefix="K",
            ),
            yaxis2=dict(
                title=dict(text="Profit (K)", font=dict(size=13)),
                overlaying="y",
                side="right",
                showgrid=False,
                zeroline=True,
                zerolinecolor="rgba(148,163,184,0.35)",
                tickfont=dict(size=13),
                tickprefix="K",
            ),
            legend=dict(
                orientation="h",
                y=1.07,
                yanchor="bottom",
                x=0,
                xanchor="left",
                font=dict(size=12.5),
                bgcolor="rgba(15,23,42,0)",
            ),
            bargap=0.32, hovermode="x unified",
        )
        st.plotly_chart(fig, use_container_width=True)

        # Insight strip
        if s["first_loss_month"]:
            st.markdown(
                f'<div style="background:rgba(239,68,68,0.1);border-left:3px solid {THEME["red"]};'
                f'padding:10px 14px;border-radius:6px;font-size:13px;color:#FCA5A5;margin:4px 0 8px;">'
                f'⚠️ Expenses may exceed revenue from <strong>{s["first_loss_month"]}</strong> under this scenario.</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div style="background:rgba(16,185,129,0.08);border-left:3px solid {THEME["emerald"]};'
                f'padding:10px 14px;border-radius:6px;font-size:13px;color:#6EE7B7;margin:4px 0 8px;">'
                f'✅ All 6 months profitable. Month 6 margin: <strong>{s["final_margin"]:.1f}%</strong>.</div>',
                unsafe_allow_html=True,
            )

        # Monthly breakdown table
        with st.expander("📋 Monthly Breakdown Table", expanded=False):
            def _tr(r: dict, prev_profit: float | None) -> str:
                pc = THEME["emerald"] if r["profit"] >= 0 else THEME["red"]
                if prev_profit is not None:
                    diff = r["profit"] - prev_profit
                    trend_html = (
                        f' <span style="color:{THEME["emerald"]}">↑K{diff:,.0f}</span>'
                        if diff >= 0
                        else f' <span style="color:{THEME["red"]}">↓K{abs(diff):,.0f}</span>'
                    )
                else:
                    trend_html = ""
                warn = " ⚠️" if not r["profitable"] else ""
                return (
                    f'<tr style="border-bottom:1px solid rgba(148,163,184,0.1);">'
                    f'<td style="padding:8px 12px;font-weight:600;">{r["month_name"]}</td>'
                    f'<td style="padding:8px 12px;color:{THEME["emerald"]};">K{r["revenue"]:,.0f}</td>'
                    f'<td style="padding:8px 12px;color:{THEME["orange"]};">K{r["expenses"]:,.0f}</td>'
                    f'<td style="padding:8px 12px;color:{pc};font-weight:600;">K{r["profit"]:,.0f}{warn}{trend_html}</td>'
                    f'<td style="padding:8px 12px;color:{pc};">{r["margin"]:.1f}%</td>'
                    f'</tr>'
                )

            table_html = "".join(
                _tr(r, rows[i-1]["profit"] if i > 0 else None)
                for i, r in enumerate(rows)
            )
            st.markdown(
                '<table style="width:100%;border-collapse:collapse;font-size:13px;color:#E2E8F0;">'
                '<thead><tr style="border-bottom:1px solid rgba(148,163,184,0.25);color:#94A3B8;'
                'font-size:11px;text-transform:uppercase;letter-spacing:.06em;">'
                '<th style="padding:8px 12px;text-align:left;">Month</th>'
                '<th style="padding:8px 12px;text-align:left;">Revenue</th>'
                '<th style="padding:8px 12px;text-align:left;">Expenses</th>'
                '<th style="padding:8px 12px;text-align:left;">Profit</th>'
                '<th style="padding:8px 12px;text-align:left;">Margin</th>'
                f'</tr></thead><tbody>{table_html}</tbody></table>',
                unsafe_allow_html=True,
            )

    # ── AI analyst summary (right side of dashboard) ───────────────────────────
    with col_ai:
        st.markdown(
            '<div class="duka-forecast-analyst">'
            '<div style="color:#94A3B8;font-size:0.72rem;font-weight:800;letter-spacing:0.12em;text-transform:uppercase;margin-bottom:0.35rem;">'
            "AI analyst</div>"
            '<div style="color:#F8FAFC;font-size:1.05rem;font-weight:700;line-height:1.25;">Forecast snapshot</div>'
            '<div style="color:#94A3B8;font-size:0.82rem;margin-top:0.35rem;line-height:1.45;">'
            "A plain-English read of the chart on the left.</div>"
            "</div>",
            unsafe_allow_html=True,
        )

        cache_key = f"forecast_summary_{scenario_key}"
        if cache_key not in st.session_state:
            with st.spinner("🔮 Writing AI forecast narrative…"):
                st.session_state[cache_key] = _generate_forecast_ai_summary(
                    baseline, forecast, scenario_key
                )
        st.markdown(
            '<div class="duka-forecast-narrative-card">' + html.escape(
                st.session_state[cache_key]
            ).replace("\n", "<br>")
            + "</div>",
            unsafe_allow_html=True,
        )

        sugg_pills = [
            ("📈 Profit trend", "Show profit trend"),
            ("📊 Revenue vs expenses", "Revenue vs expenses"),
            ("⚠️ Riskiest months", "Riskiest months"),
            ("💰 Expense breakdown", "Expense breakdown"),
        ]
        st.caption("Quick prompts · tap to ask")
        for idx, (label, prompt_text) in enumerate(sugg_pills):
            if st.button(label, key=f"fpill_{idx}", use_container_width=True):
                st.session_state["forecast_question"] = prompt_text
                st.session_state["forecast_pin_to_question"] = True
                st.rerun()

    # End dashboard wrapper
    st.markdown("</div>", unsafe_allow_html=True)

    # ── Conversation (centered, full-width row) ────────────────────────────────
    if "forecast_chat" not in st.session_state:
        st.session_state["forecast_chat"] = []

    st.markdown(
        '<div id="duka-ai-forecast-analysis-anchor"></div>',
        unsafe_allow_html=True,
    )

    # Centre the conversation in a narrower column so it doesn't stretch wide
    convo_left, convo_main, convo_right = st.columns([1, 4, 1], gap="small")
    with convo_main:
        st.markdown(
            '<div class="duka-forecast-convo-card">'
            '<div class="duka-forecast-convo-head">'
            '<span class="duka-forecast-convo-title">💬 Chat with the analyst</span>'
            '<span class="duka-forecast-convo-sub">Ask anything — answers use the numbers from this scenario.</span>'
            "</div>",
            unsafe_allow_html=True,
        )

        if not st.session_state["forecast_chat"]:
            st.markdown(
                '<div class="duka-forecast-empty">No questions yet — try a quick prompt above '
                'or type below.</div>',
                unsafe_allow_html=True,
            )

        for _msg_idx, msg in enumerate(st.session_state["forecast_chat"]):
            if msg["role"] == "user":
                escaped = (
                    msg["content"]
                    .replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                st.markdown(
                    f'<div class="user-bubble-wrap"><div class="user-bubble">{escaped}</div></div>',
                    unsafe_allow_html=True,
                )
            else:
                with st.chat_message("assistant", avatar="📊"):
                    body = sanitize_forecast_answer_text(msg["content"])
                    st.markdown(body)
                    if msg.get("plotly_chart"):
                        st.plotly_chart(
                            build_plotly_chart(msg["plotly_chart"]),
                            use_container_width=True,
                            key=f"forecast_chat_plotly_{_msg_idx}",
                        )
                    elif msg.get("chart"):
                        render_followup_chart(msg["chart"])

        st.markdown("</div>", unsafe_allow_html=True)

        user_q = st.chat_input("Ask about your forecast...", key="forecast_chat_input")
        if st.session_state.get("forecast_question"):
            user_q = st.session_state.pop("forecast_question")

        if user_q:
            st.session_state["forecast_chat"].append({"role": "user", "content": user_q})
            with st.spinner("📊 Analyzing your forecast…"):
                reply = _handle_forecast_question(
                    user_q, baseline, forecast, scenario_key,
                    st.session_state["forecast_chat"],
                )
            clean_text, chart_data = parse_and_render_chart(reply)
            clean_text = sanitize_forecast_answer_text(clean_text)
            if not chart_data:
                fallback = _build_forecast_page_chart(
                    user_q, baseline, scenario_key, forecast,
                    st.session_state["forecast_chat"],
                )
                if fallback:
                    chart_data = fallback

            # Make sure there is always readable prose AND a real number
            # breakdown alongside the chart. If the AI wrote a thin one-liner
            # (e.g. "Here's a visual representation of the trends:") prepend
            # the auto-generated multi-line breakdown so the user sees real
            # figures before the chart.
            scen_label = forecast.get("scenario", {}).get("label")
            if chart_data:
                short_or_empty = (not clean_text) or len(clean_text) < 80
                lacks_kwacha = "K" not in (clean_text or "")
                if short_or_empty or lacks_kwacha:
                    breakdown = _caption_for_forecast_chart(chart_data, scen_label)
                    if breakdown:
                        if clean_text and len(clean_text) >= 10:
                            clean_text = breakdown + "\n\n" + clean_text
                        else:
                            clean_text = breakdown

            assistant_msg: dict[str, Any] = {"role": "assistant", "content": clean_text}
            if chart_data:
                assistant_msg["plotly_chart"] = chart_data
            st.session_state["forecast_chat"].append(assistant_msg)
            st.session_state["forecast_pin_to_question"] = True
            st.rerun()

        # ChatGPT-style scroll: pin the user's last question to top after a reply
        if st.session_state.pop("forecast_pin_to_question", False):
            inject_pin_to_question_scroll()


def render_scenario_planner_page() -> None:
    render_page_header(
        "Scenario Planner",
        "Adjust sliders to stress-test what happens if sales rise, costs shift, or debt increases.",
        "What-if",
    )

    baseline = st.session_state.get("baseline") or get_baseline(st.session_state)
    base_metrics = st.session_state.get("metrics")

    if not baseline:
        st.markdown(
            '<div class="empty-state">'
            '<div class="empty-state-kicker">Analysis required</div>'
            '<div class="empty-state-title">Run Chat Advisor first</div>'
            '<div class="empty-state-copy">The Scenario Planner needs your actual revenue and expenses before it can calculate anything useful.</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        if st.button("→ Go to Chat Advisor"):
            st.session_state.active_page = "Chat Advisor"
            st.rerun()
        return

    if not base_metrics:
        base_metrics = compute_metrics(baseline)

    # Current baseline banner
    st.markdown(
        f'<div style="background:rgba(37,99,235,0.08);border:1px solid rgba(37,99,235,0.2);'
        f'border-radius:8px;padding:10px 16px;font-size:13px;color:#93C5FD;margin-bottom:16px;">'
        f'<strong>📊 Current baseline:</strong> {baseline["source"]}'
        f' &nbsp;·&nbsp; Revenue <strong>K{baseline["monthly_revenue"]:,.2f}</strong>'
        f' &nbsp;·&nbsp; Expenses <strong>K{baseline["monthly_expenses"]:,.2f}</strong>'
        f' &nbsp;·&nbsp; Profit <strong>K{baseline["monthly_profit"]:,.2f}</strong>'
        f' &nbsp;·&nbsp; Margin <strong>{baseline["profit_margin"]:.1f}%</strong>'
        f'</div>',
        unsafe_allow_html=True,
    )

    render_tool_tip("How to use", "Move the sliders to model a what-if scenario. All metrics recalculate instantly in Python — no AI required.")

    # ── Sliders ────────────────────────────────────────────────────────────────
    ctrl1, ctrl2, ctrl3 = st.columns(3, gap="large")
    with ctrl1:
        revenue_delta = st.slider("Revenue change", -50, 75, 0, 5, format="%d%%",
                                  help="What if your sales grew or fell by this percentage?")
    with ctrl2:
        expense_delta = st.slider("Expense change", -50, 50, 0, 5, format="%d%%",
                                  help="What if your total costs changed by this percentage?")
    with ctrl3:
        extra_debt = st.number_input(
            "Extra monthly debt payment (K)", min_value=0.0, value=0.0, step=100.0,
            help="Add a hypothetical loan repayment to see the impact.",
        )

    # ── Python does ALL the math ───────────────────────────────────────────────
    new_rev = baseline["monthly_revenue"] * (1 + revenue_delta / 100)
    new_exp = baseline["monthly_expenses"] * (1 + expense_delta / 100) + extra_debt
    scenario_baseline = {
        **baseline,
        "monthly_revenue": round(new_rev, 2),
        "monthly_expenses": round(new_exp, 2),
        "monthly_profit": round(new_rev - new_exp, 2),
        "profit_margin": round((new_rev - new_exp) / new_rev * 100, 1) if new_rev > 0 else 0.0,
    }
    new_metrics = compute_metrics(scenario_baseline)
    delta_profit = new_metrics["monthly_profit"] - base_metrics["monthly_profit"]
    delta_health = new_metrics["health_score"] - base_metrics["health_score"]
    delta_margin = new_metrics["profit_margin_pct"] - base_metrics["profit_margin_pct"]

    # ── KPI cards ──────────────────────────────────────────────────────────────
    cols = st.columns(4, gap="small")
    cols[0].metric("New Revenue", f"K{new_rev:,.0f}", f"{revenue_delta:+.0f}%")
    cols[1].metric("New Expenses", f"K{new_exp:,.0f}", f"{expense_delta:+.0f}%")
    cols[2].metric(
        "New Profit", f"K{new_metrics['monthly_profit']:,.0f}",
        f"K{delta_profit:+,.0f}",
        delta_color="normal" if delta_profit >= 0 else "inverse",
    )
    cols[3].metric(
        "Health Score", f"{new_metrics['health_score']}/100",
        f"{delta_health:+.0f} pts",
        delta_color="normal" if delta_health >= 0 else "inverse",
    )

    # Margin and loan label change
    loan_label, loan_color, loan_icon = get_loan_label(new_metrics["profit_margin_pct"])
    m_col1, m_col2 = st.columns(2, gap="large")
    m_col1.metric(
        "New Margin", f"{new_metrics['profit_margin_pct']:.1f}%",
        f"{delta_margin:+.1f}pp vs current",
        delta_color="normal" if delta_margin >= 0 else "inverse",
    )
    m_col2.metric(
        "Loan Readiness", f"{loan_icon} {loan_label}",
        f"Safe borrowing: K{new_metrics['safe_loan_amount']:,.0f}",
        delta_color="off",
    )

    # Verdict
    if delta_health > 10:
        st.success(f"✅ This scenario improves your business health by {delta_health} points.")
    elif delta_health < -10:
        st.error(f"❌ This scenario reduces your business health by {abs(delta_health)} points.")
    elif new_metrics["monthly_profit"] <= 0:
        st.error("❌ This scenario results in a loss. Expenses exceed revenue.")
    else:
        st.info("ℹ️ This scenario has minimal impact on overall business health.")

    # Safe loan formula display
    st.caption(
        f"Safe borrowing amount under this scenario: K{new_metrics['safe_loan_amount']:,.2f} "
        f"({new_metrics['safe_loan_formula']}) — a conservative SME lending rule of thumb."
    )

    # ── Send to Chat button ────────────────────────────────────────────────────
    st.divider()
    if st.button("💬 Ask Duka AI to explain this scenario", type="primary"):
        scenario_desc = (
            f"I modelled a scenario: revenue {revenue_delta:+.0f}%, "
            f"expenses {expense_delta:+.0f}%"
            + (f", extra debt payment K{extra_debt:,.0f}/month" if extra_debt > 0 else "")
            + f". New profit would be K{new_metrics['monthly_profit']:,.0f} "
            f"(was K{base_metrics['monthly_profit']:,.0f}). "
            f"Health score changes from {base_metrics['health_score']} to {new_metrics['health_score']}. "
            f"Margin changes from {base_metrics['profit_margin_pct']:.1f}% to {new_metrics['profit_margin_pct']:.1f}%. "
            "What should I focus on given these numbers?"
        )
        st.session_state["pending_chat_question"] = scenario_desc
        st.session_state.active_page = "Chat Advisor"
        st.rerun()

def render_loan_calculator_page() -> None:
    render_page_header(
        "Loan Calculator",
        "Estimate monthly repayment and compare it against current profit.",
        "Repayment",
    )
    report = require_report()
    if not report:
        return
    loan = report["loan"]
    cashflow = report["cashflow"]
    render_tool_tip("Borrowing guardrail", "A repayment above roughly 20-30% of current profit is usually a warning sign for a small business.")
    control_cols = st.columns(3, gap="large")
    with control_cols[0]:
        amount = st.number_input("Loan amount (K)", min_value=0.0, value=float(loan.get("suggested_loan_amount", 0.0)), step=500.0)
    with control_cols[1]:
        annual_rate = st.slider("Annual interest rate", 0.0, 60.0, 28.0, 0.5, format="%.1f%%")
    with control_cols[2]:
        months = st.slider("Repayment period (months)", 1, 60, 12)
    monthly_rate = annual_rate / 100 / 12
    payment = amount / months if monthly_rate == 0 else amount * monthly_rate / (1 - (1 + monthly_rate) ** -months)
    monthly_profit = max(float(cashflow.get("profit", 0.0)), 0.0)
    burden = (payment / monthly_profit * 100) if monthly_profit else 0
    cols = st.columns(3)
    cols[0].metric("Monthly repayment", format_currency(payment))
    cols[1].metric("Total repayment", format_currency(payment * months))
    cols[2].metric("Profit used", f"{burden:.1f}%")
    if monthly_profit <= 0:
        st.error("Current profit is not positive, so new borrowing is not advisable from these numbers.")
    elif burden > 30:
        st.warning("This repayment would use a high share of current profit.")
    else:
        st.success("This repayment is within a more manageable range based on current profit.")
    render_loan_readiness_card(report)


def get_expense_breakdown(report: dict[str, Any]) -> dict[str, float]:
    parsed = report.get("parsed_data", {})
    breakdown = {str(k).replace("_", " ").title(): float(v or 0) for k, v in (parsed.get("expenses_breakdown") or {}).items() if float(v or 0) > 0}
    if breakdown:
        return breakdown
    doc_breakdown = report.get("document_analysis", {}).get("expenses_breakdown") or {}
    return {str(k).replace("_", " ").title(): float(v or 0) for k, v in doc_breakdown.items() if float(v or 0) > 0}


def _expense_identity_reply() -> str:
    return (
        "I'm **Duka AI's Expense Analyst** — I focus only on **where your Kwacha is going**. "
        "I work from the verified expense breakdown on this page (no guesses): I can rank your top "
        "categories, show what % of revenue each one eats, and suggest **specific** cost cuts.\n\n"
        "Try: *Which category is hurting me most?* · *How can I cut transport costs?* · "
        "*Where should I trim 10% of expenses?*"
    )


def _build_expense_chat_context(
    report: dict[str, Any], breakdown: dict[str, float], frame: "pd.DataFrame"
) -> str:
    cf = report.get("cashflow") or {}
    revenue = float(cf.get("revenue", 0) or 0)
    total_exp = float(frame["Amount"].sum())
    rows = frame.sort_values("Amount", ascending=False).to_dict("records")

    bp = report.get("business_profile") or {}
    bt = (bp.get("business_type") or "").strip() or "SME"
    loc = (bp.get("location") or "").strip() or "Zambia"

    lines: list[str] = []
    lines.append("VERIFIED EXPENSE BREAKDOWN (Python-computed, do NOT invent figures):\n")
    lines.append(
        f"- Business: {bt} · Location: {loc}\n"
        f"- Monthly revenue: K{revenue:,.0f}\n"
        f"- Total monthly expenses: K{total_exp:,.0f}"
        + (f" ({total_exp / revenue * 100:.1f}% of revenue)\n" if revenue > 0 else "\n")
    )
    lines.append("- Categories (largest → smallest):\n")
    for r in rows:
        cat = str(r["Category"])
        amt = float(r["Amount"])
        pct_rev = (amt / revenue * 100) if revenue > 0 else 0.0
        pct_exp = (amt / total_exp * 100) if total_exp > 0 else 0.0
        lines.append(
            f"  · {cat}: K{amt:,.0f}  ({pct_exp:.1f}% of expenses, {pct_rev:.1f}% of revenue)\n"
        )
    return "".join(lines)


def _classify_expense_chat_input(question: str) -> str | None:
    low = (question or "").lower().strip()
    if not low:
        return "empty"
    if any(p in low for p in ("who are you", "what are you", "what do you do", "introduce yourself")):
        return "identity"
    if low in ("thanks", "thank you", "ty", "cheers", "ta"):
        return "thanks"
    return None


def _handle_expense_question(
    question: str,
    report: dict,
    breakdown: dict[str, float],
    frame: "pd.DataFrame",
    chat_history: list[dict],
) -> str:
    from agents import request_llm_chat

    canned = _classify_expense_chat_input(question)
    if canned == "identity":
        return _expense_identity_reply()
    if canned == "thanks":
        return "Glad it helped — ask anything else about your expense breakdown."
    if canned == "empty":
        return "Type a question about your expense categories or how to cut them."

    ctx = _build_expense_chat_context(report, breakdown, frame)
    system = (
        "You are **Duka AI's Expense Analyst** for **Zambian SMEs**. "
        "You ONLY answer using the VERIFIED EXPENSE BREAKDOWN below — never invent categories or amounts. "
        "When you reference a category, **always** quote its exact K amount and its % of revenue or % of expenses. "
        "Be specific with cost-cutting advice (which line, by how much in K, and why it's realistic for that business type/location). "
        "Use **K** for Kwacha. Reply in 3–6 short sentences or short bullets — no fluff.\n\n"
        f"{ctx}"
    )

    messages = [
        {"role": "system", "content": system},
        *[{"role": m["role"], "content": m["content"]} for m in chat_history[-8:]],
        {"role": "user", "content": question},
    ]
    result = request_llm_chat(messages, temperature=0.2, max_tokens=500)
    return result or "I couldn't generate a response right now. Please try again."


def render_expense_analyzer_page() -> None:
    render_page_header(
        "Expense Analyzer",
        "Break down where money is going and spot the biggest cost pressure.",
        "Donut + AI agent",
    )
    report = require_report()
    if not report:
        return

    st.markdown(
        '<div id="duka-ai-expense-analyzer-page-marker" aria-hidden="true" '
        'style="position:absolute;width:0;height:0;overflow:hidden"></div>',
        unsafe_allow_html=True,
    )

    cashflow = report["cashflow"]
    breakdown = get_expense_breakdown(report)
    if not breakdown:
        breakdown = {"Total expenses": float(cashflow["expenses"])}
        st.info("Detailed expense categories were not provided, so this view shows total expenses only.")
    frame = pd.DataFrame(
        [{"Category": k, "Amount": v} for k, v in breakdown.items()]
    ).sort_values("Amount", ascending=False)
    top_row = frame.iloc[0]

    # Two-column layout: snapshot/chart left, AI agent chat right
    col_chart, col_chat = st.columns([3, 2], gap="large")

    with col_chart:
        metric_cols = st.columns(3)
        metric_cols[0].metric("Total expenses", format_currency(float(frame["Amount"].sum())))
        metric_cols[1].metric("Largest category", str(top_row["Category"]))
        metric_cols[2].metric("Largest amount", format_currency(float(top_row["Amount"])))
        st.dataframe(frame, use_container_width=True, hide_index=True)
        fig = go.Figure(data=[go.Pie(labels=frame["Category"], values=frame["Amount"], hole=0.58)])
        fig.update_layout(
            height=380,
            margin=dict(l=8, r=8, t=16, b=8),
            paper_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#E2E8F0", family="Segoe UI, Arial, sans-serif"),
            legend=dict(orientation="h", y=-0.05),
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_chat:
        st.markdown(
            '<div id="duka-ai-expense-analyzer-chat-anchor"></div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="duka-forecast-convo-card">'
            '<div class="duka-forecast-convo-head">'
            '<span class="duka-forecast-convo-title">🔍 Ask the Expense Analyst</span>'
            '<span class="duka-forecast-convo-sub">'
            "Grounded in your verified categories — no invented numbers."
            "</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        sugg_pills = [
            "🥇 Which category hurts me most?",
            "✂️ Where can I cut 10%?",
            "🚚 How do I reduce transport?",
            "👥 Are wages too high?",
            "💡 Which line saves me the most?",
            "📉 What % of revenue goes to expenses?",
        ]
        p1, p2 = st.columns(2)
        for idx, pill in enumerate(sugg_pills):
            col = p1 if idx % 2 == 0 else p2
            if col.button(pill, key=f"epill_{idx}", use_container_width=True):
                st.session_state["expense_question"] = pill

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        if "expense_chat" not in st.session_state:
            st.session_state["expense_chat"] = []

        if not st.session_state["expense_chat"]:
            st.markdown(
                '<div class="duka-forecast-empty">No messages yet — tap a suggestion '
                "or type below.</div>",
                unsafe_allow_html=True,
            )

        for msg in st.session_state["expense_chat"]:
            if msg["role"] == "user":
                escaped = (
                    msg["content"]
                    .replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                st.markdown(
                    f'<div class="user-bubble-wrap"><div class="user-bubble">{escaped}</div></div>',
                    unsafe_allow_html=True,
                )
            else:
                with st.chat_message("assistant", avatar="🔍"):
                    st.markdown(msg["content"])

        st.markdown("</div>", unsafe_allow_html=True)

        user_q = st.chat_input("Ask about your expenses...", key="expense_chat_input")
        if st.session_state.get("expense_question"):
            user_q = st.session_state.pop("expense_question")

        if user_q:
            st.session_state["expense_chat"].append({"role": "user", "content": user_q})
            with st.spinner("🔍 Analyzing your expenses…"):
                reply = _handle_expense_question(
                    user_q, report, breakdown, frame, st.session_state["expense_chat"]
                )
            st.session_state["expense_chat"].append({"role": "assistant", "content": reply})
            st.session_state["expense_pin_to_question"] = True
            st.rerun()

        if st.session_state.pop("expense_pin_to_question", False):
            inject_pin_to_question_scroll()


def build_report_text(report: dict[str, Any]) -> str:
    cashflow = report["cashflow"]
    loan = report["loan"]
    actions = "\n".join(f"- {action}" for action in report.get("final_recommended_actions", []))
    return (
        "Duka AI Summary\n\n"
        f"{report['final_summary']}\n\n"
        f"Revenue: {format_currency(cashflow['revenue'])}\n"
        f"Expenses: {format_currency(cashflow['expenses'])}\n"
        f"{'Net Loss' if cashflow['profit'] < 0 else 'Profit'}: {format_currency(abs(cashflow['profit']))}\n"
        f"{'Loss' if cashflow['profit'] < 0 else 'Profit'} margin: {cashflow.get('profit_margin', 0):.1f}%\n"
        f"Loan readiness: {loan['loan_readiness_score']}/100\n\n"
        f"Recommended actions:\n{actions}\n"
    )


def build_pdf_report_bytes(report: dict[str, Any]) -> bytes:
    text = build_report_text(report)
    lines: list[str] = []
    for paragraph in text.splitlines():
        wrapped = textwrap.wrap(paragraph, width=88) if paragraph.strip() else [""]
        lines.extend(wrapped)

    content_parts = ["BT", "/F1 11 Tf", "50 790 Td", "14 TL"]
    for line in lines[:52]:
        safe = line.encode("latin-1", "replace").decode("latin-1")
        safe = safe.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content_parts.append(f"({safe}) Tj")
        content_parts.append("T*")
    content_parts.append("ET")
    stream = "\n".join(content_parts).encode("latin-1")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{i} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii"))
    return bytes(pdf)


def _build_plaintext_pdf_from_full(full: dict) -> bytes:
    """Raw-bytes PDF fallback used when ReportLab is unavailable or broken."""
    snap  = full.get("snapshot", {})
    loan  = full.get("loan", {})
    fc6   = full.get("forecast_6m", {})
    mkt   = full.get("market", {})

    paragraphs: list[str] = [
        "Duka AI - Financial Health Report",
        f"{full.get('business_name','Your Business')} | "
        f"{full.get('location','Zambia')} | Generated {full.get('generated_at','')}",
        "",
        "EXECUTIVE SUMMARY",
        full.get("executive_summary", "—"),
        "",
        "FINANCIAL SNAPSHOT",
        f"Revenue:      K{snap.get('revenue', 0):,.0f}",
        f"Expenses:     K{snap.get('expenses', 0):,.0f}  ({snap.get('expense_pct', 0):.0f}% of revenue)",
        f"{'Net Loss' if snap.get('profit',0) < 0 else 'Profit'}:       K{abs(snap.get('profit', 0)):,.0f}  ({snap.get('margin_pct', 0):.1f}% margin)",
        f"Health Score: {snap.get('health_score', 0)}/100 — {snap.get('health_label', 'N/A')}",
        f"Cash Flow:    {snap.get('cash_flow_status', '—')}",
        "",
        "LOAN READINESS",
        f"Score:          {loan.get('score', 0)}/100 — {loan.get('status', '—')}",
        f"Safe borrowing: K{loan.get('safe_amount', 0):,.0f}",
        f"Formula:        {loan.get('formula', '—')}",
        "",
        "6-MONTH FORECAST (BASE CASE)",
        f"Month 6 Profit: K{fc6.get('base_profit_m6', 0):,.0f}",
        f"Total Profit:   K{fc6.get('total_profit', 0):,.0f}",
        f"Outlook: {'All 6 months profitable' if fc6.get('all_profitable') else 'Risk from ' + str(fc6.get('first_loss_month',''))}",
        "",
        "MARKET CONDITIONS",
        mkt.get("summary", "—"),
        "",
        "TOP 3 RECOMMENDATIONS",
    ]
    for i, r in enumerate(full.get("recommendations", []), 1):
        paragraphs.append(
            f"{i}. {r.get('action','—')}  →  {r.get('impact','—')}  ({r.get('timeline','—')})"
        )
    risks = full.get("risks", [])
    if risks:
        paragraphs += ["", "RISKS TO WATCH"] + [f"!  {r}" for r in risks]
    paragraphs += ["", "Generated by Duka AI | AMD MI300X | Qwen Model"]

    paragraphs = [_pdf_safe_text(p) for p in paragraphs]

    # Wrap and build raw PDF stream
    lines: list[str] = []
    for para in paragraphs:
        wrapped = textwrap.wrap(para, width=88) if para.strip() else [""]
        lines.extend(wrapped)

    content_parts = ["BT", "/F1 11 Tf", "50 790 Td", "14 TL"]
    for line in lines[:52]:
        safe = line.encode("latin-1", "replace").decode("latin-1")
        safe = safe.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content_parts += [f"({safe}) Tj", "T*"]
    content_parts.append("ET")
    stream = "\n".join(content_parts).encode("latin-1")

    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842]"
        b" /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode() + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for i, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{i} 0 obj\n".encode())
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects)+1}\n".encode())
    pdf.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        pdf.extend(f"{off:010d} 00000 n \n".encode())
    pdf.extend(
        f"trailer\n<< /Size {len(objects)+1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode()
    )
    return bytes(pdf)


def _report_text_escape(s: object) -> str:
    """Escape LLM/user text for safe insertion into report HTML."""
    return html.escape(str(s or ""), quote=False)


def _pdf_safe_text(s: object) -> str:
    """Normalize text for PDF built-in fonts (WinAnsi); avoids '?' from latin-1 mapping."""
    t = str(s or "")
    for a, b in (
        ("\u2014", "-"), ("\u2013", "-"), ("\u2212", "-"), ("\u2022", "*"),
        ("\u2192", "->"), ("\u2019", "'"), ("\u2018", "'"), ("\u201c", '"'),
        ("\u201d", '"'), ("\u2026", "..."), ("\u00a0", " "),
    ):
        t = t.replace(a, b)
    return t.encode("latin-1", "replace").decode("latin-1")


def _pdf_escape(s: object) -> str:
    """Escape for ReportLab Paragraph XML/HTML subset."""
    from xml.sax.saxutils import escape

    return escape(_pdf_safe_text(s), {"'": "&apos;", '"': "&quot;"})


def _pdf_para(text: object, style):
    """ReportLab Paragraph with safe body text (newlines -> br)."""
    from reportlab.platypus import Paragraph

    t = _pdf_escape(text).replace("\n", "<br/>")
    return Paragraph(t, style)


def build_full_report_pdf(full: dict) -> bytes:
    """Build a styled PDF from generate_full_report(); falls back if ReportLab is missing or fails."""
    try:
        return _render_reportlab_pdf(full)
    except Exception:
        return _build_plaintext_pdf_from_full(full)


def _render_reportlab_pdf(full: dict) -> bytes:
    """Premium ReportLab layout (requires `reportlab` package)."""
    from io import BytesIO

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        HRFlowable,
        KeepTogether,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.8 * cm,
    )

    base = getSampleStyleSheet()
    C_DARK = colors.HexColor("#0F172A")
    C_BLUE = colors.HexColor("#2563EB")
    C_SLATE = colors.HexColor("#1e293b")
    C_MID = colors.HexColor("#64748B")
    C_LITE = colors.HexColor("#F1F5F9")
    C_CARD = colors.HexColor("#f8fafc")
    C_GRID = colors.HexColor("#CBD5E1")
    C_GRN = colors.HexColor("#059669")
    C_RED = colors.HexColor("#DC2626")
    C_AMB = colors.HexColor("#D97706")

    def _style(name, **kw):
        return ParagraphStyle(name, parent=base["Normal"], **kw)

    ST_SUB = _style(
        "RP_SUB",
        fontSize=9.5,
        textColor=colors.HexColor("#cbd5e1"),
        leading=13,
    )
    NM = _style("RP_NM", fontSize=10, textColor=C_DARK, leading=15)
    NM_SM = _style("RP_NM_SM", fontSize=9, textColor=C_DARK, leading=13)
    SM = _style("RP_SM", fontSize=9, textColor=C_MID, leading=13)
    H3 = _style(
        "RP_H3",
        fontSize=10,
        textColor=C_MID,
        fontName="Helvetica-Bold",
        spaceBefore=6,
        spaceAfter=2,
    )
    FT = _style("RP_FT", fontSize=8, textColor=C_MID, alignment=1)
    TITLE = _style(
        "RP_TITLE",
        fontSize=20,
        textColor=colors.white,
        fontName="Helvetica-Bold",
        leading=24,
        spaceAfter=4,
    )

    snap = full.get("snapshot", {})
    h_score = int(snap.get("health_score", 0) or 0)
    h_color = C_GRN if h_score >= 70 else (C_AMB if h_score >= 40 else C_RED)

    loan = full.get("loan", {})
    l_score = int(loan.get("score", 0) or 0)
    l_color = C_GRN if l_score >= 70 else (C_AMB if l_score >= 40 else C_RED)

    def _hr(thickness=0.8, c=C_GRID):
        return HRFlowable(width="100%", thickness=thickness, color=c, spaceAfter=8, spaceBefore=8)

    def _tbl(data, col_widths, style_cmds):
        t = Table(data, colWidths=col_widths)
        t.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    (
                        "ROWBACKGROUNDS",
                        (0, 0),
                        (-1, -1),
                        [colors.white, C_CARD],
                    ),
                    ("GRID", (0, 0), (-1, -1), 0.35, C_GRID),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ]
                + style_cmds
            )
        )
        return t

    def _section_bar(label: str):
        lab = _style(
            "RP_SEC",
            fontSize=9,
            textColor=colors.white,
            fontName="Helvetica-Bold",
            leading=11,
            letterSpacing=1.2,
        )
        tb = Table([[Paragraph(_pdf_escape(label.upper()), lab)]], colWidths=[17.4 * cm])
        tb.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), C_BLUE),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ]
            )
        )
        return tb

    story = []

    # ── Hero header ───────────────────────────────────────────────────────────
    hero = Table(
        [
            [Paragraph(_pdf_escape("Duka AI Financial Health Report"), TITLE)],
            [
                _pdf_para(
                    f"{full.get('business_name', 'Your Business')} | "
                    f"{full.get('location', 'Zambia')} | "
                    f"Generated {full.get('generated_at', '')}",
                    ST_SUB,
                )
            ],
        ],
        colWidths=[17.4 * cm],
    )
    hero.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), C_SLATE),
                ("TOPPADDING", (0, 0), (-1, -1), 16),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
                ("LEFTPADDING", (0, 0), (-1, -1), 14),
                ("RIGHTPADDING", (0, 0), (-1, -1), 14),
                ("LINEABOVE", (0, 0), (-1, 0), 3, C_BLUE),
            ]
        )
    )
    story.append(hero)

    _snap_profit = float(snap.get("profit", 0) or 0)
    _pr_col = C_RED if _snap_profit < 0 else C_GRN
    _rev = float(snap.get("revenue", 0) or 0)
    _exp = float(snap.get("expenses", 0) or 0)

    kpi = Table(
        [
            [
                "MONTHLY REVENUE",
                "MONTHLY EXPENSES",
                "NET " + ("LOSS" if _snap_profit < 0 else "PROFIT"),
            ],
            [
                f"K{_rev:,.0f}",
                f"K{_exp:,.0f}",
                f"K{abs(_snap_profit):,.0f}",
            ],
        ],
        colWidths=[5.8 * cm, 5.8 * cm, 5.8 * cm],
    )
    kpi.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), C_LITE),
                ("BOX", (0, 0), (-1, -1), 0.75, C_GRID),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 7.5),
                ("TEXTCOLOR", (0, 0), (-1, 0), C_MID),
                ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 1), (-1, 1), 14),
                ("TEXTCOLOR", (0, 1), (1, 1), C_DARK),
                ("TEXTCOLOR", (2, 1), (2, 1), _pr_col),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ]
        )
    )
    story.append(kpi)
    story.append(Spacer(1, 0.35 * cm))

    # ── Executive Summary ─────────────────────────────────────────────────────
    story.append(_section_bar("Executive summary"))
    story.append(Spacer(1, 0.15 * cm))
    story.append(_pdf_para(full.get("executive_summary", ""), NM))
    story.append(Spacer(1, 0.3 * cm))

    # ── Financial Snapshot ───────────────────────────────────────────────────
    story.append(_section_bar("Financial snapshot"))
    story.append(Spacer(1, 0.15 * cm))
    _snap_p_label = "Net Loss" if _snap_profit < 0 else "Profit"
    _snap_p_value = (
        f"K{abs(_snap_profit):,.0f}  ({float(snap.get('margin_pct', 0) or 0):.1f}% margin)"
    )
    snap_data = [
        ["Revenue", f"K{float(snap.get('revenue', 0) or 0):,.0f}"],
        [
            "Expenses",
            f"K{float(snap.get('expenses', 0) or 0):,.0f}  "
            f"({float(snap.get('expense_pct', 0) or 0):.0f}% of revenue)",
        ],
        [_snap_p_label, _snap_p_value],
        [
            "Health Score",
            f"{h_score}/100 - {_pdf_safe_text(snap.get('health_label', 'N/A'))}",
        ],
        ["Cash Flow", _pdf_safe_text(snap.get("cash_flow_status", "N/A"))],
        ["Data Source", _pdf_safe_text(snap.get("source", "-"))],
    ]
    story.append(
        KeepTogether(
            [
                _tbl(
                    snap_data,
                    [5.2 * cm, 11.8 * cm],
                    [
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("TEXTCOLOR", (1, 2), (1, 2), _pr_col),
                        ("TEXTCOLOR", (1, 3), (1, 3), h_color),
                    ],
                )
            ]
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    # ── Expense Breakdown ────────────────────────────────────────────────────
    cats = full.get("expense_categories", [])
    if cats:
        story.append(_section_bar("Expense breakdown"))
        story.append(Spacer(1, 0.15 * cm))
        exp_data = [["Category", "Amount", "% of Expenses"]] + [
            [
                _pdf_safe_text(c.get("name", "")),
                f"K{float(c.get('amount', 0) or 0):,.0f}",
                f"{float(c.get('pct_of_expenses', 0) or 0):.1f}%",
            ]
            for c in cats[:6]
        ]
        story.append(
            _tbl(
                exp_data,
                [7 * cm, 4 * cm, 4 * cm],
                [
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BACKGROUND", (0, 0), (-1, 0), C_SLATE),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, C_CARD],
                    ),
                ],
            )
        )
        story.append(Spacer(1, 0.3 * cm))

    # ── Cash Flow Analysis ───────────────────────────────────────────────────
    cf_sect = full.get("cash_flow", {})
    if cf_sect.get("summary"):
        story.append(_section_bar("Cash flow analysis"))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_pdf_para(cf_sect["summary"], NM))
        _ins = cf_sect.get("insight") or cf_sect.get("reasoning")
        if _ins:
            story.append(Spacer(1, 0.12 * cm))
            story.append(
                Paragraph(f"<i>{_pdf_escape(_ins)}</i>", SM)
            )
        story.append(Spacer(1, 0.3 * cm))

    # ── Loan Readiness ───────────────────────────────────────────────────────
    story.append(_section_bar("Loan readiness"))
    story.append(Spacer(1, 0.15 * cm))
    loan_data = [
        ["Score", f"{l_score}/100 - {_pdf_safe_text(loan.get('status', ''))}"],
        ["Risk Level", _pdf_safe_text(loan.get("risk_level", "-"))],
        ["Safe Borrowing", f"K{float(loan.get('safe_amount', 0) or 0):,.0f}"],
        ["Formula", _pdf_safe_text(loan.get("formula", "-"))],
    ]
    story.append(
        _tbl(
            loan_data,
            [5.2 * cm, 11.8 * cm],
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("TEXTCOLOR", (1, 0), (1, 0), l_color),
            ],
        )
    )
    improve = loan.get("improve", [])
    if improve:
        story.append(Spacer(1, 0.18 * cm))
        story.append(Paragraph("<b>How to improve</b>", H3))
        for item in improve:
            story.append(
                Paragraph(f"- {_pdf_escape(item)}", NM)
            )
    story.append(Spacer(1, 0.3 * cm))

    # ── Market Conditions ─────────────────────────────────────────────────────
    mkt = full.get("market", {})
    if mkt.get("summary"):
        story.append(_section_bar("Market conditions"))
        story.append(Spacer(1, 0.15 * cm))
        story.append(_pdf_para(mkt["summary"], NM))
        if mkt.get("opportunity"):
            story.append(Spacer(1, 0.12 * cm))
            story.append(
                Paragraph(
                    f"<b>Opportunity:</b> {_pdf_escape(mkt['opportunity'])}",
                    NM,
                )
            )
        story.append(Spacer(1, 0.3 * cm))

    # ── 6-Month Forecast ──────────────────────────────────────────────────────
    fc6 = full.get("forecast_6m", {})
    story.append(_section_bar("6-month forecast (base case)"))
    story.append(Spacer(1, 0.15 * cm))
    fc_status = (
        "All 6 months profitable"
        if fc6.get("all_profitable")
        else f"Risk from {_pdf_safe_text(fc6.get('first_loss_month', 'month unknown'))}"
    )
    fc_data = [
        ["Month 6 profit", f"K{float(fc6.get('base_profit_m6', 0) or 0):,.0f}"],
        ["Total 6-mo profit", f"K{float(fc6.get('total_profit', 0) or 0):,.0f}"],
        ["Profitable months", f"{fc6.get('profitable_months', 0)}/6"],
        ["Outlook", fc_status],
    ]
    story.append(
        _tbl(
            fc_data,
            [6 * cm, 11 * cm],
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ],
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    # ── Recommendations ───────────────────────────────────────────────────────
    recs = full.get("recommendations", [])
    if recs:
        story.append(_section_bar("Top recommendations"))
        story.append(Spacer(1, 0.15 * cm))
        hdr = _style(
            "RH",
            fontSize=9,
            textColor=colors.white,
            fontName="Helvetica-Bold",
            leading=11,
        )
        rec_rows = [
            [
                Paragraph(_pdf_escape("#"), hdr),
                Paragraph(_pdf_escape("Action"), hdr),
                Paragraph(_pdf_escape("Impact"), hdr),
                Paragraph(_pdf_escape("Timeline"), hdr),
            ]
        ]
        for i, r in enumerate(recs):
            rec_rows.append(
                [
                    Paragraph(_pdf_escape(str(i + 1)), NM_SM),
                    Paragraph(_pdf_escape(_pdf_safe_text(r.get("action", ""))), NM_SM),
                    Paragraph(_pdf_escape(_pdf_safe_text(r.get("impact", ""))), NM_SM),
                    Paragraph(_pdf_escape(_pdf_safe_text(r.get("timeline", ""))), NM_SM),
                ]
            )
        rt = Table(rec_rows, colWidths=[0.9 * cm, 7.3 * cm, 4.6 * cm, 4.6 * cm])
        rt.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), C_DARK),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, C_CARD]),
                    ("GRID", (0, 0), (-1, -1), 0.35, C_GRID),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(rt)
        story.append(Spacer(1, 0.3 * cm))

    # ── Risks ────────────────────────────────────────────────────────────────
    risks = full.get("risks", [])
    if risks:
        story.append(_section_bar("Risks to watch"))
        story.append(Spacer(1, 0.15 * cm))
        NM_RISK = _style("RP_RISK", fontSize=10, textColor=C_AMB, leading=15)
        risk_rows = [
            [Paragraph(f"<b>!</b> {_pdf_escape(_pdf_safe_text(risk))}", NM_RISK)]
            for risk in risks
        ]
        rt2 = Table(risk_rows, colWidths=[17.4 * cm])
        rt2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fffbeb")),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#fbbf24")),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(rt2)
        story.append(Spacer(1, 0.3 * cm))

    # ── Footer ───────────────────────────────────────────────────────────────
    story.append(_hr(1, C_GRID))
    story.append(
        Paragraph(
            _pdf_escape(
                "Generated by Duka AI  |  Powered by AMD MI300X  |  Qwen Model"
            ),
            FT,
        )
    )

    doc.build(story)
    return buf.getvalue()


def build_excel_report_bytes(full: dict) -> tuple[bytes, str]:
    """Return (bytes, mime_type) for an Excel or CSV of the report data."""
    from io import BytesIO, StringIO

    snap = full.get("snapshot", {})
    loan = full.get("loan", {})
    fc6  = full.get("forecast_6m", {})

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # ── Sheet 1: Overview ──────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Overview"

        hdr_fill = PatternFill("solid", fgColor="0F172A")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        sub_fill = PatternFill("solid", fgColor="1D4ED8")
        sub_font = Font(bold=True, color="FFFFFF", size=10)
        bold     = Font(bold=True)
        thin     = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        def _hdr(ws, row, col, text, span=1):
            ws.cell(row, col, text).font = hdr_font
            ws.cell(row, col).fill = hdr_fill
            ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="center")
            if span > 1:
                ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)

        def _sub(ws, row, col, text):
            ws.cell(row, col, text).font = sub_font
            ws.cell(row, col).fill = sub_fill

        def _row(ws, row, label, value):
            ws.cell(row, 1, label).font = bold
            ws.cell(row, 2, value)
            for c in (1, 2):
                ws.cell(row, c).border = thin

        r = 1
        _hdr(ws, r, 1, "Duka AI — Financial Health Report", 2); r += 1
        ws.cell(r, 1, f"{full.get('business_name','')} | {full.get('location','')} | {full.get('generated_at','')}"); r += 2

        _sub(ws, r, 1, "EXECUTIVE SUMMARY"); ws.cell(r, 2).fill = sub_fill; r += 1
        ws.cell(r, 1, full.get("executive_summary","")).alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2); r += 2

        _sub(ws, r, 1, "FINANCIAL SNAPSHOT"); ws.cell(r, 2).fill = sub_fill; r += 1
        _xls_profit = snap.get("profit", 0)
        _xls_p_label = "Net Loss" if _xls_profit < 0 else "Profit"
        _row(ws, r, "Revenue",     f"K{snap.get('revenue',0):,.0f}"); r += 1
        _row(ws, r, "Expenses",    f"K{snap.get('expenses',0):,.0f} ({snap.get('expense_pct',0):.0f}% of revenue)"); r += 1
        _row(ws, r, _xls_p_label,  f"K{abs(_xls_profit):,.0f} ({snap.get('margin_pct',0):.1f}% margin)"); r += 1
        _row(ws, r, "Health Score",f"{snap.get('health_score',0)}/100 — {snap.get('health_label','')}"); r += 1
        _row(ws, r, "Cash Flow",   snap.get("cash_flow_status","—")); r += 2

        _sub(ws, r, 1, "LOAN READINESS"); ws.cell(r, 2).fill = sub_fill; r += 1
        _row(ws, r, "Score",          f"{loan.get('score',0)}/100 — {loan.get('status','')}"); r += 1
        _row(ws, r, "Risk Level",     loan.get("risk_level","—")); r += 1
        _row(ws, r, "Safe Borrowing", f"K{loan.get('safe_amount',0):,.0f}"); r += 1
        _row(ws, r, "Formula",        loan.get("formula","—")); r += 2

        _sub(ws, r, 1, "6-MONTH FORECAST (BASE CASE)"); ws.cell(r, 2).fill = sub_fill; r += 1
        outlook = "All months profitable" if fc6.get("all_profitable") else f"Risk from {fc6.get('first_loss_month','')}"
        _row(ws, r, "Month 6 Profit",    f"K{fc6.get('base_profit_m6',0):,.0f}"); r += 1
        _row(ws, r, "Total 6-Mo Profit", f"K{fc6.get('total_profit',0):,.0f}"); r += 1
        _row(ws, r, "Outlook",           outlook); r += 2

        # Recommendations
        recs = full.get("recommendations", [])
        if recs:
            _sub(ws, r, 1, "TOP RECOMMENDATIONS"); ws.cell(r, 2).fill = sub_fill; ws.cell(r, 3).fill = sub_fill; r += 1
            ws.cell(r, 1, "Action").font = bold; ws.cell(r, 2, "Impact").font = bold; ws.cell(r, 3, "Timeline").font = bold; r += 1
            for rec in recs:
                ws.cell(r, 1, rec.get("action",""))
                ws.cell(r, 2, rec.get("impact",""))
                ws.cell(r, 3, rec.get("timeline",""))
                r += 1
            r += 1

        # Risks
        risks = full.get("risks", [])
        if risks:
            _sub(ws, r, 1, "RISKS TO WATCH"); ws.cell(r, 2).fill = sub_fill; r += 1
            for risk in risks:
                ws.cell(r, 1, f"⚠  {risk}"); r += 1
            r += 1

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18

        # ── Sheet 2: Expense Breakdown ────────────────────────────────────────
        cats = full.get("expense_categories", [])
        if cats:
            ws2 = wb.create_sheet("Expense Breakdown")
            _hdr(ws2, 1, 1, "Category")
            _hdr(ws2, 1, 2, "Amount (K)")
            _hdr(ws2, 1, 3, "% of Expenses")
            for i, c in enumerate(cats, start=2):
                ws2.cell(i, 1, c.get("name",""))
                ws2.cell(i, 2, round(c.get("amount",0), 2))
                ws2.cell(i, 3, f"{c.get('pct_of_expenses',0):.1f}%")
            ws2.column_dimensions["A"].width = 28
            ws2.column_dimensions["B"].width = 18
            ws2.column_dimensions["C"].width = 18

        out = BytesIO()
        wb.save(out)
        return out.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    except ImportError:
        # openpyxl not installed — produce a UTF-8 CSV instead
        import csv

        sio = StringIO()
        w = csv.writer(sio)
        w.writerow(["Duka AI — Financial Health Report"])
        w.writerow([full.get("business_name",""), full.get("location",""), full.get("generated_at","")])
        w.writerow([])
        w.writerow(["EXECUTIVE SUMMARY"])
        w.writerow([full.get("executive_summary","")])
        w.writerow([])
        w.writerow(["FINANCIAL SNAPSHOT", ""])
        _csv_profit = snap.get("profit", 0)
        _csv_p_label = "Net Loss" if _csv_profit < 0 else "Profit"
        w.writerow(["Revenue",     f"K{snap.get('revenue',0):,.0f}"])
        w.writerow(["Expenses",    f"K{snap.get('expenses',0):,.0f}"])
        w.writerow([_csv_p_label,  f"K{abs(_csv_profit):,.0f}"])
        w.writerow(["Health Score",f"{snap.get('health_score',0)}/100"])
        w.writerow([])
        w.writerow(["LOAN READINESS", ""])
        w.writerow(["Score",         f"{loan.get('score',0)}/100"])
        w.writerow(["Safe Borrowing",f"K{loan.get('safe_amount',0):,.0f}"])
        w.writerow([])
        w.writerow(["RECOMMENDATIONS", "Impact", "Timeline"])
        for r in full.get("recommendations",[]):
            w.writerow([r.get("action",""), r.get("impact",""), r.get("timeline","")])
        w.writerow([])
        w.writerow(["RISKS TO WATCH"])
        for risk in full.get("risks",[]):
            w.writerow([risk])
        return sio.getvalue().encode("utf-8"), "text/csv"


def render_generate_report_page() -> None:
    render_page_header(
        "Generate Report",
        "AI-compiled financial health report — ready to share with your bank or accountant.",
        "Report",
    )
    report = require_report()
    if not report:
        return

    baseline = st.session_state.get("baseline") or get_baseline(st.session_state)
    if not baseline:
        st.warning("No financial baseline found. Run an analysis in Chat Advisor first.")
        return

    metrics = st.session_state.get("metrics") or compute_metrics(baseline)
    bp = {
        "business_type":     st.session_state.get("business_type", ""),
        "location":          st.session_state.get("location", ""),
        "products_services": st.session_state.get("products_services", ""),
    }

    # ── Trigger / cache logic ────────────────────────────────────────────────
    full: dict | None = st.session_state.get("generated_report")

    if full is None:
        if not st.session_state.get("_report_generating"):
            st.info(
                "Click below to compile a full financial health report from your analysis data. "
                "The AI will write the narrative; all numbers come from verified calculations."
            )
            if st.button("⚡ Generate Full Report", type="primary", use_container_width=True):
                st.session_state["_report_generating"] = True
                st.rerun()
            return

        # Generating — phased progress (matches agents/report_agent.generate_full_report)
        from agents.report_agent import generate_full_report

        with st.status("📄 Starting financial report pipeline…", expanded=True) as status:
            st.markdown(
                "**Typical duration:** about **15–60 seconds**, depending on API latency. "
                "Every Kwacha amount is taken from Python-verified calculations first; "
                "the AI only writes narrative text grounded in those numbers."
            )
            st.caption(
                "Pipeline: verified snapshot → expense categories → 6-month base-case forecast "
                "→ AI executive summary → AI recommendations & risks → merge for PDF/Excel export."
            )
            prog = st.progress(0)

            def _on_report_progress(step: int, total: int, detail: str) -> None:
                pct = min(step / max(total, 1), 1.0)
                try:
                    prog.progress(pct, text=f"Step {step}/{total}")
                except TypeError:
                    prog.progress(pct)
                short = detail if len(detail) <= 80 else detail[:77] + "…"
                status.update(label=f"📄 {short}", state="running")
                st.markdown(f"##### Step {step} of {total}")
                st.markdown(detail)
                st.divider()

            try:
                full = generate_full_report(
                    baseline, metrics, report, bp, on_progress=_on_report_progress
                )
            except Exception as exc:
                st.session_state["_report_generating"] = False
                try:
                    prog.progress(0, text="Stopped")
                except TypeError:
                    prog.progress(0)
                status.update(
                    label="❌ Report generation failed — see details below",
                    state="error",
                    expanded=True,
                )
                st.exception(exc)
                return

            try:
                prog.progress(1.0, text="Done")
            except TypeError:
                prog.progress(1.0)
            status.update(
                label="✅ Report compiled — scroll down for preview and downloads.",
                state="complete",
                expanded=False,
            )

        st.session_state["generated_report"] = full
        st.session_state["_report_generating"] = False
        st.rerun()

    # ── Action bar (top) ─────────────────────────────────────────────────────
    snap   = full.get("snapshot", {})
    loan_d = full.get("loan", {})
    fc6    = full.get("forecast_6m", {})

    pdf_bytes             = build_full_report_pdf(full)
    excel_bytes, xls_mime = build_excel_report_bytes(full)
    xls_ext  = "xlsx" if "spreadsheet" in xls_mime else "csv"
    xls_label = "📊 Download Excel" if xls_ext == "xlsx" else "📊 Download CSV"

    col_pdf, col_xls, col_regen = st.columns([3, 3, 2], gap="small")
    col_pdf.download_button(
        "📄 Download PDF", pdf_bytes, "duka-ai-report.pdf", "application/pdf",
        use_container_width=True, type="primary", key="rpt_dl_pdf_top",
    )
    col_xls.download_button(
        xls_label, excel_bytes, f"duka-ai-report.{xls_ext}", xls_mime,
        use_container_width=True, key="rpt_dl_xls_top",
    )
    if col_regen.button("🔄 Regenerate", use_container_width=True):
        st.session_state.pop("generated_report", None)
        st.rerun()

    st.markdown("<div style='height:12px;'></div>", unsafe_allow_html=True)

    _rn = _report_text_escape(full.get("business_name", "Your Business"))
    _rloc = _report_text_escape(full.get("location", "Zambia"))
    _rat = _report_text_escape(full.get("generated_at", ""))
    _snap_pr = float(snap.get("profit", 0) or 0)
    _kpi_pc = "loss" if _snap_pr < 0 else "gain"

    st.markdown(
        '<div id="duka-ai-report-page-marker" aria-hidden="true"></div>'
        '<div class="duka-report-premium">',
        unsafe_allow_html=True,
    )

    # ── Hero + KPI ribbon ────────────────────────────────────────────────────
    st.markdown(
        f"""
        <div class="duka-report-hero">
          <div class="duka-report-hero-kicker">Financial Health Report</div>
          <div class="duka-report-hero-title">{_rn}</div>
          <div class="duka-report-hero-meta">{_rloc} &nbsp;·&nbsp; Generated {_rat}</div>
          <span class="duka-report-hero-badge">Verified numbers · AI narrative</span>
        </div>
        <div class="duka-report-kpi-row">
          <div class="duka-report-kpi">
            <div class="duka-report-kpi-lab">Monthly revenue</div>
            <div class="duka-report-kpi-val">K{snap.get("revenue", 0):,.0f}</div>
          </div>
          <div class="duka-report-kpi">
            <div class="duka-report-kpi-lab">Monthly expenses</div>
            <div class="duka-report-kpi-val">K{snap.get("expenses", 0):,.0f}</div>
          </div>
          <div class="duka-report-kpi">
            <div class="duka-report-kpi-lab">Net profit</div>
            <div class="duka-report-kpi-val {_kpi_pc}">K{abs(_snap_pr):,.0f}</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        f"""
        <div class="duka-report-exec">
          <div class="duka-report-section-h">Executive summary</div>
          <div class="duka-report-exec-body">{_report_text_escape(full.get("executive_summary", "—"))}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Financial Snapshot + Expense Breakdown ───────────────────────────────
    def _score_color(score: int) -> str:
        return "#10B981" if score >= 70 else ("#F59E0B" if score >= 40 else "#EF4444")

    h_color = _score_color(snap.get("health_score", 0))
    l_color = _score_color(loan_d.get("score", 0))

    col_snap, col_exp = st.columns([1, 1], gap="medium")

    with col_snap:
        _cfs = _report_text_escape(snap.get("cash_flow_status", "—"))
        _hl = _report_text_escape(snap.get("health_label", "N/A"))
        _src = _report_text_escape(snap.get("source", "—"))
        _pr_disp = float(snap.get("profit", 0) or 0)
        _pr_col = "#FCA5A5" if _pr_disp < 0 else "#6EE7B7"
        st.markdown(
            f"""
            <div class="duka-report-card" style="height:100%;">
              <div class="duka-report-section-h">Financial snapshot</div>
              <table style="width:100%;border-collapse:collapse;font-size:13px;color:#E2E8F0;">
                <tr><td style="padding:7px 0;color:#94A3B8;">Revenue</td>
                    <td style="text-align:right;font-weight:700;">K{snap.get("revenue",0):,.0f}</td></tr>
                <tr><td style="padding:7px 0;color:#94A3B8;">Expenses</td>
                    <td style="text-align:right;font-weight:700;color:#F87171;">
                      K{snap.get("expenses",0):,.0f}
                      <span style="font-size:11px;color:#94A3B8;"> ({snap.get("expense_pct",0):.0f}% of rev.)</span>
                    </td></tr>
                <tr><td style="padding:7px 0;color:#94A3B8;">Profit</td>
                    <td style="text-align:right;font-weight:700;color:{_pr_col};">
                      K{_pr_disp:,.0f}
                      <span style="font-size:11px;color:#94A3B8;"> ({snap.get("margin_pct",0):.1f}% margin)</span>
                    </td></tr>
                <tr><td colspan="2" style="border-top:1px solid rgba(148,163,184,0.12);padding-top:10px;"></td></tr>
                <tr><td style="padding:7px 0;color:#94A3B8;">Health score</td>
                    <td style="text-align:right;font-weight:800;color:{h_color};">
                      {snap.get("health_score",0)}/100 — {_hl}
                    </td></tr>
                <tr><td style="padding:7px 0;color:#94A3B8;">Cash flow</td>
                    <td style="text-align:right;font-weight:600;">{_cfs}</td></tr>
                <tr><td style="padding:7px 0;color:#94A3B8;">Data source</td>
                    <td style="text-align:right;font-size:12px;color:#CBD5E1;">{_src}</td></tr>
              </table>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_exp:
        cats = full.get("expense_categories", [])
        if cats:
            rows_html = "".join(
                f"""<tr>
                  <td style="padding:6px 0;color:#CBD5E1;">{i + 1}. {_report_text_escape(c.get("name", ""))}</td>
                  <td style="text-align:right;color:#F87171;font-weight:600;">K{c.get("amount", 0):,.0f}</td>
                  <td style="text-align:right;color:#94A3B8;font-size:11px;">{c.get("pct_of_expenses", 0):.1f}%</td>
                </tr>"""
                for i, c in enumerate(cats[:6])
            )
            st.markdown(
                f"""
                <div class="duka-report-card" style="height:100%;">
                  <div class="duka-report-section-h">Expense breakdown</div>
                  <table style="width:100%;border-collapse:collapse;font-size:13px;color:#E2E8F0;">
                    {rows_html}
                  </table>
                </div>
                """,
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div class="duka-report-card" style="color:#94A3B8;font-size:13px;line-height:1.55;">'
                "No expense category breakdown available for this analysis."
                "</div>",
                unsafe_allow_html=True,
            )

    st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)

    # ── Cash Flow + Loan side by side ────────────────────────────────────────
    col_cf, col_loan = st.columns([1, 1], gap="medium")

    with col_cf:
        cf_sect = full.get("cash_flow", {})
        _cf_sum = _report_text_escape(cf_sect.get("summary", "—"))
        _cf_ins = cf_sect.get("insight") or cf_sect.get("reasoning") or ""
        _cf_ins_html = (
            f"<div style='margin-top:12px;font-size:12px;color:#94A3B8;line-height:1.55;font-style:italic;border-top:1px solid rgba(148,163,184,0.12);padding-top:10px;'>{_report_text_escape(_cf_ins)}</div>"
            if _cf_ins
            else ""
        )
        st.markdown(
            f"""
            <div class="duka-report-card">
              <div class="duka-report-section-h">Cash flow analysis</div>
              <div style="font-size:14px;color:#E2E8F0;line-height:1.68;font-weight:450;">
                {_cf_sum}
              </div>
              {_cf_ins_html}
            </div>
            """,
            unsafe_allow_html=True,
        )

    with col_loan:
        improve_html = "".join(
            f'<li style="color:#CBD5E1;font-size:12px;margin-bottom:5px;">{_report_text_escape(item)}</li>'
            for item in loan_d.get("improve", [])
        )
        _lst = _report_text_escape(loan_d.get("status", "—"))
        _lfr = _report_text_escape(loan_d.get("formula", ""))
        st.markdown(
            f"""
            <div class="duka-report-card">
              <div class="duka-report-section-h">Loan readiness</div>
              <div style="font-size:22px;font-weight:850;color:{l_color};margin-bottom:6px;letter-spacing:-0.02em;">
                {loan_d.get("score", 0)}/100 — {_lst}
              </div>
              <div style="font-size:13px;color:#E2E8F0;margin-bottom:8px;line-height:1.55;">
                Safe borrowing: <strong>K{loan_d.get("safe_amount", 0):,.0f}</strong>
                <span style="font-size:11px;color:#94A3B8;"> · {_lfr}</span>
              </div>
              {"<ul style='padding-left:18px;margin-top:8px;list-style:disc;'>" + improve_html + "</ul>" if improve_html else ""}
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # ── 6-Month outlook ────────────────────────────────────────────────────────
    _fc_out = (
        "All 6 months profitable (base case)."
        if fc6.get("all_profitable")
        else f"Attention: losses possible from {fc6.get('first_loss_month', '—')} upward."
    )
    st.markdown(
        f"""
        <div class="duka-report-card" style="margin-bottom:18px;">
          <div class="duka-report-section-h">6-month outlook (base case)</div>
          <div class="duka-report-fc-grid">
            <div class="duka-report-fc-stat">
              <div class="duka-report-fc-lab">Month 6 profit</div>
              <div class="duka-report-fc-val">K{fc6.get("base_profit_m6", 0):,.0f}</div>
            </div>
            <div class="duka-report-fc-stat">
              <div class="duka-report-fc-lab">Total 6-mo profit</div>
              <div class="duka-report-fc-val">K{fc6.get("total_profit", 0):,.0f}</div>
            </div>
            <div class="duka-report-fc-stat">
              <div class="duka-report-fc-lab">Profitable months</div>
              <div class="duka-report-fc-val">{fc6.get("profitable_months", 0)}/6</div>
            </div>
            <div class="duka-report-fc-stat">
              <div class="duka-report-fc-lab">Outlook</div>
              <div class="duka-report-fc-val" style="font-size:0.95rem;line-height:1.35;color:#CBD5E1;">
                {_report_text_escape(_fc_out)}
              </div>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

    # ── Market Conditions ────────────────────────────────────────────────────
    mkt = full.get("market", {})
    if mkt.get("summary"):
        monitor_html = "".join(
            f'<span style="background:rgba(37,99,235,0.18);border:1px solid rgba(37,99,235,0.35);'
            f'border-radius:999px;padding:3px 11px;font-size:11px;color:#93C5FD;margin:4px 6px 0 0;display:inline-block;">'
            f'{_report_text_escape(m)}</span>'
            for m in mkt.get("monitor", [])
        )
        _opp = mkt.get("opportunity", "")
        _opp_blk = (
            f"<div style='font-size:13px;color:#6EE7B7;margin-top:10px;line-height:1.55;'><strong>Opportunity:</strong> {_report_text_escape(_opp)}</div>"
            if _opp
            else ""
        )
        st.markdown(
            f"""
            <div class="duka-report-card" style="margin-bottom:18px;">
              <div class="duka-report-section-h">Market conditions</div>
              <div style="font-size:14px;color:#E2E8F0;line-height:1.68;margin-bottom:6px;">
                {_report_text_escape(mkt.get("summary", "—"))}
              </div>
              {_opp_blk}
              {("<div style='margin-top:12px;'>" + monitor_html + "</div>") if monitor_html else ""}
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Recommendations ──────────────────────────────────────────────────────
    recs = full.get("recommendations", [])
    if recs:
        TIMELINE_COLORS = {"Now": "#10B981", "This month": "#2563EB"}
        rec_cards = ""
        for i, r in enumerate(recs):
            t_color = TIMELINE_COLORS.get(r.get("timeline", ""), "#64748B")
            _act = _report_text_escape(r.get("action", "—"))
            _imp = _report_text_escape(r.get("impact", "—"))
            _tl = _report_text_escape(r.get("timeline", "—"))
            rec_cards += f"""
            <div class="duka-report-rec-item">
              <div class="duka-report-rec-num">{i + 1}</div>
              <div style="flex:1;min-width:0;">
                <div class="duka-report-rec-action">{_act}</div>
                <div class="duka-report-rec-meta">
                  <span class="duka-report-rec-impact">→ {_imp}</span>
                  <span class="duka-report-rec-pill" style="color:{t_color};border-color:{t_color};">{_tl}</span>
                </div>
              </div>
            </div>
            """
        st.markdown(
            f"""
            <div class="duka-report-card" style="margin-bottom:18px;">
              <div class="duka-report-section-h">Top recommendations</div>
              <div class="duka-report-rec-list">{rec_cards}</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Risks to Watch ───────────────────────────────────────────────────────
    risks = full.get("risks", [])
    if risks:
        risk_html = "".join(
            f'<div class="duka-report-risk-row">'
            f'<span class="duka-report-risk-ic" aria-hidden="true">⚠</span>'
            f'<span class="duka-report-risk-txt">{_report_text_escape(r)}</span></div>'
            for r in risks
        )
        st.markdown(
            f"""
            <div class="duka-report-risk-card" style="margin-bottom:18px;">
              <div class="duka-report-section-h" style="color:#FBBF24;">Risks to watch</div>
              {risk_html}
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── Footer / branding ────────────────────────────────────────────────────
    st.markdown(
        '<div class="duka-report-foot">'
        "Generated by Duka AI &nbsp;·&nbsp; Powered by AMD MI300X &nbsp;·&nbsp; Qwen Model"
        "</div>"
        "</div>",
        unsafe_allow_html=True,
    )

    # ── Share section ────────────────────────────────────────────────────────
    st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)
    with st.expander("📤 Share this report", expanded=False):
        st.markdown(
            '<div style="font-size:13px;color:#94A3B8;margin-bottom:12px;">'
            'Share this report with your bank manager or accountant via WhatsApp or email. '
            'Download the PDF above and attach it to your message.'
            '</div>',
            unsafe_allow_html=True,
        )

        # Plain-text copy of the report
        report_text = (
            f"Duka AI — Financial Health Report\n"
            f"{full.get('business_name','')} | {full.get('location','')} | {full.get('generated_at','')}\n"
            f"{'='*60}\n\n"
            f"EXECUTIVE SUMMARY\n{full.get('executive_summary','')}\n\n"
            f"FINANCIAL SNAPSHOT\n"
            f"Revenue:      K{snap.get('revenue',0):,.0f}\n"
            f"Expenses:     K{snap.get('expenses',0):,.0f} ({snap.get('expense_pct',0):.0f}%)\n"
            f"{'Net Loss' if snap.get('profit',0) < 0 else 'Profit'}:       K{abs(snap.get('profit',0)):,.0f} ({snap.get('margin_pct',0):.1f}% margin)\n"
            f"Health Score: {snap.get('health_score',0)}/100 — {snap.get('health_label','')}\n\n"
            f"LOAN READINESS\n"
            f"Score:          {loan_d.get('score',0)}/100 — {loan_d.get('status','')}\n"
            f"Safe borrowing: K{loan_d.get('safe_amount',0):,.0f}\n\n"
            + "TOP RECOMMENDATIONS\n"
            + "\n".join(
                f"{i+1}. {r.get('action','')} → {r.get('impact','')} ({r.get('timeline','')})"
                for i, r in enumerate(recs)
            ) + "\n\n"
            + ("RISKS TO WATCH\n" + "\n".join(f"⚠  {r}" for r in risks) + "\n\n" if risks else "")
            + f"{'='*60}\nGenerated by Duka AI | Powered by AMD MI300X | Qwen Model"
        )

        st.text_area("Copy report text", report_text, height=200, key="report_share_text")

        sh_pdf, sh_xls = st.columns(2)
        sh_pdf.download_button(
            "📄 Download PDF",
            pdf_bytes,
            "duka-ai-report.pdf",
            "application/pdf",
            use_container_width=True,
            key="rpt_dl_pdf_share",
        )
        sh_xls.download_button(
            xls_label,
            excel_bytes,
            f"duka-ai-report.{xls_ext}",
            xls_mime,
            use_container_width=True,
            key="rpt_dl_xls_share",
        )

    # ── Schedule automatic reports ────────────────────────────────────────────
    st.markdown("<div style='height:20px;'></div>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="background: linear-gradient(135deg, rgba(29,158,117,0.12) 0%,
                    rgba(37,99,235,0.08) 100%); border: 1px solid rgba(29,158,117,0.35);
                    border-radius: 12px; padding: 20px 24px; margin-bottom: 16px;">
          <div style="font-size: 18px; font-weight: 800; color: #E2E8F0; margin-bottom: 6px;">
            &#128236; Want this report automatically?
          </div>
          <div style="font-size: 13px; color: #94A3B8; line-height: 1.65;">
            I can analyze your business and email you a complete financial report on a
            schedule you choose &mdash; no action needed from you.
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        "<div style='font-size:12px;font-weight:700;text-transform:uppercase;"
        "letter-spacing:.1em;color:#94A3B8;margin-bottom:10px;'>"
        "Set up automatic reports:</div>",
        unsafe_allow_html=True,
    )

    sched_freq = st.session_state.get("schedule_frequency")
    btn_w, btn_m, btn_c = st.columns(3, gap="small")

    if btn_w.button(
        "📅 Weekly — every Monday",
        use_container_width=True,
        type="primary" if sched_freq == "weekly" else "secondary",
        key="sched_weekly",
    ):
        st.session_state["schedule_frequency"] = "weekly"
        st.session_state.pop("schedule_confirmed", None)
        st.rerun()

    if btn_m.button(
        "📅 Monthly — 1st of month",
        use_container_width=True,
        type="primary" if sched_freq == "monthly" else "secondary",
        key="sched_monthly",
    ):
        st.session_state["schedule_frequency"] = "monthly"
        st.session_state.pop("schedule_confirmed", None)
        st.rerun()

    if btn_c.button(
        "✏️ Custom schedule",
        use_container_width=True,
        type="primary" if sched_freq == "custom" else "secondary",
        key="sched_custom",
    ):
        st.session_state["schedule_frequency"] = "custom"
        st.session_state.pop("schedule_confirmed", None)
        st.rerun()

    if sched_freq and not st.session_state.get("schedule_confirmed"):
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

        if sched_freq == "custom":
            _custom_map = {
                "Every 2 weeks (bi-weekly)": "biweekly",
                "Quarterly (every 3 months)": "quarterly",
            }
            _chosen = st.selectbox(
                "Choose your schedule:",
                list(_custom_map.keys()),
                key="sched_custom_choice",
            )
            effective_freq = _custom_map[_chosen]
        else:
            effective_freq = sched_freq

        sched_email = st.text_input(
            "Your email address",
            placeholder="you@example.com",
            key="sched_email_input",
        )

        if st.button(
            "Set up automatic reports ✅",
            type="primary",
            use_container_width=True,
            key="sched_confirm",
        ):
            import re as _re
            if not sched_email or not _re.match(r"[^@]+@[^@]+\.[^@]+", sched_email.strip()):
                st.error("Please enter a valid email address.")
            else:
                _biz_ctx = {
                    "business_type":     st.session_state.get("business_type", ""),
                    "location":          st.session_state.get("location", "Zambia"),
                    "products_services": st.session_state.get("products_services", ""),
                    "manual_notes":      st.session_state.get("manual_notes", ""),
                    "manual_revenue":    st.session_state.get("manual_revenue", 0.0),
                    "manual_expenses":   st.session_state.get("manual_expenses", 0.0),
                    "manual_debt":       st.session_state.get("manual_debt", 0.0),
                    "manual_staff":      st.session_state.get("manual_staff", 0),
                }
                try:
                    from tools.scheduler import save_schedule
                    _sid = save_schedule(
                        email=sched_email.strip(),
                        frequency=effective_freq,
                        business_context=_biz_ctx,
                    )
                    st.session_state["schedule_confirmed"] = True
                    st.session_state["schedule_email"] = sched_email.strip()
                    st.session_state["schedule_id"] = _sid
                    st.rerun()
                except Exception as _exc:
                    st.error(f"Could not save schedule: {_exc}")

    if st.session_state.get("schedule_confirmed"):
        _conf_email = st.session_state.get("schedule_email", "")
        _conf_freq = st.session_state.get("schedule_frequency", "weekly")
        _freq_display = {
            "weekly":    "every Monday at 8 AM",
            "monthly":   "on the 1st of each month at 8 AM",
            "biweekly":  "every 2 weeks at 8 AM",
            "quarterly": "every quarter at 8 AM",
            "custom":    "on your chosen schedule",
        }.get(_conf_freq, "on your chosen schedule")
        st.success(
            f"Scheduled! Your financial report will be emailed to **{_conf_email}** {_freq_display}."
        )
        if st.button("Cancel / change schedule", key="sched_cancel"):
            for _k in ["schedule_confirmed", "schedule_frequency", "schedule_email", "schedule_id"]:
                st.session_state.pop(_k, None)
            st.rerun()


def _market_intel_identity_reply() -> str:
    return (
        "I'm **Duka AI's Market Intelligence Agent** — part of **Duka AI**, built for **Zambian SMEs**. "
        "I connect **your business profile and numbers** from your last analysis to **market conditions** "
        "(this page’s snapshot and, when enabled, web sources from when your report was built). "
        "I'm not a general chatbot for every topic — I stay focused on **your market and your financial picture**.\n\n"
        "Try: *What's the biggest risk for my type of business right now?* or "
        "*How does my revenue pattern line up with what the market intel says?*"
    )


def _build_market_chat_business_context(report: dict[str, Any]) -> str:
    """Ground Market Intel chat in the same numbers as the rest of the report."""
    lines: list[str] = []
    cf = report.get("cashflow") or {}
    lines.append("VERIFIED BUSINESS NUMBERS (from last analysis — cite these for sales/profit/trend questions):\n")
    lines.append(
        f"- Revenue K{float(cf.get('revenue', 0)):,.0f} · Expenses K{float(cf.get('expenses', 0)):,.0f} · "
        f"Profit K{float(cf.get('profit', 0)):,.0f} · Margin {float(cf.get('profit_margin', 0)):.1f}% · "
        f"Status: {cf.get('cash_flow_status', 'n/a')}\n"
    )
    if (cf.get("summary") or "").strip():
        lines.append(f"- Summary: {(cf.get('summary') or '')[:500]}\n")

    parsed = report.get("parsed_data") or {}
    ac = parsed.get("_analysis_context") or {}
    doc = report.get("document_analysis") or ac.get("document_analysis") or {}
    bp = report.get("business_profile") or ac.get("business_profile") or {}

    bt = (parsed.get("business_type") or bp.get("business_type") or "").strip()
    loc = (parsed.get("location") or bp.get("location") or "").strip()
    products = (parsed.get("products_services") or bp.get("products_services") or "").strip()
    if bt or loc:
        lines.append(f"- Business: {bt or 'SME'} · Location: {loc or 'Zambia'}\n")
    if products:
        lines.append(f"- Products/services listed by owner: {products[:600]}\n")

    mb = doc.get("monthly_breakdown") if isinstance(doc.get("monthly_breakdown"), list) else []
    if len(mb) > 1:
        lines.append(
            "MONTHLY REVENUE PATTERN (use this for “trending”, seasonality, up/down — not individual SKUs):\n"
        )
        for entry in mb[:14]:
            if not isinstance(entry, dict):
                continue
            mn = entry.get("month") or entry.get("Month") or ""
            rev = entry.get("revenue", entry.get("Revenue"))
            if rev is not None:
                try:
                    lines.append(f"  · {mn}: revenue K{float(rev):,.0f}\n")
                except (TypeError, ValueError):
                    pass

    eb = parsed.get("expenses_breakdown") or {}
    if isinstance(eb, dict) and eb:
        top = sorted(
            ((k, float(v or 0)) for k, v in eb.items() if float(v or 0) > 0),
            key=lambda x: x[1],
            reverse=True,
        )[:5]
        if top:
            lines.append(
                "TOP EXPENSE CATEGORIES (if user asks where money goes):\n"
                + "".join(f"  · {k}: K{v:,.0f}\n" for k, v in top)
            )

    ts = report.get("transaction_summary") or {}
    if ts.get("average_daily_sales"):
        lines.append(
            f"- Transactions (if connected): avg daily sales ~{format_currency(ts['average_daily_sales'])}, "
            f"risk {ts.get('cash_flow_risk', 'n/a')}.\n"
        )

    lines.append(
        "\nSCOPE NOTE: There is usually **no per-product SKU sales feed**. "
        "For “are my products trending”, use **monthly revenue trend** above + owner’s product list + market intel. "
        "If monthly data is missing, say revenue is only known as totals above.\n"
    )
    return "".join(lines)[:7500]


def _classify_market_chat_input(question: str) -> str | None:
    """Return a canned reply key, or None to use the LLM."""
    low = (question or "").lower().strip()
    if not low:
        return "empty"
    if any(
        p in low
        for p in (
            "who are you",
            "what are you",
            "what do you do",
            "introduce yourself",
        )
    ):
        return "identity"
    if low in ("thanks", "thank you", "ty", "cheers", "ta"):
        return "thanks"
    return None


def _handle_market_question(
    question: str,
    market: dict,
    chat_history: list[dict],
    report: dict[str, Any] | None = None,
) -> str:
    """Answer using the market block + verified business numbers from the same report."""
    from agents import request_llm_chat

    canned = _classify_market_chat_input(question)
    if canned == "identity":
        return _market_intel_identity_reply()
    if canned == "thanks":
        return "Glad it helped — ask anything else about your market snapshot or your numbers."
    if canned == "empty":
        return "Type a question about your market or how it relates to your business."

    sources = market.get("sources", [])
    queries = market.get("queries_used", [])
    web_used = market.get("web_search_used", False)

    # Build context from real search data
    source_context = ""
    if web_used and sources:
        source_context = "\n\nWEB SEARCH SOURCES (real, retrieved data — use these):\n"
        seen: set[str] = set()
        for s in sources:
            url = s.get("url", "")
            if url in seen:
                continue
            seen.add(url)
            source_context += f"- {s.get('title', '')} ({url})\n"
            if s.get("content"):
                source_context += f"  {s['content'][:300]}\n"

    market_context = (
        f"MARKET INTELLIGENCE RESULTS:\n"
        f"Business type: {market.get('business_type', 'SME')}\n"
        f"Location: {market.get('location', 'Zambia')}\n"
        f"Market summary: {market.get('market_summary', '')}\n"
        f"Key risks: {'; '.join(market.get('risk_factors', []))}\n"
        f"Opportunity: {market.get('opportunity', '')}\n"
        f"Recommendation: {market.get('recommendation', '')}\n"
        f"What to monitor: {'; '.join(market.get('monitor_next', []))}\n"
        f"Search queries used: {'; '.join(queries)}\n"
        f"{source_context}"
    )

    biz_block = _build_market_chat_business_context(report) if report else ""

    if web_used:
        grounding = (
            "Live Tavily web search was used when this report was built — snippets and titles below are real. "
            "Answer using the MARKET INTELLIGENCE block + VERIFIED BUSINESS NUMBERS + web excerpts. "
            "Do NOT invent statistics. Cite or paraphrase sources when you mention external facts."
        )
    else:
        grounding = (
            "No live web search ran for this report (missing TAVILY_API_KEY or search failed). "
            "The MARKET INTELLIGENCE RESULTS below are best-effort synthesis from business context — "
            "do NOT claim you ran a fresh web search. "
            "Still use VERIFIED BUSINESS NUMBERS for anything about **their** sales, profit, or month-to-month trend.\n"
            "If something isn't in the data, say exactly what's missing (e.g. no SKU-level sales)."
        )

    system = (
        "You are **Duka AI's Market Intelligence Agent** inside **Duka AI** for **Zambian SMEs**. "
        "You are NOT a generic assistant for “the economy” or random topics — you tie **market conditions** to "
        "**this owner's** business type, location, and **verified numbers**.\n\n"
        "WHO YOU ARE (if asked): You help this user interpret **their** market snapshot (this page) and how it "
        "relates to **their revenue/cash flow** — not broad unrelated advice.\n\n"
        "PRODUCT / “TRENDING” QUESTIONS: You normally do **not** have per-SKU sales. "
        "Answer using **monthly revenue** (if provided), overall totals, their stated products/services, "
        "and the market intel risks/opportunities. Say clearly when you're inferring vs when data proves it.\n\n"
        f"{grounding}\n"
        "Be concise (2–6 sentences unless they ask for detail). "
        "Never refuse to discuss trends if monthly revenue or totals exist — use them.\n\n"
        f"{biz_block}\n"
        f"{market_context}"
    )

    messages = [
        {"role": "system", "content": system},
        *[{"role": m["role"], "content": m["content"]} for m in chat_history[-8:]],
        {"role": "user", "content": question},
    ]
    result = request_llm_chat(messages, temperature=0.2, max_tokens=500)
    return result or "I couldn't generate a response right now. Please try again."


def render_market_intel_page() -> None:
    render_page_header(
        "Market Intel",
        "Live web search for Zambian market conditions, risks, and opportunities.",
        "Live",
    )
    report = require_report()
    if not report:
        return

    market = report.get("market", {})

    st.markdown(
        '<div id="duka-ai-market-intel-page-marker" aria-hidden="true" '
        'style="position:absolute;width:0;height:0;overflow:hidden"></div>',
        unsafe_allow_html=True,
    )

    # Two-column layout: intel card left, chat panel right
    col_intel, col_chat = st.columns([3, 2], gap="large")

    with col_intel:
        # Market data was computed during Chat Advisor analysis — we don't re-search here.
        render_market_intelligence_section(report)

    with col_chat:
        web_used = market.get("web_search_used", False)
        sources = market.get("sources", [])
        if web_used:
            sub_line = (
                f"Grounded in {len(sources)} live web sources · "
                "Ask anything about your market"
            )
        else:
            sub_line = "Ask anything about your business's market conditions in Zambia"

        st.markdown(
            '<div id="duka-ai-market-intel-chat-anchor"></div>',
            unsafe_allow_html=True,
        )

        st.markdown(
            '<div class="duka-forecast-convo-card">'
            '<div class="duka-forecast-convo-head">'
            '<span class="duka-forecast-convo-title">🌍 Ask the Market Agent</span>'
            '<span class="duka-forecast-convo-sub">'
            f"{html.escape(sub_line)}"
            "</span>"
            "</div>",
            unsafe_allow_html=True,
        )

        # Suggestion pills
        sugg_pills = [
            "📈 What's the market trend?",
            "⚠️ What are the main risks?",
            "🎯 What's my best opportunity?",
            "💰 How do prices affect me?",
            "🗓️ What's the best season to grow?",
            "🏆 How do I beat competition?",
        ]
        p1, p2 = st.columns(2)
        for idx, pill in enumerate(sugg_pills):
            col = p1 if idx % 2 == 0 else p2
            if col.button(pill, key=f"mpill_{idx}", use_container_width=True):
                st.session_state["market_question"] = pill

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # Chat history (same card shell as Cash Flow Forecast conversation)
        if "market_chat" not in st.session_state:
            st.session_state["market_chat"] = []

        if not st.session_state["market_chat"]:
            st.markdown(
                '<div class="duka-forecast-empty">No messages yet — tap a suggestion '
                "or type below.</div>",
                unsafe_allow_html=True,
            )

        for msg in st.session_state["market_chat"]:
            if msg["role"] == "user":
                escaped = (
                    msg["content"]
                    .replace("&", "&amp;")
                    .replace("<", "&lt;")
                    .replace(">", "&gt;")
                )
                st.markdown(
                    f'<div class="user-bubble-wrap"><div class="user-bubble">{escaped}</div></div>',
                    unsafe_allow_html=True,
                )
            else:
                with st.chat_message("assistant", avatar="🌍"):
                    st.markdown(msg["content"])

        st.markdown("</div>", unsafe_allow_html=True)

        user_q = st.chat_input("Ask about your market...", key="market_chat_input")
        if st.session_state.get("market_question"):
            user_q = st.session_state.pop("market_question")

        if user_q:
            st.session_state["market_chat"].append({"role": "user", "content": user_q})
            with st.spinner("🌍 Searching market intelligence…"):
                reply = _handle_market_question(
                    user_q,
                    market,
                    st.session_state["market_chat"],
                    report,
                )
            st.session_state["market_chat"].append({"role": "assistant", "content": reply})
            st.session_state["market_pin_to_question"] = True
            st.rerun()

        # ChatGPT-style scroll: pin the user's last question to top after a reply
        if st.session_state.pop("market_pin_to_question", False):
            inject_pin_to_question_scroll()


def render_settings_page(settings: Any) -> None:
    render_page_header(
        "Settings",
        "Control the current session, provider preference, and reset behavior.",
    )
    ai_status = "✅ AI analysis active" if settings.llm_enabled else "⚠️ AI unavailable — check API key"
    st.caption(ai_status)
    if settings.llm_provider == "amd":
        active_model = settings.amd_model
        provider_label = "AMD MI300X vLLM"
    elif settings.llm_provider == "openai":
        active_model = settings.model_name
        provider_label = "OpenAI-compatible API"
    else:
        active_model = settings.amd_model or settings.model_name
        provider_label = settings.llm_provider
    st.text_input("Provider", value=provider_label, disabled=True)
    st.text_input("Model", value=active_model, disabled=True)
    if st.button("Clear session", type="primary"):
        for key in ["analysis_report", "messages", "manual_notes", "selected_example_prompt", "selected_example_prompt_text"]:
            st.session_state[key] = [] if key == "messages" else None if key == "analysis_report" else ""
        st.session_state.active_page = "Chat Advisor"
        st.rerun()


def render_selected_page(page: str, settings: Any) -> None:
    if page == "Dashboard":
        render_dashboard_page()
    elif page == "Cash Flow Forecast":
        render_cash_flow_forecast_page()
    elif page == "Scenario Planner":
        render_scenario_planner_page()
    elif page == "Loan Calculator":
        render_loan_calculator_page()
    elif page == "Expense Analyzer":
        render_expense_analyzer_page()
    elif page == "Generate Report":
        render_generate_report_page()
    elif page == "Market Intel":
        render_market_intel_page()
    elif page == "Settings":
        render_settings_page(settings)


def apply_custom_styles() -> None:
    st.markdown(
        f"""
        <style>
            html, body, .stApp,
            [data-testid="stAppViewContainer"],
            [data-testid="stAppViewContainer"] > .main,
            [data-testid="stMain"],
            [data-testid="stMainBlockContainer"],
            [data-testid="stHeader"] {{
                background: #212121 !important;
                background-color: #212121 !important;
                background-image: none !important;
            }}
            [data-testid="stHeader"] {{
                box-shadow: none !important;
            }}
            .block-container {{
                max-width: 1280px;
                padding-top: 0.75rem;
                padding-bottom: 6rem;  /* room above sticky composer */
                padding-left: 2rem;
                padding-right: 2rem;
            }}
            /* ChatGPT-style narrow conversation column once analysis is live */
            .chatgpt-stage .block-container,
            body:has(#duka-ai-chat-anchor) .block-container {{
                max-width: 920px;
                padding-left: 1rem;
                padding-right: 1rem;
            }}
            /* Cash Flow Forecast: use a comfortable wide canvas */
            body:has(#duka-ai-forecast-page-marker) .block-container {{
                max-width: min(1480px, 94vw) !important;
                padding-left: 1.35rem !important;
                padding-right: 1.35rem !important;
                padding-bottom: 6rem !important;
            }}
            body:has(#duka-ai-market-intel-page-marker) .block-container {{
                max-width: min(1480px, 94vw) !important;
                padding-left: 1.35rem !important;
                padding-right: 1.35rem !important;
                padding-bottom: 6rem !important;
            }}
            body:has(#duka-ai-expense-analyzer-page-marker) .block-container {{
                max-width: min(1480px, 94vw) !important;
                padding-left: 1.35rem !important;
                padding-right: 1.35rem !important;
                padding-bottom: 6rem !important;
            }}
            .duka-forecast-dashboard {{
                margin-bottom: 0.4rem;
            }}
            /* Generate Report — premium document layout */
            body:has(#duka-ai-report-page-marker) .block-container {{
                max-width: min(920px, 96vw) !important;
                padding-left: 1.25rem !important;
                padding-right: 1.25rem !important;
            }}
            .duka-report-premium {{
                position: relative;
            }}
            .duka-report-hero {{
                position: relative;
                overflow: hidden;
                background: linear-gradient(152deg, #080d18 0%, #101827 45%, #0b1324 100%);
                border: 1px solid rgba(148, 163, 184, 0.22);
                border-radius: 18px;
                padding: 26px 32px 22px;
                margin-bottom: 22px;
                box-shadow:
                    0 28px 90px rgba(0, 0, 0, 0.5),
                    0 0 0 1px rgba(255, 255, 255, 0.04) inset;
            }}
            .duka-report-hero::after {{
                content: "";
                position: absolute;
                top: -20%;
                right: -10%;
                width: 52%;
                height: 140%;
                background: radial-gradient(ellipse at 70% 20%, rgba(59, 130, 246, 0.14), transparent 58%);
                pointer-events: none;
            }}
            .duka-report-hero-kicker {{
                font-size: 0.68rem;
                font-weight: 800;
                letter-spacing: 0.16em;
                text-transform: uppercase;
                color: #6EE7B7;
                margin-bottom: 6px;
                position: relative;
                z-index: 1;
            }}
            .duka-report-hero-title {{
                font-size: clamp(1.45rem, 3.5vw, 1.85rem);
                font-weight: 850;
                color: #F8FAFC;
                line-height: 1.15;
                margin-bottom: 6px;
                letter-spacing: -0.02em;
                position: relative;
                z-index: 1;
            }}
            .duka-report-hero-meta {{
                font-size: 0.88rem;
                color: rgba(148, 163, 184, 0.95);
                position: relative;
                z-index: 1;
            }}
            .duka-report-hero-badge {{
                display: inline-block;
                margin-top: 12px;
                padding: 4px 12px;
                border-radius: 999px;
                font-size: 0.65rem;
                font-weight: 800;
                letter-spacing: 0.12em;
                text-transform: uppercase;
                color: #93C5FD;
                background: rgba(37, 99, 235, 0.15);
                border: 1px solid rgba(96, 165, 250, 0.35);
                position: relative;
                z-index: 1;
            }}
            .duka-report-kpi-row {{
                display: grid;
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 12px;
                margin-bottom: 22px;
            }}
            @media (max-width: 700px) {{
                .duka-report-kpi-row {{ grid-template-columns: 1fr; }}
            }}
            .duka-report-kpi {{
                background: linear-gradient(180deg, rgba(30, 41, 59, 0.85) 0%, rgba(15, 23, 42, 0.65) 100%);
                border: 1px solid rgba(148, 163, 184, 0.18);
                border-radius: 14px;
                padding: 14px 16px;
                text-align: center;
                box-shadow: 0 8px 32px rgba(0, 0, 0, 0.25);
            }}
            .duka-report-kpi-lab {{
                font-size: 0.65rem;
                font-weight: 800;
                letter-spacing: 0.11em;
                text-transform: uppercase;
                color: #94A3B8;
                margin-bottom: 4px;
            }}
            .duka-report-kpi-val {{
                font-size: 1.15rem;
                font-weight: 800;
                color: #F1F5F9;
            }}
            .duka-report-kpi-val.loss {{ color: #FCA5A5; }}
            .duka-report-kpi-val.gain {{ color: #6EE7B7; }}
            .duka-report-exec {{
                position: relative;
                background: rgba(15, 23, 42, 0.75);
                border: 1px solid rgba(51, 65, 85, 0.45);
                border-radius: 16px;
                padding: 20px 24px 22px;
                margin-bottom: 22px;
                box-shadow: 0 16px 48px rgba(0, 0, 0, 0.28);
            }}
            .duka-report-exec::before {{
                content: "";
                position: absolute;
                left: 0;
                top: 14px;
                bottom: 14px;
                width: 4px;
                border-radius: 4px;
                background: linear-gradient(180deg, #10B981, #2563EB);
            }}
            .duka-report-section-h {{
                font-size: 0.68rem;
                font-weight: 800;
                letter-spacing: 0.14em;
                text-transform: uppercase;
                color: #94A3B8;
                margin-bottom: 10px;
            }}
            .duka-report-exec-body {{
                font-size: 1.02rem;
                color: #E2E8F0;
                line-height: 1.72;
                font-weight: 450;
                padding-left: 12px;
            }}
            .duka-report-card {{
                background: rgba(15, 23, 42, 0.72);
                border: 1px solid rgba(148, 163, 184, 0.16);
                border-radius: 14px;
                padding: 18px 22px;
                box-shadow: 0 12px 40px rgba(0, 0, 0, 0.22);
            }}
            .duka-report-fc-grid {{
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 14px;
                margin-top: 8px;
            }}
            @media (max-width: 540px) {{
                .duka-report-fc-grid {{ grid-template-columns: 1fr; }}
            }}
            .duka-report-fc-stat {{
                background: rgba(30, 41, 59, 0.5);
                border: 1px solid rgba(148, 163, 184, 0.12);
                border-radius: 12px;
                padding: 12px 14px;
            }}
            .duka-report-fc-lab {{
                font-size: 0.62rem;
                font-weight: 700;
                letter-spacing: 0.08em;
                text-transform: uppercase;
                color: #94A3B8;
            }}
            .duka-report-fc-val {{
                font-size: 1.1rem;
                font-weight: 800;
                color: #F8FAFC;
                margin-top: 4px;
            }}
            .duka-report-foot {{
                text-align: center;
                padding: 18px 12px 8px;
                font-size: 0.72rem;
                letter-spacing: 0.06em;
                text-transform: uppercase;
                color: rgba(148, 163, 184, 0.55);
                border-top: 1px solid rgba(148, 163, 184, 0.12);
                margin-top: 12px;
            }}
            .duka-report-rec-list {{
                margin-top: 4px;
            }}
            .duka-report-rec-item {{
                display: flex;
                align-items: flex-start;
                gap: 14px;
                padding: 14px 0;
                border-bottom: 1px solid rgba(148, 163, 184, 0.1);
            }}
            .duka-report-rec-item:first-child {{ padding-top: 4px; }}
            .duka-report-rec-item:last-child {{
                border-bottom: none;
                padding-bottom: 4px;
            }}
            .duka-report-rec-num {{
                width: 30px;
                height: 30px;
                border-radius: 50%;
                background: linear-gradient(145deg, rgba(37, 99, 235, 0.35), rgba(15, 23, 42, 0.8));
                border: 1px solid rgba(96, 165, 250, 0.45);
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 800;
                color: #93C5FD;
                font-size: 0.82rem;
                flex-shrink: 0;
                box-shadow: 0 4px 14px rgba(37, 99, 235, 0.2);
            }}
            .duka-report-rec-action {{
                font-size: 0.95rem;
                color: #F1F5F9;
                font-weight: 650;
                line-height: 1.45;
                margin-bottom: 6px;
            }}
            .duka-report-rec-meta {{
                font-size: 0.78rem;
                line-height: 1.5;
            }}
            .duka-report-rec-impact {{
                color: #6EE7B7;
            }}
            .duka-report-rec-pill {{
                display: inline-block;
                background: rgba(0, 0, 0, 0.28);
                border: 1px solid currentColor;
                border-radius: 999px;
                padding: 2px 10px;
                font-size: 0.68rem;
                font-weight: 700;
                margin-left: 8px;
            }}
            .duka-report-risk-card {{
                background: linear-gradient(160deg, rgba(245, 158, 11, 0.08) 0%, rgba(15, 23, 42, 0.75) 100%);
                border: 1px solid rgba(245, 158, 11, 0.28);
                border-radius: 14px;
                padding: 18px 22px;
                box-shadow: 0 12px 40px rgba(0, 0, 0, 0.22);
            }}
            .duka-report-risk-row {{
                display: flex;
                gap: 12px;
                align-items: flex-start;
                margin-bottom: 12px;
            }}
            .duka-report-risk-row:last-child {{ margin-bottom: 0; }}
            .duka-report-risk-ic {{
                color: #FBBF24;
                font-size: 1.1rem;
                flex-shrink: 0;
                line-height: 1.4;
            }}
            .duka-report-risk-txt {{
                font-size: 0.9rem;
                color: #FCD34D;
                line-height: 1.6;
            }}
            [data-testid="stSidebar"] {{
                background: #171717;
                border-right: 1px solid rgba(255,255,255,0.06);
            }}
            [data-testid="stSidebar"] [data-testid="stSidebarContent"] {{
                padding: 1rem 1rem 1.2rem;
            }}
            .sidebar-brand {{
                padding: 0.15rem 0 0.85rem;
                margin-bottom: 0.15rem;
            }}
            .sidebar-brand-row {{
                display: flex;
                align-items: center;
                gap: 0.65rem;
            }}
            .sidebar-brand-mark {{
                width: 2.35rem;
                height: 2.35rem;
                min-width: 2.35rem;
                border-radius: 0.55rem;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 1.15rem;
                background: linear-gradient(135deg, #1D9E75, #378ADD);
                box-shadow: 0 4px 14px rgba(29, 158, 117, 0.28);
            }}
            .sidebar-brand-copy {{
                min-width: 0;
            }}
            .sidebar-brand-title {{
                color: #FFFFFF;
                font-size: 1.2rem;
                font-weight: 800;
                letter-spacing: -0.02em;
                line-height: 1.15;
            }}
            .sidebar-brand-subtitle {{
                color: rgba(255,255,255,0.58);
                font-size: 0.84rem;
                font-weight: 600;
                margin-top: 0.18rem;
                letter-spacing: 0.01em;
            }}
            .sidebar-status {{
                display: inline-flex;
                align-items: center;
                color: #A7F3D0;
                background: rgba(16,185,129,0.10);
                border: 1px solid rgba(16,185,129,0.18);
                border-radius: 999px;
                padding: 0.24rem 0.62rem;
                font-size: 0.68rem;
                font-weight: 700;
                margin: 0.4rem 0 1.1rem;
            }}
            .sidebar-section {{
                color: rgba(255,255,255,0.58);
                border-top: 1px solid rgba(255,255,255,0.10);
                margin-top: 0.95rem;
                padding-top: 0.85rem;
                margin-bottom: 0.45rem;
                font-size: 0.72rem;
                font-weight: 800;
                letter-spacing: 0.11em;
            }}
            .sidebar-footer {{
                border-top: 1px solid rgba(255,255,255,0.10);
                color: rgba(255,255,255,0.48);
                font-size: 0.72rem;
                line-height: 1.45;
                margin-top: 1rem;
                padding-top: 0.9rem;
            }}
            [data-testid="stSidebar"] .stButton > button {{
                min-height: 2.25rem !important;
                border-radius: 12px !important;
                justify-content: flex-start !important;
                padding: 0.35rem 0.7rem !important;
                font-size: 0.84rem !important;
                font-weight: 700 !important;
                box-shadow: none !important;
                margin-bottom: 0.2rem;
                transition: opacity 0.15s ease, transform 0.15s ease,
                    border-color 0.18s ease, background 0.18s ease !important;
            }}
            [data-testid="stSidebar"] .stButton > button:active {{
                opacity: 0.9 !important;
                transform: scale(0.988) !important;
            }}
            [data-testid="stSidebar"] .stButton > button[kind="primary"] {{
                background: linear-gradient(135deg, #2563EB 0%, #14B8A6 100%) !important;
                border: 1px solid rgba(255,255,255,0.16) !important;
                color: #FFFFFF !important;
            }}
            [data-testid="stSidebar"] .stButton > button[kind="secondary"] {{
                background: rgba(255,255,255,0.055) !important;
                border: 1px solid rgba(255,255,255,0.11) !important;
                color: rgba(255,255,255,0.86) !important;
            }}
            [data-testid="stSidebar"] .stButton > button[kind="secondary"]:hover {{
                background: rgba(255,255,255,0.10) !important;
                border-color: rgba(20,184,166,0.42) !important;
                color: #FFFFFF !important;
            }}
            .page-header {{
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                gap: 1rem;
                padding: 1rem 0 0.75rem;
                margin-bottom: 0.65rem;
                border-bottom: 1px solid rgba(255,255,255,0.08);
            }}
            .page-title {{
                color: #FFFFFF;
                font-size: 1.65rem;
                line-height: 1.15;
                font-weight: 850;
            }}
            .page-subtitle {{
                color: rgba(226,232,240,0.72);
                font-size: 0.9rem;
                line-height: 1.55;
                margin-top: 0.3rem;
                max-width: 760px;
            }}
            .page-badge {{
                flex: 0 0 auto;
                color: #A7F3D0;
                background: rgba(16,185,129,0.10);
                border: 1px solid rgba(16,185,129,0.22);
                border-radius: 999px;
                padding: 0.35rem 0.7rem;
                font-size: 0.74rem;
                font-weight: 800;
            }}
            .tool-tip, .empty-state {{
                background: rgba(255,255,255,0.055);
                border: 1px solid rgba(255,255,255,0.09);
                border-radius: 16px;
                padding: 0.9rem 1rem;
                margin-bottom: 0.9rem;
            }}
            .tool-tip strong {{
                display: block;
                color: #FFFFFF;
                font-size: 0.9rem;
                margin-bottom: 0.18rem;
            }}
            .tool-tip span {{
                color: rgba(226,232,240,0.72);
                font-size: 0.82rem;
                line-height: 1.5;
            }}
            .empty-state {{
                max-width: 620px;
                margin-top: 1rem;
            }}
            .empty-state-kicker {{
                color: #FBBF24;
                font-size: 0.72rem;
                font-weight: 800;
                letter-spacing: 0.11em;
                text-transform: uppercase;
                margin-bottom: 0.35rem;
            }}
            .empty-state-title {{
                color: #FFFFFF;
                font-size: 1.15rem;
                font-weight: 800;
                margin-bottom: 0.2rem;
            }}
            .empty-state-copy {{
                color: rgba(226,232,240,0.72);
                font-size: 0.9rem;
                line-height: 1.55;
            }}
            h1, h2, h3, h4, h5, p, label {{
                color: {THEME["white"]};
            }}
            .hero-header {{
                background:
                    radial-gradient(circle at top right, rgba(255,255,255,0.18), transparent 26%),
                    linear-gradient(135deg, #0B1220 0%, #12365E 45%, #0E7C66 100%);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 18px;
                padding: 1.5rem 2rem;
                box-shadow: 0 8px 24px rgba(2, 6, 23, 0.22);
                margin-top: 1rem;
                margin-bottom: 0.75rem;
                overflow: visible;
            }}
            .hero-kicker {{
                text-transform: uppercase;
                letter-spacing: 0.12em;
                font-size: 0.78rem;
                color: rgba(255,255,255,0.72);
                margin-bottom: 0.45rem;
                margin-top: 0;
                font-weight: 700;
            }}
            .hero-title {{
                color: #FFFFFF;
                font-size: 2rem;
                line-height: 1.02;
                font-weight: 800;
                margin-bottom: 0.2rem;
            }}
            .hero-subtitle {{
                color: rgba(255,255,255,0.9);
                font-size: 0.96rem;
                margin-bottom: 0.2rem;
            }}
            .hero-statement {{
                color: rgba(255,255,255,0.74);
                font-size: 0.92rem;
                max-width: 860px;
            }}
            .mode-line {{
                color: rgba(255,255,255,0.74);
                font-size: 0.82rem;
                margin-bottom: 0.55rem;
            }}
            .section-intro {{
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 18px;
                padding: 0.95rem 1rem;
                margin-bottom: 0.75rem;
            }}
            .section-intro-title {{
                color: #FFFFFF;
                font-size: 0.96rem;
                font-weight: 800;
                margin-bottom: 0.22rem;
            }}
            .section-intro-subtitle {{
                color: rgba(255,255,255,0.74);
                line-height: 1.5;
                font-size: 0.86rem;
            }}
            .subsection-label {{
                color: rgba(255,255,255,0.94);
                font-size: 0.92rem;
                font-weight: 700;
                margin-bottom: 0.55rem;
            }}
            .example-description {{
                color: rgba(255,255,255,0.78);
                font-size: 0.82rem;
                font-style: italic;
                min-height: 1.3rem;
                margin-top: 0.45rem;
                margin-bottom: 0.2rem;
            }}
            .numbers-helper, .analyze-note {{
                color: rgba(255,255,255,0.62);
                font-size: 0.78rem;
                line-height: 1.5;
            }}
            .mode-helper {{
                color: rgba(255,255,255,0.72);
                font-size: 0.8rem;
                line-height: 1.5;
                margin-bottom: 0.65rem;
            }}
            /* Data source / upload section polish */
            [data-testid="stSelectbox"] > div[data-baseweb="select"] {{
                background: rgba(15, 23, 42, 0.82) !important;
                border: 1px solid rgba(148, 163, 184, 0.28) !important;
                border-radius: 12px !important;
            }}
            [data-testid="stFileUploader"] {{
                background: rgba(15, 23, 42, 0.55);
                border: 1px solid rgba(148, 163, 184, 0.22);
                border-radius: 16px;
                padding: 0.72rem 0.78rem 0.82rem;
                margin-top: 0.22rem;
            }}
            [data-testid="stFileUploaderDropzone"] {{
                background: rgba(2, 6, 23, 0.66) !important;
                border: 1px dashed rgba(148, 163, 184, 0.42) !important;
                border-radius: 12px !important;
            }}
            [data-testid="stFileUploaderDropzone"] section {{
                padding: 0.55rem 0.35rem !important;
            }}
            [data-testid="stFileUploader"] small {{
                color: rgba(203, 213, 225, 0.78) !important;
            }}
            .analyze-note {{
                text-align: center;
                margin-top: 0.38rem;
                margin-bottom: 0.15rem;
            }}
            .analysis-hero {{
                background: rgba(255,255,255,0.05);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 18px;
                padding: 1rem 1.05rem;
            }}
            .analysis-kicker {{
                color: rgba(255,255,255,0.62);
                font-size: 0.72rem;
                font-weight: 800;
                letter-spacing: 0.12em;
                text-transform: uppercase;
                margin-bottom: 0.35rem;
            }}
            .analysis-headline {{
                color: #FFFFFF;
                font-size: 1.1rem;
                line-height: 1.45;
                font-weight: 700;
                margin-bottom: 0.3rem;
            }}
            .analysis-support {{
                color: rgba(255,255,255,0.74);
                font-size: 0.9rem;
                line-height: 1.6;
            }}
            .analysis-guidance {{
                color: rgba(255,255,255,0.84);
                font-size: 0.92rem;
                line-height: 1.65;
                background: rgba(255,255,255,0.04);
                border: 1px solid rgba(255,255,255,0.07);
                border-radius: 16px;
                padding: 0.9rem 1rem;
            }}
            .analysis-guidance strong {{
                color: #FFFFFF;
            }}
            .chat-stage-header {{
                background: transparent;
                border: none;
                border-radius: 0;
                padding: 0.4rem 0 0.6rem;
            }}
            .chat-stage-title {{
                color: #ECECEC;
                font-size: 1.1rem;
                font-weight: 600;
                margin-bottom: 0.15rem;
                letter-spacing: -0.01em;
            }}
            .chat-stage-copy {{
                color: rgba(236,236,236,0.55);
                font-size: 0.82rem;
                line-height: 1.5;
            }}
            .chat-placeholder {{
                background: rgba(255,255,255,0.04);
                border: 1px dashed rgba(255,255,255,0.12);
                border-radius: 18px;
                padding: 1rem 1.05rem;
                color: rgba(255,255,255,0.68);
                font-size: 0.9rem;
                line-height: 1.55;
                margin-top: 1rem;
            }}
            .chat-block-gap {{
                height: 1rem;
            }}
            .summary-panel, .loan-card, .metric-card, .prompt-card, .preview-card, .agent-card, .mini-card, .advisor-panel, .chat-card, .visual-card {{
                background: linear-gradient(180deg, #1B1F2A 0%, #161A23 100%);
                border: 1px solid rgba(148, 163, 184, 0.18);
                border-radius: 18px;
                box-shadow: 0 8px 22px rgba(0, 0, 0, 0.28);
                color: #E5E7EB;
            }}
            .summary-panel *, .loan-card *, .preview-card *, .agent-card *, .mini-card *, .advisor-panel *, .chat-card *, .visual-card * {{
                color: inherit;
            }}
            .advisor-panel {{
                padding: 1.2rem 1.25rem;
                margin-bottom: 1rem;
            }}
            .advisor-headline {{
                color: {THEME["text"]};
                font-size: 1.08rem;
                line-height: 1.7;
                font-weight: 600;
            }}
            .advisor-grid {{
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 0.8rem;
                margin-top: 0.9rem;
            }}
            .advisor-insight {{
                background: #FFFFFF;
                border: 1px solid {THEME["border"]};
                border-radius: 18px;
                padding: 0.9rem;
            }}
            .advisor-insight span {{
                display: block;
                color: {THEME["muted"]};
                font-size: 0.76rem;
                text-transform: uppercase;
                letter-spacing: 0.08em;
                margin-bottom: 0.28rem;
            }}
            .advisor-insight strong {{
                color: {THEME["text"]};
                line-height: 1.55;
            }}
            .summary-panel {{
                display: flex;
                justify-content: space-between;
                gap: 1rem;
                align-items: flex-start;
                padding: 1.15rem 1.2rem;
                margin-bottom: 1rem;
            }}
            .summary-kicker {{
                color: #7DD3FC;
                text-transform: uppercase;
                letter-spacing: 0.12em;
                font-size: 0.74rem;
                font-weight: 800;
                margin-bottom: 0.42rem;
            }}
            .summary-headline {{
                color: #E5E7EB;
                font-size: 0.96rem;
                line-height: 1.65;
                font-weight: 500;
                max-width: 760px;
            }}
            .source-chip-row {{
                display: flex;
                flex-wrap: wrap;
                gap: 0.5rem;
                margin-top: 0.8rem;
            }}
            .source-chip {{
                color: #93C5FD;
                background: rgba(59, 130, 246, 0.10);
                border: 1px solid rgba(96, 165, 250, 0.32);
                border-radius: 999px;
                padding: 0.34rem 0.7rem;
                font-size: 0.78rem;
                font-weight: 700;
                letter-spacing: 0.01em;
            }}
            .source-list {{
                min-width: 290px;
                display: grid;
                gap: 0.55rem;
            }}
            .source-line {{
                display: flex;
                align-items: center;
                gap: 0.55rem;
                color: #E5E7EB;
                font-weight: 600;
                font-size: 0.88rem;
            }}
            .source-check {{
                width: 20px;
                height: 20px;
                border-radius: 999px;
                background: rgba(16, 185, 129, 0.12);
                border: 1px solid rgba(16, 185, 129, 0.45);
                color: #6EE7B7;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-weight: 800;
                flex-shrink: 0;
            }}
            #duka-ai-forecast-analysis-anchor {{
                scroll-margin-top: 5.5rem;
            }}
            #duka-ai-market-intel-chat-anchor {{
                scroll-margin-top: 5.5rem;
            }}
            .duka-forecast-analyst {{
                border: 1px solid rgba(148, 163, 184, 0.2);
                border-radius: 14px;
                background: rgba(15, 23, 42, 0.72);
                padding: 1rem 1.05rem;
                margin-bottom: 0.85rem;
            }}
            .duka-forecast-narrative-card {{
                border: 1px solid rgba(148, 163, 184, 0.16);
                border-radius: 14px;
                background: rgba(30, 41, 59, 0.65);
                padding: 1.1rem 1.15rem;
                margin-bottom: 1rem;
                color: #F1F5F9;
                font-size: 1.02rem;
                line-height: 1.65;
                font-weight: 500;
            }}
            .duka-forecast-convo-card {{
                margin-top: 1.4rem;
                padding: 1.1rem 1.2rem 0.4rem;
                background: rgba(15, 23, 42, 0.55);
                border: 1px solid rgba(148, 163, 184, 0.18);
                border-radius: 18px;
                box-shadow: 0 10px 30px rgba(0,0,0,0.18);
            }}
            .duka-forecast-convo-head {{
                display: flex;
                align-items: baseline;
                gap: 0.85rem;
                flex-wrap: wrap;
                padding-bottom: 0.7rem;
                margin-bottom: 0.6rem;
                border-bottom: 1px solid rgba(148, 163, 184, 0.16);
            }}
            .duka-forecast-convo-title {{
                color: #F8FAFC;
                font-weight: 800;
                font-size: 1.05rem;
                letter-spacing: 0.01em;
            }}
            .duka-forecast-convo-sub {{
                color: rgba(148, 163, 184, 0.95);
                font-size: 0.84rem;
            }}
            .duka-forecast-empty {{
                color: rgba(148, 163, 184, 0.85);
                font-size: 0.92rem;
                padding: 0.9rem 0.5rem 1.1rem;
                text-align: center;
            }}
            body:has(#duka-ai-forecast-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] {{
                background: rgba(15, 23, 42, 0.7) !important;
                border: 1px solid rgba(148, 163, 184, 0.18) !important;
                border-radius: 14px !important;
                padding: 0.6rem 0.8rem 0.7rem !important;
                margin-bottom: 0.7rem !important;
            }}
            body:has(#duka-ai-forecast-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] p,
            body:has(#duka-ai-forecast-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] li {{
                font-size: 1.02rem !important;
                line-height: 1.62 !important;
                color: #F1F5F9 !important;
            }}
            body:has(#duka-ai-forecast-page-marker) [data-testid="stChatInput"] {{
                background: rgba(15, 23, 42, 0.85) !important;
                border-radius: 14px !important;
                border: 1px solid rgba(148, 163, 184, 0.22) !important;
            }}
            body:has(#duka-ai-market-intel-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] {{
                background: rgba(15, 23, 42, 0.7) !important;
                border: 1px solid rgba(148, 163, 184, 0.18) !important;
                border-radius: 14px !important;
                padding: 0.6rem 0.8rem 0.7rem !important;
                margin-bottom: 0.7rem !important;
            }}
            body:has(#duka-ai-market-intel-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] p,
            body:has(#duka-ai-market-intel-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] li {{
                font-size: 1.02rem !important;
                line-height: 1.62 !important;
                color: #F1F5F9 !important;
            }}
            body:has(#duka-ai-expense-analyzer-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] {{
                background: rgba(15, 23, 42, 0.7) !important;
                border: 1px solid rgba(148, 163, 184, 0.18) !important;
                border-radius: 14px !important;
                padding: 0.6rem 0.8rem 0.7rem !important;
                margin-bottom: 0.7rem !important;
            }}
            body:has(#duka-ai-expense-analyzer-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] p,
            body:has(#duka-ai-expense-analyzer-page-marker) .duka-forecast-convo-card [data-testid="stChatMessage"] [data-testid="stMarkdownContainer"] li {{
                font-size: 1.02rem !important;
                line-height: 1.62 !important;
                color: #F1F5F9 !important;
            }}
            body:has(#duka-ai-expense-analyzer-page-marker) [data-testid="stChatInput"] {{
                background: rgba(15, 23, 42, 0.85) !important;
                border-radius: 14px !important;
                border: 1px solid rgba(148, 163, 184, 0.22) !important;
            }}
            body:has(#duka-ai-market-intel-page-marker) [data-testid="stChatInput"] {{
                background: rgba(15, 23, 42, 0.85) !important;
                border-radius: 14px !important;
                border: 1px solid rgba(148, 163, 184, 0.22) !important;
            }}
            .summary-statuses {{
                display: grid;
                gap: 0.6rem;
                min-width: 220px;
            }}
            .status-card {{
                background: rgba(255,255,255,0.04);
                border: 1px solid rgba(148, 163, 184, 0.22);
                border-radius: 14px;
                padding: 0.75rem 0.9rem;
            }}
            .status-card span {{
                display: block;
                color: rgba(226, 232, 240, 0.62);
                font-size: 0.7rem;
                text-transform: uppercase;
                letter-spacing: 0.1em;
                margin-bottom: 0.28rem;
                font-weight: 700;
            }}
            .status-card strong {{
                color: #F8FAFC;
                font-size: 0.95rem;
                font-weight: 700;
            }}
            .metric-card {{
                padding: 0.95rem 1rem;
                min-height: 136px;
            }}
            .metric-row {{
                display: flex;
                align-items: center;
                gap: 0.45rem;
                margin-bottom: 0.55rem;
            }}
            .metric-icon {{
                width: 28px;
                height: 28px;
                border-radius: 10px;
                background: rgba(255,255,255,0.06);
                color: #F8FAFC;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 800;
            }}
            .metric-label {{
                color: rgba(226, 232, 240, 0.7);
                font-weight: 700;
                letter-spacing: 0.02em;
                font-size: 0.78rem;
                text-transform: uppercase;
            }}
            .metric-value {{
                color: #F8FAFC;
                font-size: 1.65rem;
                font-weight: 800;
                margin-bottom: 0.2rem;
                letter-spacing: -0.01em;
            }}
            .metric-note {{
                color: rgba(226, 232, 240, 0.62);
                line-height: 1.45;
                font-size: 0.84rem;
            }}
            .prompt-card {{
                padding: 0.95rem;
                min-height: 180px;
                margin-bottom: 0.5rem;
                background: linear-gradient(180deg, #172033 0%, #111827 100%);
                border: 1px solid #334155;
                box-shadow: 0 16px 30px rgba(2, 6, 23, 0.22);
            }}
            .prompt-title {{
                color: #F8FAFC;
                font-weight: 800;
                font-size: 1.05rem;
                margin-bottom: 0.55rem;
            }}
            .prompt-preview {{
                color: #CBD5E1;
                line-height: 1.65;
            }}
            .preview-card {{
                padding: 1rem 1.05rem;
                background: linear-gradient(180deg, #131A28 0%, #0F1522 100%);
                border: 1px solid rgba(148, 163, 184, 0.26);
                box-shadow: 0 10px 26px rgba(0, 0, 0, 0.32);
            }}
            .preview-title {{
                color: #F8FAFC;
                font-size: 1rem;
                font-weight: 800;
                margin-bottom: 0.75rem;
            }}
            .preview-summary {{
                color: rgba(226, 232, 240, 0.82);
                line-height: 1.58;
                margin-top: 0.65rem;
            }}
            .preview-grid {{
                display: grid;
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 0.7rem 0.8rem;
            }}
            .preview-grid span {{
                display: block;
                color: rgba(148, 163, 184, 0.95);
                font-size: 0.72rem;
                text-transform: uppercase;
                letter-spacing: 0.09em;
                margin-bottom: 0.18rem;
            }}
            .preview-grid strong {{
                color: #F8FAFC;
                font-size: 0.97rem;
                font-weight: 700;
            }}
            .preview-warning {{
                background: #FFF7ED;
                border: 1px solid #FED7AA;
                border-left: 4px solid {THEME["gold"]};
                color: {THEME["text"]};
                border-radius: 12px;
                padding: 0.7rem 0.8rem;
                margin-top: 0.45rem;
                font-size: 0.9rem;
                line-height: 1.5;
            }}
            .provider-selected {{
                display: flex;
                align-items: center;
                gap: 0.75rem;
            }}
            .provider-selected-label {{
                color: {THEME["muted"]};
                font-size: 0.74rem;
                text-transform: uppercase;
                letter-spacing: 0.08em;
            }}
            .provider-selected-name {{
                color: {THEME["text"]};
                font-weight: 800;
                font-size: 1rem;
            }}
            .provider-tile {{
                position: relative;
                overflow: hidden;
                min-height: 158px;
                background:
                    linear-gradient(145deg, rgba(255,255,255,0.085) 0%, rgba(255,255,255,0.035) 100%),
                    #111827;
                border: 1px solid rgba(255,255,255,0.12);
                border-radius: 18px;
                padding: 0.95rem;
                margin-bottom: 0.55rem;
                transition: transform 0.18s ease, border-color 0.18s ease, background 0.18s ease;
            }}
            .provider-tile:hover {{
                transform: translateY(-2px);
                border-color: color-mix(in srgb, var(--provider) 62%, white 8%);
                background:
                    linear-gradient(145deg, rgba(255,255,255,0.11) 0%, rgba(255,255,255,0.045) 100%),
                    #111827;
            }}
            .provider-tile-connected {{
                background:
                    linear-gradient(145deg, color-mix(in srgb, var(--provider) 18%, transparent) 0%, rgba(255,255,255,0.045) 100%),
                    #111827;
                border-color: color-mix(in srgb, var(--provider) 72%, white 6%);
            }}
            .provider-accent {{
                position: absolute;
                inset: 0 0 auto 0;
                height: 4px;
            }}
            .provider-topline {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 0.65rem;
                margin-bottom: 0.78rem;
            }}
            .provider-logo {{
                width: 48px;
                height: 48px;
                border-radius: 15px;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                font-size: 0.72rem;
                font-weight: 900;
                box-shadow: inset 0 1px 0 rgba(255,255,255,0.24), 0 10px 22px rgba(0,0,0,0.24);
            }}
            .provider-status {{
                color: rgba(255,255,255,0.72);
                border: 1px solid rgba(255,255,255,0.13);
                background: rgba(255,255,255,0.055);
                border-radius: 999px;
                padding: 0.22rem 0.5rem;
                font-size: 0.66rem;
                font-weight: 800;
            }}
            .provider-tile-connected .provider-status {{
                color: #D1FAE5;
                border-color: rgba(16,185,129,0.28);
                background: rgba(16,185,129,0.12);
            }}
            .provider-name {{
                color: #FFFFFF;
                font-size: 0.95rem;
                font-weight: 850;
                line-height: 1.25;
                margin-bottom: 0.2rem;
            }}
            .provider-kind {{
                color: color-mix(in srgb, var(--provider) 82%, white 18%);
                font-size: 0.74rem;
                font-weight: 800;
                margin-bottom: 0.5rem;
            }}
            .provider-signal {{
                color: rgba(226,232,240,0.68);
                min-height: 2.1rem;
                font-size: 0.73rem;
                line-height: 1.45;
            }}
            .provider-meta-row {{
                display: flex;
                flex-wrap: wrap;
                gap: 0.35rem;
                margin-top: 0.75rem;
            }}
            .provider-meta-row span {{
                color: rgba(255,255,255,0.72);
                border: 1px solid rgba(255,255,255,0.10);
                background: rgba(255,255,255,0.045);
                border-radius: 999px;
                padding: 0.2rem 0.45rem;
                font-size: 0.64rem;
                font-weight: 700;
            }}
            .provider-connected-summary {{
                display: flex;
                align-items: center;
                gap: 0.75rem;
                background: rgba(255,255,255,0.055);
                border: 1px solid rgba(255,255,255,0.10);
                border-left: 4px solid;
                border-radius: 14px;
                padding: 0.75rem 0.85rem;
                margin: 0.55rem 0;
            }}
            .provider-dropdown-preview {{
                display: flex;
                align-items: center;
                gap: 0.75rem;
                background:
                    linear-gradient(135deg, color-mix(in srgb, var(--provider) 13%, transparent), rgba(255,255,255,0.045)),
                    #111827;
                border: 1px solid rgba(255,255,255,0.10);
                border-left: 4px solid;
                border-radius: 16px;
                padding: 0.75rem 0.85rem;
                min-height: 70px;
                margin: 0.4rem 0 0.75rem;
            }}
            .provider-connected-mark {{
                width: 42px;
                height: 42px;
                border-radius: 13px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 0.7rem;
                font-weight: 900;
                flex: 0 0 auto;
            }}
            .provider-connected-title {{
                color: #FFFFFF;
                font-size: 0.9rem;
                font-weight: 850;
                margin-bottom: 0.16rem;
            }}
            .provider-connected-copy {{
                color: rgba(226,232,240,0.70);
                font-size: 0.78rem;
                line-height: 1.45;
            }}
            .provider-badge {{
                width: 42px;
                height: 42px;
                border-radius: 14px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 0.74rem;
                font-weight: 800;
                flex-shrink: 0;
            }}
            .provider-badge-large {{
                width: 54px;
                height: 54px;
                border-radius: 16px;
                font-size: 0.85rem;
            }}
            .loan-card, .agent-card, .mini-card, .chat-card, .visual-card {{
                padding: 1rem 1.05rem;
            }}
            .card-title {{
                color: #F8FAFC;
                font-size: 0.96rem;
                font-weight: 800;
                margin-bottom: 0.7rem;
                letter-spacing: 0.01em;
            }}
            .loan-score-row {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 0.8rem;
                margin-bottom: 0.75rem;
            }}
            .loan-score {{
                color: #F8FAFC;
                font-size: 1.9rem;
                font-weight: 800;
                letter-spacing: -0.01em;
            }}
            .loan-badge {{
                background: rgba(255,255,255,0.04);
                border: 1px solid;
                border-radius: 999px;
                padding: 0.36rem 0.72rem;
                font-size: 0.8rem;
                font-weight: 800;
            }}
            .progress-shell {{
                width: 100%;
                height: 10px;
                background: rgba(148, 163, 184, 0.18);
                border-radius: 999px;
                overflow: hidden;
                margin-bottom: 0.8rem;
            }}
            .progress-fill {{
                height: 100%;
                border-radius: 999px;
            }}
            .loan-explainer, .loan-meta {{
                color: rgba(226, 232, 240, 0.72);
                line-height: 1.55;
                margin-bottom: 0.55rem;
                font-size: 0.86rem;
            }}
            .loan-meta strong {{
                color: #F8FAFC;
            }}
            .loan-risk-note {{
                background: #FEF2F2;
                border: 1px solid #FECACA;
                border-left: 4px solid {THEME["red"]};
                color: {THEME["text"]};
                border-radius: 14px;
                padding: 0.8rem 0.9rem;
                line-height: 1.55;
            }}
            .risk-card {{
                display: flex;
                gap: 0.75rem;
                align-items: flex-start;
                background: #FFF7ED;
                border: 1px solid #FED7AA;
                border-left: 5px solid {THEME["red"]};
                border-radius: 16px;
                padding: 0.9rem;
                margin-bottom: 0.7rem;
            }}
            .risk-icon {{
                width: 22px;
                height: 22px;
                border-radius: 999px;
                background: {THEME["red"]};
                color: #FFFFFF;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 0.8rem;
                font-weight: 800;
                flex-shrink: 0;
            }}
            .risk-text, .agent-summary, .agent-detail, .chat-question, .chat-answer {{
                color: {THEME["text"]};
                line-height: 1.58;
                font-size: 0.94rem;
            }}
            .ok-card {{
                background: #ECFDF5;
                border: 1px solid #A7F3D0;
                border-left: 5px solid {THEME["emerald"]};
                color: {THEME["text"]};
                border-radius: 16px;
                padding: 0.9rem;
                line-height: 1.55;
            }}
            .action-card {{
                display: flex;
                gap: 0.85rem;
                align-items: flex-start;
                background: #EFF6FF;
                border: 1px solid #BFDBFE;
                border-radius: 16px;
                padding: 0.9rem;
                margin-bottom: 0.65rem;
            }}
            .action-number {{
                width: 28px;
                height: 28px;
                border-radius: 10px;
                background: {THEME["blue"]};
                color: #FFFFFF;
                display: flex;
                align-items: center;
                justify-content: center;
                font-weight: 800;
                flex-shrink: 0;
            }}
            .action-text {{
                color: {THEME["text"]};
                line-height: 1.55;
                font-size: 0.94rem;
            }}
            .agent-detail {{
                background: rgba(255,255,255,0.9);
                border: 1px solid {THEME["border"]};
                border-radius: 14px;
                padding: 0.78rem 0.88rem;
                color: {THEME["muted"]};
                margin-bottom: 0.5rem;
            }}
            .market-subhead {{
                color: #F8FAFC;
                font-weight: 800;
                font-size: 0.78rem;
                letter-spacing: 0.08em;
                text-transform: uppercase;
                margin: 1rem 0 0.5rem;
                display: flex;
                align-items: center;
                gap: 0.5rem;
            }}
            .market-subhead::before {{
                content: "";
                display: inline-block;
                width: 4px;
                height: 14px;
                background: linear-gradient(180deg, #34D399, #2563EB);
                border-radius: 2px;
            }}
            .market-pill-row {{
                display: flex;
                flex-wrap: wrap;
                gap: 0.5rem;
                margin-bottom: 0.2rem;
            }}
            .market-pill {{
                background: rgba(239, 68, 68, 0.14);
                border: 1px solid rgba(248, 113, 113, 0.45);
                color: #FCA5A5;
                border-radius: 999px;
                padding: 0.42rem 0.78rem;
                font-size: 0.84rem;
                font-weight: 600;
            }}
            .market-pill-blue {{
                background: rgba(37, 99, 235, 0.16);
                border-color: rgba(96, 165, 250, 0.45);
                color: #93C5FD;
            }}
            .agent-card .agent-summary {{
                color: #E5E7EB !important;
                background: rgba(15, 23, 42, 0.55);
                border: 1px solid rgba(148, 163, 184, 0.18);
                border-left: 3px solid #34D399;
                border-radius: 12px;
                padding: 0.85rem 1rem;
                font-size: 0.96rem;
                line-height: 1.6;
                margin: 0.4rem 0 0.4rem;
            }}
            .agent-card .agent-detail {{
                background: rgba(15, 23, 42, 0.55) !important;
                border: 1px solid rgba(148, 163, 184, 0.18) !important;
                color: #E5E7EB !important;
                border-radius: 12px;
                padding: 0.78rem 0.95rem;
                font-size: 0.94rem;
                line-height: 1.55;
                margin-bottom: 0.4rem;
            }}
            .chat-label {{
                color: {THEME["muted"]};
                font-size: 0.74rem;
                text-transform: uppercase;
                letter-spacing: 0.08em;
                margin-bottom: 0.2rem;
                font-weight: 800;
            }}
            .chat-question {{
                margin-bottom: 0.7rem;
                font-weight: 700;
            }}
            .visual-row {{
                display: flex;
                justify-content: space-between;
                gap: 1rem;
                padding: 0.65rem 0;
                border-top: 1px solid {THEME["border"]};
                color: #E5E7EB;
                font-size: 0.92rem;
            }}
            .visual-row:first-of-type {{
                border-top: none;
                padding-top: 0.2rem;
            }}
            .visual-row span {{
                color: rgba(226, 232, 240, 0.72);
                font-weight: 500;
            }}
            .visual-row strong {{
                color: #F8FAFC;
                font-weight: 700;
                text-align: right;
            }}
            .chat-invite {{
                color: rgba(255,255,255,0.52);
                font-size: 0.92rem;
                margin-top: 0.95rem;
                margin-bottom: 0.2rem;
                font-style: italic;
                line-height: 1.65;
            }}
            .agent-label {{
                color: rgba(255,255,255,0.38);
                font-size: 0.7rem;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.1em;
                margin-bottom: 0.3rem;
            }}
            /* User message bubble */
            .user-bubble-wrap {{
                display: flex;
                justify-content: flex-end;
                margin: 0.6rem 0 0.9rem 0;
                padding-left: 3rem;
                gap: 0.55rem;
                align-items: flex-end;
            }}
            .user-bubble {{
                background: linear-gradient(135deg, #1D9E75 0%, #14B783 60%, #2563EB 140%);
                color: #FFFFFF;
                border-radius: 18px 18px 4px 18px;
                padding: 0.7rem 1.05rem;
                max-width: 75%;
                font-size: 0.95rem;
                line-height: 1.55;
                word-wrap: break-word;
                box-shadow: 0 4px 14px rgba(29, 158, 117, 0.22);
                border: 1px solid rgba(255,255,255,0.08);
                font-weight: 500;
            }}
            /* Polished avatars for st.chat_message (assistant) */
            [data-testid="stChatMessageAvatarAssistant"] {{
                background: linear-gradient(135deg, #1D9E75, #2563EB) !important;
                color: #FFFFFF !important;
                border: 1px solid rgba(255,255,255,0.15) !important;
                box-shadow: 0 4px 12px rgba(37, 99, 235, 0.28) !important;
                width: 34px !important;
                height: 34px !important;
                border-radius: 50% !important;
                display: flex !important;
                align-items: center !important;
                justify-content: center !important;
                font-size: 1.05rem !important;
            }}
            [data-testid="stChatMessageAvatarUser"] {{
                background: linear-gradient(135deg, #475569, #1E293B) !important;
                color: #F8FAFC !important;
                border: 1px solid rgba(255,255,255,0.12) !important;
                width: 34px !important;
                height: 34px !important;
                border-radius: 50% !important;
                font-size: 1rem !important;
            }}
            /* Inline metrics strip */
            .inline-metrics {{
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.1);
                border-radius: 10px;
                padding: 0.42rem 0.92rem;
                color: rgba(255,255,255,0.68);
                font-size: 0.83rem;
                font-family: 'SF Mono', 'Segoe UI Mono', 'Courier New', monospace;
                margin: 0.65rem 0 0.72rem;
                display: inline-block;
                letter-spacing: 0.01em;
            }}
            /* Suggestion pills label */
            .sug-pills-label {{
                color: rgba(255,255,255,0.35);
                font-size: 0.7rem;
                font-weight: 700;
                text-transform: uppercase;
                letter-spacing: 0.12em;
                margin: 0.5rem 0 0.45rem;
            }}
            /* Assistant message — flat ChatGPT-style, no card */
            [data-testid="stChatMessage"] {{
                background: transparent !important;
                border: none !important;
                border-radius: 0 !important;
                margin: 0.4rem 0 1.2rem 0 !important;
                padding: 0 !important;
                gap: 0.75rem !important;
            }}
            [data-testid="stChatMessage"] [data-testid="stChatMessageAvatarUser"],
            [data-testid="stChatMessage"] [data-testid="stChatMessageAvatarAssistant"] {{
                background: #19C37D !important;
                border: none !important;
                box-shadow: none !important;
            }}
            [data-testid="stChatMessageContent"] {{
                background: transparent !important;
                border: none !important;
                padding: 0 !important;
            }}
            [data-testid="stChatMessageContent"] p,
            [data-testid="stChatMessageContent"] li {{
                color: #ECECEC !important;
                font-size: 0.96rem !important;
                line-height: 1.7 !important;
            }}
            [data-testid="stChatMessageContent"] p {{
                margin: 0 0 0.75rem 0 !important;
            }}
            [data-testid="stChatMessageContent"] p:last-child {{
                margin-bottom: 0 !important;
            }}
            [data-testid="stChatMessageContent"] h1,
            [data-testid="stChatMessageContent"] h2,
            [data-testid="stChatMessageContent"] h3,
            [data-testid="stChatMessageContent"] h4 {{
                color: #FFFFFF !important;
            }}
            /* "Next step:" callout — only when the agent ends with a real action */
            [data-testid="stChatMessageContent"] p:has(> strong:first-child:only-child),
            [data-testid="stChatMessageContent"] p:has(> strong:first-child) {{
                /* base paragraph style preserved */
            }}
            .next-step-callout {{
                margin-top: 0.85rem !important;
                padding: 0.7rem 0.9rem !important;
                background: rgba(25, 195, 125, 0.08) !important;
                border-left: 3px solid #19C37D !important;
                border-radius: 8px !important;
                color: #ECECEC !important;
                font-size: 0.92rem !important;
                line-height: 1.55 !important;
            }}
            .next-step-callout strong {{
                color: #19C37D !important;
                margin-right: 0.4rem;
            }}
            [data-testid="stChatMessageContent"] code {{
                background: rgba(255,255,255,0.06) !important;
                color: #F8FAFC !important;
                padding: 0.1rem 0.35rem;
                border-radius: 6px;
                font-size: 0.88em;
            }}
            [data-testid="stMetric"] {{
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.08);
                border-radius: 16px;
                padding: 0.95rem 1rem;
                min-height: 108px;
                box-shadow: inset 0 1px 0 rgba(255,255,255,0.04);
            }}
            [data-testid="stMetricLabel"] p {{
                color: rgba(255,255,255,0.64) !important;
                font-size: 0.8rem !important;
                line-height: 1.2 !important;
            }}
            [data-testid="stMetricValue"] {{
                color: #F8FAFC !important;
                font-size: 1.45rem !important;
                line-height: 1.15 !important;
            }}
            [data-testid="stMetricDelta"] {{
                display: none;
            }}
            [data-testid="stCaptionContainer"] p {{
                color: rgba(255,255,255,0.58) !important;
                font-size: 0.76rem !important;
            }}
            .stDivider {{
                margin: 0.9rem 0 1rem 0;
            }}
            /* Static composer panel — solid background, anchored at bottom of column */
            [data-testid="stBottom"],
            [data-testid="stBottom"] > div {{
                background: #212121 !important;
                background-color: #212121 !important;
                background-image: none !important;
                border: none !important;
                box-shadow: none !important;
                backdrop-filter: none !important;
                -webkit-backdrop-filter: none !important;
            }}
            [data-testid="stBottom"] {{
                padding: 0 !important;
            }}
            /* Constrain the bottom composer to the chat column width */
            [data-testid="stBottomBlockContainer"] {{
                max-width: 780px !important;
                margin: 0 auto !important;
                padding: 0.85rem 1rem 1rem !important;
                background: #212121 !important;
            }}
            /* Composer panel — solid card, NOT transparent */
            [data-testid="stChatInput"],
            [data-testid="stChatInput"] > div,
            [data-testid="stChatInput"] section,
            [data-testid="stChatInput"] form,
            [data-testid="stChatInputContainer"] {{
                background: #2F2F2F !important;
                background-color: #2F2F2F !important;
                background-image: none !important;
                border: 1px solid #3A3A3A !important;
                border-radius: 24px !important;
                box-shadow: none !important;
                backdrop-filter: none !important;
                -webkit-backdrop-filter: none !important;
            }}
            [data-testid="stChatInput"] {{
                margin: 0 !important;
                padding: 0 !important;
            }}
            /* The textarea sits flush inside the panel — no double border */
            [data-testid="stChatInput"] textarea,
            [data-testid="stChatInputTextArea"] {{
                background: transparent !important;
                border: none !important;
                border-radius: 24px !important;
                color: #ECECEC !important;
                font-size: 1rem !important;
                line-height: 1.55 !important;
                min-height: 52px !important;
                padding: 0.95rem 3rem 0.95rem 1.15rem !important;
                box-shadow: none !important;
                outline: none !important;
                resize: none !important;
            }}
            [data-testid="stChatInput"]:hover {{
                border-color: #4A4A4A !important;
            }}
            [data-testid="stChatInput"]:focus-within {{
                border-color: #5A5A5A !important;
            }}
            [data-testid="stChatInput"] textarea:focus,
            [data-testid="stChatInput"] textarea:focus-visible {{
                background: transparent !important;
                border: none !important;
                box-shadow: none !important;
                outline: none !important;
            }}
            [data-testid="stChatInput"] textarea::placeholder {{
                color: rgba(236,236,236,0.45) !important;
            }}
            /* Send button — flat white circle inside the input pill */
            [data-testid="stChatInput"] button,
            [data-testid="stChatInputSubmitButton"] {{
                background: #ECECEC !important;
                color: #1F1F1F !important;
                border-radius: 50% !important;
                border: none !important;
                box-shadow: none !important;
                width: 30px !important;
                height: 30px !important;
                min-width: 30px !important;
                min-height: 30px !important;
                margin: 0.55rem !important;
            }}
            [data-testid="stChatInput"] button:hover {{
                background: #FFFFFF !important;
                box-shadow: none !important;
            }}
            [data-testid="stChatInput"] button:disabled {{
                background: #3A3A3A !important;
                color: rgba(236,236,236,0.4) !important;
            }}
            [data-testid="stChatInput"] button svg {{
                fill: #1F1F1F !important;
                color: #1F1F1F !important;
            }}
            /* Form inputs */
            .stTextInput label, .stTextArea label, .stSelectbox label, .stFileUploader label {{
                color: #E5E7EB !important;
                font-weight: 700;
                font-size: 0.86rem !important;
            }}
            .stTextInput input,
            .stTextArea textarea,
            .stSelectbox div[data-baseweb="select"] > div,
            .stSelectbox input {{
                background: #111827 !important;
                color: #F8FAFC !important;
                border: 1px solid #334155 !important;
                border-radius: 14px !important;
                box-shadow: inset 0 1px 0 rgba(255,255,255,0.02);
            }}
            .stTextInput input,
            .stSelectbox div[data-baseweb="select"] > div {{
                min-height: 3rem;
            }}
            .stTextArea textarea {{
                min-height: 150px;
                line-height: 1.6;
            }}
            .stTextInput input::placeholder,
            .stTextArea textarea::placeholder {{
                color: #94A3B8 !important;
                opacity: 1 !important;
            }}
            .stTextInput input:focus,
            .stTextArea textarea:focus,
            .stSelectbox div[data-baseweb="select"] > div:focus-within {{
                border-color: #10B981 !important;
                box-shadow: 0 0 0 1px #10B981, 0 0 0 4px rgba(16, 185, 129, 0.12) !important;
                outline: none !important;
            }}
            .stSelectbox svg {{ fill: #CBD5E1 !important; }}
            div[data-baseweb="menu"] {{
                background: #111827 !important;
                border: 1px solid #334155 !important;
                border-radius: 14px !important;
            }}
            div[data-baseweb="menu"] [role="option"] {{ background: transparent !important; }}
            div[data-baseweb="menu"] [role="option"]:hover {{ background: rgba(16, 185, 129, 0.14) !important; }}
            div[data-baseweb="menu"] * {{ color: #F8FAFC !important; }}
            [data-testid="stFileUploader"] {{
                background: #1F2937;
                border: 1px solid rgba(148, 163, 184, 0.3);
                border-radius: 18px;
                padding: 0.75rem;
            }}
            [data-testid="stFileUploaderDropzone"] {{
                background: #111827 !important;
                border: 1px dashed rgba(148, 163, 184, 0.45) !important;
                border-radius: 16px !important;
                color: #F8FAFC !important;
            }}
            [data-testid="stFileUploaderDropzone"]:hover {{
                border-color: #10B981 !important;
                background: rgba(16, 185, 129, 0.06) !important;
            }}
            [data-testid="stFileUploaderDropzone"] * {{ color: #CBD5E1 !important; }}
            [data-testid="stFileUploader"] button {{
                background: linear-gradient(135deg, #1D9E75 0%, #10B981 100%) !important;
                border: 1px solid rgba(16, 185, 129, 0.55) !important;
                color: #FFFFFF !important;
                border-radius: 10px !important;
                font-weight: 700 !important;
                box-shadow: none !important;
            }}
            [data-testid="stFileUploader"] button:hover {{
                background: linear-gradient(135deg, #178762 0%, #0EA271 100%) !important;
                border-color: rgba(16, 185, 129, 0.72) !important;
                color: #FFFFFF !important;
            }}
            .stTabs [data-baseweb="tab-list"] {{
                gap: 0.6rem;
                background: rgba(255,255,255,0.04);
                border: 1px solid rgba(255,255,255,0.09);
                border-radius: 20px;
                padding: 0.45rem 0.5rem;
            }}
            .stTabs [data-baseweb="tab"] {{
                background: transparent;
                border-radius: 14px;
                color: rgba(255,255,255,0.55);
                min-height: 2.9rem;
                padding: 0 1.2rem;
                font-weight: 700;
                font-size: 0.92rem;
                letter-spacing: 0.01em;
                border: 1px solid transparent;
                transition: all 0.18s ease;
            }}
            .stTabs [data-baseweb="tab"]:hover {{
                background: rgba(255,255,255,0.07);
                color: rgba(255,255,255,0.85);
                border-color: rgba(255,255,255,0.10);
            }}
            .stTabs [aria-selected="true"] {{
                background: linear-gradient(135deg, #1E3A5F 0%, #0E4D3A 100%) !important;
                color: #FFFFFF !important;
                border-color: rgba(16, 185, 129, 0.35) !important;
                box-shadow: 0 4px 14px rgba(16, 185, 129, 0.18), inset 0 1px 0 rgba(255,255,255,0.10);
            }}
            .stButton > button, .stDownloadButton > button, .stFormSubmitButton > button {{
                border-radius: 999px;
                min-height: 2.35rem;
                font-weight: 600;
                border: none;
                transition: all 0.15s ease;
            }}
            .stButton > button {{
                background: rgba(255,255,255,0.06);
                color: #CBD5E1;
                border: 1px solid rgba(255,255,255,0.12);
                font-size: 0.88rem;
                padding: 0.35rem 1.1rem;
            }}
            .stButton > button:hover {{
                background: rgba(255,255,255,0.12);
                border-color: rgba(16,185,129,0.45);
                color: #FFFFFF;
            }}
            .duka-ai-example-pill-btn,
            .duka-ai-number-pill-btn,
            .duka-ai-suggestion-pill-btn {{
                background: transparent !important;
                border: 0.5px solid rgba(148, 163, 184, 0.55) !important;
                color: #E2E8F0 !important;
                box-shadow: none !important;
            }}
            .duka-ai-example-pill-btn:hover,
            .duka-ai-number-pill-btn:hover,
            .duka-ai-suggestion-pill-btn:hover {{
                background: rgba(255,255,255,0.08) !important;
                border-color: rgba(191, 219, 254, 0.85) !important;
                color: #FFFFFF !important;
            }}
            .duka-ai-example-pill-btn.is-selected,
            .duka-ai-number-pill-btn.is-selected {{
                background: rgba(37, 99, 235, 0.16) !important;
                border-color: rgba(96, 165, 250, 0.78) !important;
                color: #FFFFFF !important;
            }}
            .duka-ai-example-pill-btn {{
                min-height: 2.3rem !important;
                font-size: 0.86rem !important;
                padding: 0.3rem 0.92rem !important;
            }}
            .duka-ai-number-pill-btn {{
                min-height: 2.15rem !important;
                font-size: 0.8rem !important;
                padding: 0.26rem 0.85rem !important;
            }}
            .duka-ai-suggestion-pill-btn {{
                width: 100% !important;
                min-height: 40px !important;
                height: 40px !important;
                font-size: 12.5px !important;
                font-weight: 600 !important;
                padding: 0.25rem 0.85rem !important;
                background: rgba(255,255,255,0.05) !important;
                border-radius: 12px !important;
                white-space: nowrap !important;
                overflow: hidden !important;
                text-overflow: ellipsis !important;
                display: inline-flex !important;
                align-items: center !important;
                justify-content: center !important;
            }}
            .duka-ai-analyze-btn {{
                min-height: 4.1rem !important;
                font-size: 1.14rem !important;
                font-weight: 800 !important;
                padding: 0.95rem 1.25rem !important;
                box-shadow: 0 14px 30px rgba(16, 185, 129, 0.24) !important;
            }}
            .stButton > button[kind="primary"], .stFormSubmitButton > button {{
                background: linear-gradient(135deg, {THEME["blue"]} 0%, {THEME["emerald"]} 100%);
                color: #FFFFFF;
                box-shadow: 0 10px 24px rgba(37, 99, 235, 0.22);
                min-height: 3.5rem;
                font-size: 1.08rem;
                letter-spacing: 0.02em;
                border-radius: 999px;
            }}
            .stDownloadButton > button {{
                background: rgba(255,255,255,0.07);
                color: #E2E8F0;
                border: 1px solid rgba(255,255,255,0.14);
                font-size: 0.84rem;
            }}
            .stExpander {{
                border: 1px solid rgba(148, 163, 184, 0.18) !important;
                border-radius: 16px !important;
                background: rgba(15, 20, 32, 0.55) !important;
                box-shadow: 0 6px 18px rgba(0, 0, 0, 0.18);
            }}
            .stExpander summary {{
                color: #E5E7EB !important;
                font-weight: 700 !important;
                padding: 0.85rem 1rem !important;
            }}
            .stExpander summary:hover {{
                color: #FFFFFF !important;
                background: rgba(255,255,255,0.03) !important;
                border-radius: 14px !important;
            }}
            .stExpander [data-testid="stExpanderDetails"] {{
                background: rgba(10, 14, 22, 0.55) !important;
                border-top: 1px solid rgba(148, 163, 184, 0.14) !important;
                padding: 1rem 1rem 1.1rem !important;
                border-radius: 0 0 14px 14px !important;
            }}
            [data-testid="stPlotlyChart"] {{
                background: transparent !important;
                border-radius: 12px;
            }}
            [data-testid="stPlotlyChart"] > div {{
                background: transparent !important;
                border-radius: 12px;
            }}
            [data-testid="stStatusWidget"] {{
                background: rgba(15,23,42,0.85) !important;
                border: 1px solid rgba(255,255,255,0.12) !important;
                border-radius: 16px !important;
            }}
            [data-testid="stChatMessage"] {{
                background: transparent !important;
                box-shadow: none !important;
                padding-top: 0 !important;
                padding-bottom: 0 !important;
            }}
            [data-testid="stChatMessageContent"] {{
                background: transparent !important;
                padding: 0 !important;
            }}
            .welcome-hero {{
                text-align: center;
                padding: 3.5rem 1rem 2.25rem;
            }}
            .welcome-brand {{
                font-size: 3.8rem;
                font-weight: 900;
                line-height: 1;
                background: linear-gradient(135deg, #FFFFFF 0%, #A7F3D0 45%, #60A5FA 100%);
                -webkit-background-clip: text;
                -webkit-text-fill-color: transparent;
                background-clip: text;
                margin-bottom: 0.55rem;
            }}
            .welcome-tagline {{
                color: rgba(255,255,255,0.92);
                font-size: 1.45rem;
                font-weight: 700;
                margin-bottom: 0.75rem;
            }}
            .welcome-desc {{
                color: rgba(255,255,255,0.62);
                font-size: 1rem;
                max-width: 560px;
                margin: 0 auto;
                line-height: 1.65;
            }}
            .welcome-option-card {{
                background: rgba(255,255,255,0.06);
                border: 1px solid rgba(255,255,255,0.12);
                border-radius: 22px;
                padding: 1.5rem 1.2rem 1.2rem;
                text-align: center;
                margin-bottom: 0;
                min-height: 200px;
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: flex-start;
                box-sizing: border-box;
                transition: border-color 0.18s ease, background 0.18s ease;
            }}
            .welcome-option-card:hover {{
                border-color: rgba(16,185,129,0.35);
                background: rgba(255,255,255,0.09);
            }}
            .woc-icon {{
                font-size: 2.2rem;
                margin-bottom: 0.65rem;
                flex-shrink: 0;
            }}
            .woc-title {{
                color: #FFFFFF;
                font-size: 1.02rem;
                font-weight: 800;
                margin-bottom: 0.4rem;
                flex-shrink: 0;
            }}
            .woc-desc {{
                color: rgba(255,255,255,0.62);
                font-size: 0.82rem;
                line-height: 1.55;
                flex: 1 1 auto;
                min-height: 4.6em;
            }}
            .woc-badge {{
                display: inline-block;
                background: linear-gradient(135deg, #F59E0B, #EAB308);
                color: #1F2937;
                font-size: 0.65rem;
                font-weight: 800;
                letter-spacing: 0.08em;
                text-transform: uppercase;
                padding: 0.18rem 0.55rem;
                border-radius: 999px;
                margin-bottom: 0.6rem;
                box-shadow: 0 2px 6px rgba(0,0,0,0.18);
            }}
            /* Welcome triad: equal column height, buttons on one baseline */
            div[data-testid="stHorizontalBlock"]:has(.welcome-option-card) {{
                align-items: stretch !important;
            }}
            div[data-testid="column"]:has(.welcome-option-card) {{
                display: flex !important;
                flex-direction: column !important;
                align-items: stretch !important;
            }}
            div[data-testid="column"]:has(.welcome-option-card) > *:first-child {{
                flex: 1 1 auto !important;
            }}
            div[data-testid="column"]:has(.welcome-option-card) > *:last-child {{
                margin-top: auto !important;
                width: 100% !important;
                padding-top: 0.65rem !important;
            }}
            div[data-testid="column"]:has(.welcome-option-card) .stButton > button {{
                width: 100% !important;
                min-height: 2.85rem !important;
                padding-top: 0.45rem !important;
                padding-bottom: 0.45rem !important;
            }}
            @media (max-width: 900px) {{
                .summary-panel, .advisor-grid {{
                    flex-direction: column;
                    grid-template-columns: 1fr;
                }}
                .summary-statuses, .source-list {{ width: 100%; }}
                .preview-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
                .hero-title {{ font-size: 2.1rem; }}
            }}
            @media (max-width: 400px) {{
                .welcome-brand {{ font-size: 2.4rem !important; }}
                .welcome-tagline {{ font-size: 1.1rem !important; }}
                .welcome-desc {{ font-size: 0.82rem !important; }}
                .welcome-option-card {{ min-height: 11rem; padding: 1rem 0.9rem 0.85rem; }}
                .hero-title {{ font-size: 1.5rem !important; }}
                .hero-subtitle {{ font-size: 0.82rem !important; }}
                .block-container {{ padding-left: 0.5rem !important; padding-right: 0.5rem !important; }}
                .preview-grid {{ grid-template-columns: 1fr !important; }}
                .metric-card {{ padding: 0.65rem 0.7rem !important; }}
                .summary-panel {{ padding: 0.75rem 0.8rem !important; }}
                .loan-card, .agent-card, .mini-card, .chat-card, .visual-card {{ padding: 0.75rem 0.8rem !important; }}
                .page-title {{ font-size: 1.2rem !important; }}
                .page-subtitle {{ font-size: 0.78rem !important; }}
                .analysis-headline {{ font-size: 0.95rem !important; }}
                .loan-score {{ font-size: 1.5rem !important; }}
                .summary-statuses {{ flex-direction: column !important; gap: 0.4rem !important; }}
                .woc-title {{ font-size: 0.9rem !important; }}
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )

    # ── Polish CSS: fonts, cards, button hover, scrollbar, footer ────────────
    st.markdown(
        """
        <style>
        html, body, [class*="css"] {
            font-family: 'Inter', 'Arial', sans-serif;
        }
        [data-testid="stSidebar"] {
            background-color: #0D1117 !important;
            border-right: 1px solid #1A3A5C;
        }
        .sidebar-section-label {
            font-size: 10px;
            font-weight: 600;
            letter-spacing: 1.5px;
            color: #4A5568;
            text-transform: uppercase;
            padding: 12px 12px 4px;
            display: block;
        }
        .hero-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            background: rgba(29,158,117,0.15);
            border: 1px solid rgba(29,158,117,0.3);
            color: #1D9E75;
            font-size: 12px;
            font-weight: 500;
            padding: 4px 12px;
            border-radius: 20px;
            margin-bottom: 8px;
        }
        .trust-item {
            text-align: center;
            padding: 12px 8px;
            color: #6B7280;
            font-size: 12px;
        }
        .trust-item span {
            display: block;
            font-size: 18px;
            margin-bottom: 4px;
        }
        .sample-card-featured {
            border: 1.5px solid #1D9E75 !important;
            border-radius: 12px;
            padding: 16px;
            background: rgba(29,158,117,0.05);
            position: relative;
            margin-bottom: 8px;
        }
        .sample-card-normal {
            border: 1px solid #2D3748;
            border-radius: 12px;
            padding: 16px;
            background: rgba(255,255,255,0.02);
            margin-bottom: 8px;
        }
        .recommended-badge {
            display: inline-block;
            background: #1D9E75;
            color: white;
            font-size: 10px;
            font-weight: 700;
            padding: 2px 8px;
            border-radius: 10px;
            margin-bottom: 8px;
            letter-spacing: 0.5px;
            text-transform: uppercase;
        }
        .card-icon {
            width: 32px;
            height: 32px;
            border-radius: 8px;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 16px;
            margin-bottom: 8px;
        }
        .stButton > button {
            border-radius: 8px !important;
            font-size: 13px !important;
            transition: all 0.2s ease !important;
        }
        .stButton > button:hover {
            transform: translateY(-1px) !important;
            box-shadow: 0 4px 12px rgba(29,158,117,0.3) !important;
        }
        div[data-testid="stButton"] > button[kind="primary"] {
            background: linear-gradient(135deg, #1D9E75, #16836A) !important;
            border: none !important;
            font-size: 15px !important;
            font-weight: 600 !important;
            padding: 14px 24px !important;
            border-radius: 10px !important;
            letter-spacing: 0.3px !important;
            box-shadow: 0 4px 15px rgba(29,158,117,0.4) !important;
        }
        div[data-testid="stButton"] > button[kind="primary"]:hover {
            background: linear-gradient(135deg, #22B584, #1D9E75) !important;
            box-shadow: 0 6px 20px rgba(29,158,117,0.5) !important;
            transform: translateY(-2px) !important;
        }
        h1, h2, h3 {
            text-transform: none !important;
            letter-spacing: -0.3px;
        }
        [data-testid="stChatInput"] {
            border-radius: 12px !important;
        }
        [data-testid="metric-container"] {
            border-radius: 10px !important;
            padding: 12px !important;
        }
        footer {visibility: hidden;}
        #MainMenu {visibility: hidden;}
        ::-webkit-scrollbar { width: 4px; }
        ::-webkit-scrollbar-track { background: #0D1117; }
        ::-webkit-scrollbar-thumb { background: #1A3A5C; border-radius: 4px; }

        /* ═══════════════════════════════════════════════════════
           DUKA AI — BRANDED LOADER SYSTEM
           ═══════════════════════════════════════════════════════ */

        /* ── Keyframe animations ──────────────────────────────── */
        @keyframes duka-spin {{
            0%   {{ transform: rotate(0deg); }}
            100% {{ transform: rotate(360deg); }}
        }}
        @keyframes duka-pulse-dot {{
            0%, 80%, 100% {{ transform: scale(0.55); opacity: 0.25; }}
            40%            {{ transform: scale(1.0);  opacity: 1.0; }}
        }}
        @keyframes duka-shimmer {{
            0%   {{ background-position: -400px 0; }}
            100% {{ background-position:  400px 0; }}
        }}
        @keyframes duka-fade-up {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to   {{ opacity: 1; transform: translateY(0); }}
        }}
        @keyframes duka-glow-border {{
            0%, 100% {{ border-color: rgba(29,158,117,0.25); box-shadow: 0 0 0 0 rgba(29,158,117,0); }}
            50%       {{ border-color: rgba(29,158,117,0.55); box-shadow: 0 0 14px 2px rgba(29,158,117,0.15); }}
        }}

        /* ── Branded loader card ─────────────────────────────── */
        .duka-loader {{
            display: flex;
            align-items: center;
            gap: 18px;
            background: linear-gradient(135deg, #0D1B2E 0%, #162035 100%);
            border: 1px solid rgba(29,158,117,0.3);
            border-radius: 16px;
            padding: 18px 22px;
            margin: 10px 0;
            animation: duka-fade-up 0.28s ease-out, duka-glow-border 2.4s ease-in-out infinite;
            box-shadow: 0 6px 28px rgba(0,0,0,0.45), inset 0 1px 0 rgba(255,255,255,0.04);
        }}

        /* Spinning ring indicator */
        .duka-ring {{
            width: 38px;
            height: 38px;
            border: 3px solid rgba(29,158,117,0.18);
            border-top: 3px solid #1D9E75;
            border-right: 3px solid rgba(29,158,117,0.55);
            border-radius: 50%;
            animation: duka-spin 0.75s linear infinite;
            flex-shrink: 0;
        }}

        /* Bouncing dots indicator */
        .duka-dots {{
            display: flex;
            align-items: center;
            gap: 5px;
            flex-shrink: 0;
        }}
        .duka-dots span {{
            width: 9px;
            height: 9px;
            background: #1D9E75;
            border-radius: 50%;
            animation: duka-pulse-dot 1.35s ease-in-out infinite;
        }}
        .duka-dots span:nth-child(1) {{ animation-delay: 0s;    }}
        .duka-dots span:nth-child(2) {{ animation-delay: 0.18s; }}
        .duka-dots span:nth-child(3) {{ animation-delay: 0.36s; }}

        /* Loader text */
        .duka-loader-body {{ flex: 1; min-width: 0; }}
        .duka-loader-title {{
            font-size: 13.5px;
            font-weight: 700;
            color: #E2E8F0;
            margin-bottom: 3px;
            letter-spacing: 0.015em;
        }}
        .duka-loader-sub {{
            font-size: 11.5px;
            color: rgba(148,163,184,0.75);
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}
        /* Brand accent on loader */
        .duka-loader-badge {{
            font-size: 10px;
            font-weight: 700;
            letter-spacing: 0.08em;
            text-transform: uppercase;
            color: #1D9E75;
            opacity: 0.85;
            white-space: nowrap;
            flex-shrink: 0;
        }}

        /* ── Skeleton shimmer placeholders ──────────────────── */
        .duka-skeleton {{
            background: linear-gradient(
                90deg,
                rgba(26,39,68,0) 0%,
                rgba(55,90,140,0.35) 40%,
                rgba(26,39,68,0) 80%
            );
            background-size: 400px 100%;
            animation: duka-shimmer 1.5s ease-in-out infinite;
            border-radius: 6px;
        }}
        .duka-skeleton-line {{ height: 13px; margin: 6px 0; }}
        .duka-skeleton-line.wide  {{ width: 85%; }}
        .duka-skeleton-line.med   {{ width: 60%; }}
        .duka-skeleton-line.short {{ width: 40%; }}

        /* ── Theme Streamlit's native st.spinner ─────────────── */
        [data-testid="stSpinner"] > div {{
            background: linear-gradient(135deg, #0D1B2E 0%, #162035 100%) !important;
            border: 1px solid rgba(29,158,117,0.28) !important;
            border-radius: 14px !important;
            padding: 14px 20px !important;
            box-shadow: 0 4px 22px rgba(0,0,0,0.4) !important;
            animation: duka-glow-border 2.4s ease-in-out infinite !important;
        }}
        [data-testid="stSpinner"] p {{
            color: #94A3B8 !important;
            font-size: 13px !important;
            font-weight: 500 !important;
        }}
        /* Colour the SVG spinner icon teal */
        [data-testid="stSpinner"] svg {{
            color: #1D9E75 !important;
        }}

        /* ── Theme Streamlit's native st.status ──────────────── */
        [data-testid="stStatusWidget"] {{
            background: linear-gradient(135deg, #0D1B2E 0%, #162035 100%) !important;
            border: 1px solid rgba(29,158,117,0.28) !important;
            border-radius: 14px !important;
            box-shadow: 0 4px 24px rgba(0,0,0,0.4) !important;
        }}
        [data-testid="stStatusWidget"] summary {{
            color: #E2E8F0 !important;
            font-weight: 600 !important;
        }}
        [data-testid="stStatusWidget"] p {{
            color: #CBD5E1 !important;
            font-size: 13.5px !important;
            font-weight: 500 !important;
            line-height: 1.45 !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def inject_pin_to_question_scroll() -> None:
    """ChatGPT-style scroll: pin the LATEST .user-bubble-wrap to the top of
    the viewport and bow out the moment the user shows scroll intent.

    Reusable across Chat Advisor, Cash Flow Forecast chat, Market Intel chat,
    or any other page that uses the .user-bubble-wrap right-aligned bubble
    pattern. Safe to call multiple times per render — only the most recent
    invocation is meaningful.
    """
    components.html(
        """
        <script>
        (function() {
            var win = window.parent;
            var doc = win.document;
            function pinToQuestion() {
                var nodes = doc.querySelectorAll('.user-bubble-wrap');
                if (!nodes.length) return;
                var el = nodes[nodes.length - 1];
                try {
                    el.scrollIntoView({ behavior: 'auto', block: 'start', inline: 'nearest' });
                } catch (e) {
                    try { el.scrollIntoView(true); } catch (e2) {}
                }
            }
            var cancelled = false;
            ['wheel', 'touchmove', 'keydown', 'mousedown'].forEach(function(ev) {
                win.addEventListener(ev, function() { cancelled = true; }, { passive: true, once: true });
            });
            [0, 80, 220, 480, 800, 1300].forEach(function(d) {
                win.setTimeout(function() { if (!cancelled) pinToQuestion(); }, d);
            });
        })();
        </script>
        """,
        height=0,
        width=0,
    )


def inject_navigation_loader() -> None:
    """Full-screen loader when switching sidebar pages (injected into parent document)."""
    components.html(
        """
        <script>
        (function () {
            var root;
            try {
                root = window.parent;
                if (!root || !root.document) {
                    return;
                }
            } catch (e) {
                return;
            }
            var doc = root.document;
            // Versioned flag — bump suffix when loader markup/CSS changes so
            // existing sessions pick up the new design instead of being stuck
            // on the previous one cached on window.parent.
            var LOADER_VERSION = "duka-nav-loader-v3";
            if (root.__dukaNavLoaderVersion === LOADER_VERSION) {
                return;
            }
            // Old version present — purge stale overlay/style before re-init.
            try {
                var staleOverlay = doc.getElementById("duka-nav-loader-overlay");
                if (staleOverlay && staleOverlay.parentNode) {
                    staleOverlay.parentNode.removeChild(staleOverlay);
                }
                var staleStyle = doc.getElementById("duka-nav-loader-style");
                if (staleStyle && staleStyle.parentNode) {
                    staleStyle.parentNode.removeChild(staleStyle);
                }
            } catch (e) {}
            root.__dukaNavLoaderVersion = LOADER_VERSION;

            var css = doc.createElement("style");
            css.id = "duka-nav-loader-style";
            css.textContent =
                "#duka-nav-loader-overlay{" +
                "display:none;position:fixed;inset:0;z-index:2147483646;" +
                "background:radial-gradient(circle at 50% 45%, rgba(29,158,117,0.18) 0%, rgba(13,17,23,0.92) 60%);" +
                "backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);" +
                "flex-direction:column;align-items:center;justify-content:center;gap:1.1rem;" +
                "pointer-events:none;animation:dukaNavFadeIn 0.18s ease-out;" +
                "}" +
                "@keyframes dukaNavFadeIn{from{opacity:0;}to{opacity:1;}}" +
                "#duka-nav-loader-overlay.duka-nav-visible{display:flex !important;}" +
                ".duka-nav-loader-card{display:flex;flex-direction:column;align-items:center;" +
                "gap:0.85rem;padding:1.6rem 2.4rem 1.5rem;border-radius:22px;" +
                "background:rgba(15,23,42,0.7);border:1px solid rgba(29,158,117,0.32);" +
                "box-shadow:0 20px 50px rgba(0,0,0,0.45),inset 0 1px 0 rgba(255,255,255,0.04);}" +
                ".duka-nav-loader-orb{position:relative;width:64px;height:64px;}" +
                ".duka-nav-loader-orb::before,.duka-nav-loader-orb::after{" +
                "content:'';position:absolute;inset:0;border-radius:50%;" +
                "border:3px solid transparent;}" +
                ".duka-nav-loader-orb::before{" +
                "border-top-color:#1D9E75;border-right-color:rgba(29,158,117,0.4);" +
                "animation:dukaNavSpin 0.95s cubic-bezier(0.5,0,0.5,1) infinite;}" +
                ".duka-nav-loader-orb::after{" +
                "border-bottom-color:#378ADD;border-left-color:rgba(55,138,221,0.35);" +
                "animation:dukaNavSpin 1.4s cubic-bezier(0.5,0,0.5,1) infinite reverse;" +
                "inset:8px;}" +
                ".duka-nav-loader-mark{position:absolute;inset:0;display:flex;" +
                "align-items:center;justify-content:center;font-size:1.4rem;" +
                "background:linear-gradient(135deg,#1D9E75,#378ADD);" +
                "-webkit-background-clip:text;-webkit-text-fill-color:transparent;" +
                "background-clip:text;font-weight:900;}" +
                "@keyframes dukaNavSpin{to{transform:rotate(360deg);}}" +
                ".duka-nav-loader-dots{display:flex;gap:8px;}" +
                ".duka-nav-loader-dots span{width:7px;height:7px;border-radius:50%;" +
                "background:linear-gradient(135deg,#378ADD,#1D9E75);" +
                "animation:dukaNavDot 1.2s ease-in-out infinite;opacity:0.45;" +
                "box-shadow:0 0 8px rgba(29,158,117,0.4);}" +
                ".duka-nav-loader-dots span:nth-child(2){animation-delay:0.2s;}" +
                ".duka-nav-loader-dots span:nth-child(3){animation-delay:0.4s;}" +
                "@keyframes dukaNavDot{" +
                "0%,100%{opacity:0.35;transform:scale(0.85);}" +
                "50%{opacity:1;transform:scale(1.1);}" +
                "}" +
                ".duka-nav-loader-label{font-size:0.72rem;font-weight:800;" +
                "letter-spacing:0.24em;text-transform:uppercase;" +
                "background:linear-gradient(90deg,#A7F3D0,#FFFFFF,#A7F3D0);" +
                "-webkit-background-clip:text;-webkit-text-fill-color:transparent;" +
                "background-clip:text;background-size:200% 100%;" +
                "animation:dukaNavShine 2.4s linear infinite;}" +
                "@keyframes dukaNavShine{from{background-position:200% 0;}to{background-position:-200% 0;}}" +
                ".duka-nav-loader-sub{font-size:0.78rem;color:rgba(148,163,184,0.92);" +
                "font-weight:500;margin-top:-0.1rem;letter-spacing:0.02em;}";
            doc.head.appendChild(css);

            var overlay = doc.getElementById("duka-nav-loader-overlay");
            if (!overlay) {
                overlay = doc.createElement("div");
                overlay.id = "duka-nav-loader-overlay";
                overlay.setAttribute("aria-live", "polite");
                overlay.innerHTML =
                    '<div class="duka-nav-loader-card">' +
                    '<div class="duka-nav-loader-orb">' +
                    '<div class="duka-nav-loader-mark">D</div>' +
                    '</div>' +
                    '<div class="duka-nav-loader-dots">' +
                    "<span></span><span></span><span></span></div>" +
                    '<div class="duka-nav-loader-label">Loading workspace</div>' +
                    '<div class="duka-nav-loader-sub">Preparing your view…</div>' +
                    '</div>';
                doc.body.appendChild(overlay);
            }

            var shownAt = 0;
            var debounceTimer = null;

            function showLoader() {
                overlay.classList.add("duka-nav-visible");
                overlay.setAttribute("aria-busy", "true");
                shownAt = Date.now();
            }

            function hideLoader() {
                overlay.classList.remove("duka-nav-visible");
                overlay.setAttribute("aria-busy", "false");
            }

            function scheduleHideAfterNav() {
                if (!overlay.classList.contains("duka-nav-visible")) {
                    return;
                }
                clearTimeout(debounceTimer);
                debounceTimer = setTimeout(function () {
                    var elapsed = Date.now() - shownAt;
                    var extra = elapsed < 320 ? 320 - elapsed : 0;
                    setTimeout(hideLoader, extra);
                }, 160);
            }

            doc.addEventListener(
                "click",
                function (ev) {
                    var btn = ev.target.closest("button");
                    if (!btn) {
                        return;
                    }
                    var sb = btn.closest('[data-testid="stSidebar"]');
                    if (!sb) {
                        return;
                    }
                    showLoader();
                },
                true
            );

            var observeRoot =
                doc.querySelector('[data-testid="stAppViewContainer"]') || doc.body;
            try {
                new MutationObserver(function () {
                    scheduleHideAfterNav();
                }).observe(observeRoot, { childList: true, subtree: true });
            } catch (e) {}

            hideLoader();
        })();
        </script>
        """,
        height=0,
        width=0,
    )


def main() -> None:
    st.set_page_config(
        page_title="Duka AI",
        page_icon=":material/monitoring:",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    apply_custom_styles()
    inject_navigation_loader()

    settings = get_settings()
    sample_cases = load_sample_cases()
    init_session_state()

    from tools.scheduler import start_scheduler
    start_scheduler()

    defaults = {
        "business_type": "",
        "location": "",
        "products_services": "",
        "main_question": "",
        "manual_notes": "",
        "manual_revenue": 0.0,
        "manual_expenses": 0.0,
        "manual_debt": 0.0,
        "manual_staff": 0,
        "document_type": "Income Statement",
        "numbers_entry_mode": "upload",
        "selected_example_prompt": "",
        "selected_example_prompt_text": "",
        "scroll_target": None,
        "connected_provider": None,
        "selected_providers": set(),
        "analysis_report": None,
        "messages": [],
        "last_agent": "advisor",
        "hide_example_prompts": False,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)

    # ── Navigation ─────────────────────────────────────────────────────────────
    active_page = render_sidebar_navigation(settings)
    if active_page != "Chat Advisor":
        render_selected_page(active_page, settings)
        return

    # ── Welcome screen (shown on first load and after reset) ───────────────────
    if st.session_state.get("show_welcome") and not st.session_state.get("analysis_report"):
        render_welcome_screen(sample_cases)
        return

    if st.session_state.get("analysis_report"):
        st.session_state.pop("pending_demo_analysis_run", None)

    st.markdown(
        """
        <div style="text-align:center;padding:8px 0 4px;">
            <span class="hero-badge">
                &#9889; AI-powered &nbsp;&middot;&nbsp; &#127758; Built for Africa
            </span>
        </div>
        <div style="text-align:center;padding:16px 0 8px;">
            <h1 style="font-size:2.2rem;font-weight:800;color:#FFFFFF;
                       margin:0;line-height:1.2;">
                Your <span style="color:#1D9E75;">AI financial advisor</span>
                <br>for African businesses
            </h1>
            <p style="color:#6B7280;font-size:15px;margin:12px auto 0;
                      max-width:480px;line-height:1.6;">
                Analyze cash flow, loan readiness, and market conditions.
                Talk to your finances like a conversation, not a spreadsheet.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if st.session_state.pop("scroll_to_demo_analyze", False):
        components.html(
            """
            <script>
            window.setTimeout(function() {
                var el = window.parent.document.getElementById("duka-ai-analyze-anchor");
                if (el) { el.scrollIntoView({ behavior: "smooth", block: "start" }); }
            }, 120);
            </script>
            """,
            height=0,
            width=0,
        )

    if (
        st.session_state.get("pending_demo_analysis_run")
        and not st.session_state.get("analysis_report")
    ):
        with st.status("⚡ Demo analysis in progress…", expanded=True):
            st.write("📂 Sample income statement loaded · **Lusaka Grocery Shop** profile applied.")
            st.write("🧠 Starting all agents next — detailed steps appear in the panel below.")
            st.caption("Stay on this page — the workspace scrolls to **Run Full Business Analysis** automatically.")

    # ── Powered-by indicator ───────────────────────────────────────────────────
    _provider = os.getenv("LLM_PROVIDER", "amd").strip().upper()
    _model_display = (
        "AMD MI300X"
        if _provider == "AMD"
        else ("OpenAI-compatible" if _provider == "OPENAI" else _provider or "AMD MI300X")
    )
    st.markdown(
        f"""
        <div style="display:flex;align-items:center;justify-content:center;
                    gap:8px;padding:6px 0;margin-bottom:12px;">
            <div style="width:8px;height:8px;border-radius:50%;background:#1D9E75;
                        box-shadow:0 0 6px #1D9E75;"></div>
            <span style="color:#4A5568;font-size:12px;">
                Powered by Duka AI &nbsp;&middot;&nbsp; {_model_display}
                &nbsp;&middot;&nbsp; Talk to your finances like a conversation
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Pre-fill banner & scroll (welcome sample uses numbers anchor + hides duplicate example pills)
    _did_prefill = st.session_state.pop("form_prefilled", False)
    _scroll_numbers = st.session_state.pop("scroll_to_numbers", False)
    if _did_prefill:
        if st.session_state.get("hide_example_prompts"):
            if st.session_state.get("pending_demo_analysis_run"):
                st.success(
                    "✅ **Lusaka Grocery Shop** demo loaded — sample income statement + figures are applied. "
                    "**Running full analysis automatically** — watch the progress panel below."
                )
            else:
                st.success(
                    "✅ **Lusaka Grocery Shop** sample loaded — your story and figures are filled in above. "
                    "Upload a file below (optional) or jump to **Manual numbers**, then click **Run Full Business Analysis**."
                )
        else:
            st.success(
                "✅ Sample business loaded — review the details below and click **Run Full Business Analysis**"
            )
        if st.session_state.get("pending_demo_analysis_run"):
            _scroll_id = "duka-ai-analyze-anchor"
        else:
            _scroll_id = "duka-ai-numbers-anchor" if _scroll_numbers else "duka-ai-form-anchor"
        components.html(
            f"""
            <script>
            window.setTimeout(function() {{
                var anchor = window.parent.document.getElementById("{_scroll_id}");
                if (anchor) {{ anchor.scrollIntoView({{ behavior: "smooth", block: "start" }}); }}
            }}, 220);
            </script>
            """,
            height=0,
            width=0,
        )
    elif _scroll_numbers:
        components.html(
            """
            <script>
            window.setTimeout(function() {
                var anchor = window.parent.document.getElementById("duka-ai-numbers-anchor");
                if (anchor) { anchor.scrollIntoView({ behavior: "smooth", block: "start" }); }
            }, 180);
            </script>
            """,
            height=0,
            width=0,
        )

    has_existing_analysis = bool(st.session_state.get("analysis_report"))
    if not has_existing_analysis:
        # ── Section 1: Business Context ────────────────────────────────────────────
        st.markdown('<div id="duka-ai-form-anchor"></div>', unsafe_allow_html=True)
        render_section_intro(
            "Tell us about your business",
            "A few details help Duka AI tailor its analysis to your situation.",
        )
        context_col_1, context_col_2, context_col_3, context_col_4 = st.columns(4, gap="medium")
        with context_col_1:
            st.session_state.business_type = st.text_input(
                "Business type",
                value=st.session_state.business_type,
                placeholder="Grocery shop",
                help="What kind of business do you run?",
            )
        with context_col_2:
            st.session_state.location = st.text_input(
                "Location",
                value=st.session_state.location,
                placeholder="Lusaka",
                help="City or area where your business operates",
            )
        with context_col_3:
            st.session_state.products_services = st.text_input(
                "Main products / services",
                value=st.session_state.products_services,
                placeholder="Groceries, mealie meal, drinks",
                help="What do you sell or offer?",
            )
        with context_col_4:
            st.session_state.main_question = st.text_input(
                "Main question or concern",
                value=st.session_state.main_question,
                placeholder="Should I restock or save cash?",
                help="What do you want Duka AI to help with?",
            )

        st.session_state.manual_notes = st.text_area(
            "Additional notes (optional)",
            value=st.session_state.manual_notes,
            placeholder="I run a small grocery shop in Lusaka. I want to know whether to restock or save cash. This week I made K4,500 in sales...",
            help="Add any extra details, financial numbers, debt, or business notes.",
        )

        st.markdown('<div style="margin-top:1.5rem;"></div>', unsafe_allow_html=True)
        continue_to_data = st.button(
            "↓ Continue to data sources",
            key="continue_to_data_sources",
            use_container_width=True,
            help="Jump to the upload/connect/manual section",
        )
        st.markdown(
            '<div class="analyze-note">Want higher accuracy? Add a document or figures in the section below first.</div>',
            unsafe_allow_html=True,
        )
        if continue_to_data:
            components.html(
                """
                <script>
                window.setTimeout(function() {
                    var anchor = window.parent.document.getElementById("duka-ai-numbers-anchor");
                    if (anchor) { anchor.scrollIntoView({ behavior: "smooth", block: "start" }); }
                }, 100);
                </script>
                """,
                height=0,
                width=0,
            )

        st.divider()
        if not st.session_state.get("hide_example_prompts"):
            st.markdown('<div class="subsection-label">Example prompts</div>', unsafe_allow_html=True)
            render_example_prompt_pills(sample_cases)

        # ── Trust strip ────────────────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        _trust_items = [
            ("🔒", "Bank-grade security", "Your data stays private"),
            ("🤖", "4 AI agents", "Working together for you"),
            ("🌍", "Built for Africa", "Zambian market context"),
            ("⚡", "AMD MI300X", "Enterprise-grade GPU power"),
        ]
        for _col, (_icon, _title, _sub) in zip(st.columns(4), _trust_items):
            with _col:
                st.markdown(
                    f"""
                    <div class="trust-item">
                        <span>{_icon}</span>
                        <strong style="color:#C9D1D9;font-size:12px;">{_title}</strong>
                        <br>
                        <span style="color:#4A5568;font-size:11px;">{_sub}</span>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
        st.markdown("<br>", unsafe_allow_html=True)
        st.divider()

        # ── Section 2: Data Sources ────────────────────────────────────────────────
        st.markdown('<div id="duka-ai-numbers-anchor"></div>', unsafe_allow_html=True)
        render_section_intro(
            "📎 Want more precise analysis? Add your numbers",
            "Choose one path below. Any extra numbers help Duka AI give tighter, more specific advice.",
        )
        uploaded_analysis: dict[str, Any] | None = None
        upload_error: str | None = None
        option_cols = st.columns(3, gap="small")
        for column, (label, mode_value) in zip(option_cols, NUMBER_SOURCE_OPTIONS):
            with column:
                if st.button(label, key=f"numbers_mode_{mode_value}"):
                    st.session_state.numbers_entry_mode = mode_value
                    st.rerun()

        selected_mode = st.session_state.numbers_entry_mode
        st.markdown(
            '<div class="numbers-helper">You can switch between options at any time before you analyze.</div>',
            unsafe_allow_html=True,
        )

        if selected_mode == "manual":
            st.markdown('<div class="mode-helper">Enter only the figures you know. Even partial numbers make the analysis more precise.</div>', unsafe_allow_html=True)
            num_col1, num_col2, num_col3, num_col4 = st.columns(4, gap="medium")
            with num_col1:
                st.session_state.manual_revenue = st.number_input(
                    "Weekly or monthly revenue (K)",
                    min_value=0.0, value=float(st.session_state.manual_revenue), step=100.0, format="%.0f",
                    help="Your typical sales or income per week or month in Kwacha",
                )
            with num_col2:
                st.session_state.manual_expenses = st.number_input(
                    "Monthly expenses (K)",
                    min_value=0.0, value=float(st.session_state.manual_expenses), step=100.0, format="%.0f",
                    help="Your total monthly costs — stock, rent, wages, etc.",
                )
            with num_col3:
                st.session_state.manual_debt = st.number_input(
                    "Outstanding debt (K)",
                    min_value=0.0, value=float(st.session_state.manual_debt), step=100.0, format="%.0f",
                    help="Any loans or supplier debt you currently owe",
                )
            with num_col4:
                st.session_state.manual_staff = st.number_input(
                    "Number of staff",
                    min_value=0, value=int(st.session_state.manual_staff), step=1,
                    help="How many people work in your business (including yourself)",
                )

        elif selected_mode == "upload":
            st.markdown('<div class="subsection-label">Upload your business document</div>', unsafe_allow_html=True)
            st.markdown('<div class="mode-helper">Select the document type, upload the file, and Duka AI will extract key financial figures.</div>', unsafe_allow_html=True)
            st.caption("Accepted file types: CSV, Excel (.xlsx), TXT, PDF")
            st.caption("PDF support is experimental. For best results, upload CSV, Excel, or text records.")
            st.session_state.document_type = st.selectbox(
                "Document type",
                DOCUMENT_TYPES,
                index=DOCUMENT_TYPES.index(st.session_state.document_type) if st.session_state.document_type in DOCUMENT_TYPES else 0,
            )
            uploaded_file = st.file_uploader(
                "Upload business document",
                type=["csv", "xlsx", "txt", "pdf"],
                help="Upload an income statement, sales record, expense log, bank statement, mobile money statement, or business notes.",
            )
            if uploaded_file is not None:
                try:
                    _doc_fp = f"{uploaded_file.name}:{uploaded_file.size}"
                    _is_new_file = st.session_state.get("_doc_fingerprint") != _doc_fp
                    if _is_new_file:
                        _ext = uploaded_file.name.rsplit(".", 1)[-1].upper() if "." in uploaded_file.name else "FILE"
                        with st.status(
                            f"📂 Reading {uploaded_file.name} …",
                            expanded=True,
                        ) as _upload_status:
                            st.write(f"🔍 Detecting document structure ({_ext} format)…")
                            st.write("📊 Extracting financial figures…")
                            uploaded_analysis = parse_uploaded_file(
                                uploaded_file, st.session_state.document_type
                            )
                            _det_type = uploaded_analysis.get("document_type", "document")
                            _conf     = uploaded_analysis.get("confidence", "")
                            _conf_str = f" · {_conf} confidence" if _conf else ""
                            _upload_status.update(
                                label=f"✅ {_det_type} loaded{_conf_str}",
                                state="complete",
                                expanded=False,
                            )
                    else:
                        uploaded_analysis = parse_uploaded_file(
                            uploaded_file, st.session_state.document_type
                        )

                    # Replace stale data only when a genuinely new file is detected.
                    if _is_new_file:
                        # ── Wipe ALL previous financial state ────────────────────────
                        for _k in (
                            "baseline", "metrics", "parsed_document",
                            "manual_revenue", "manual_expenses", "manual_profit",
                            "document_revenue", "document_expenses",
                            "generated_report", "analysis_report",
                        ):
                            st.session_state.pop(_k, None)
                        for _k in [k for k in st.session_state if k.startswith("forecast_summary_")]:
                            del st.session_state[_k]
                        st.session_state["forecast_chat"] = []

                        # ── Load fresh data from the new document ────────────────────
                        st.session_state["parsed_document"] = uploaded_analysis
                        doc_rev = float(uploaded_analysis.get("revenue")  or 0)
                        doc_exp = float(uploaded_analysis.get("expenses") or 0)
                        if doc_rev > 0 or doc_exp > 0:
                            st.session_state["manual_revenue"] = doc_rev
                            st.session_state["manual_expenses"] = doc_exp
                            new_baseline = {
                                "monthly_revenue":  doc_rev,
                                "monthly_expenses": doc_exp,
                                "monthly_profit":   doc_rev - doc_exp,
                                "months_of_data":   int(uploaded_analysis.get("months_of_data") or 1),
                                "source": (
                                    f"Uploaded {uploaded_analysis.get('document_type', 'document')}: "
                                    f"{uploaded_file.name}"
                                ),
                                "expenses_breakdown": uploaded_analysis.get("expenses_breakdown", {}),
                            }
                            st.session_state["baseline"] = new_baseline
                            st.session_state["metrics"]  = compute_metrics(new_baseline)

                        st.session_state["_doc_fingerprint"] = _doc_fp

                    st.markdown('<div id="duka-ai-upload-summary-anchor"></div>', unsafe_allow_html=True)
                    render_file_preview(uploaded_analysis)
                    if st.session_state.get("analysis_report"):
                        st.info("Document updated. Click **Re-run Analysis with Current Data** below to refresh all agent insights.")
                    else:
                        st.info("Document extracted successfully. Next step: click **Run Full Business Analysis** below.")
                    if _is_new_file:
                        components.html(
                            """
                            <script>
                            window.setTimeout(function() {
                                var anchor = window.parent.document.getElementById("duka-ai-upload-summary-anchor");
                                if (anchor) { anchor.scrollIntoView({ behavior: "smooth", block: "start" }); }
                            }, 120);
                            </script>
                            """,
                            height=0,
                            width=0,
                        )
                except Exception as exc:
                    upload_error = f"Could not parse the uploaded file. {exc}"
                    st.error(upload_error)
            with st.expander("Need a sample template?", expanded=False):
                st.markdown('<div class="mode-helper">Use one of these starter files if you want a clean format to upload.</div>', unsafe_allow_html=True)
                render_download_buttons()

        else:
            st.markdown(
                '<div style="display:inline-flex;align-items:center;gap:0.5rem;background:#7C2D12;border:1px solid #F97316;'
                'border-radius:999px;padding:0.35rem 0.85rem;margin-bottom:0.75rem;">'
                '<span style="font-size:0.78rem;font-weight:800;color:#FED7AA;text-transform:uppercase;letter-spacing:0.1em;">'
                '⚠ Demo Mode — no real accounts connected</span></div>',
                unsafe_allow_html=True,
            )
            st.markdown('<div class="subsection-label">Connect a demo account</div>', unsafe_allow_html=True)
            st.markdown('<div class="mode-helper">This is a safe demo flow. Pick a provider to preview how synced account data would improve the analysis.</div>', unsafe_allow_html=True)
            render_clickable_provider_cards()
            connected_preview = get_connected_demo(st.session_state.connected_provider)
            if connected_preview:
                render_connected_preview(connected_preview)

        # ── Analyze Button (bottom — after optional data sources) ─────────────────
        st.divider()
        st.markdown('<div id="duka-ai-analyze-anchor"></div>', unsafe_allow_html=True)
        has_existing_analysis = bool(st.session_state.get("analysis_report"))
        analyze_clicked = False
        if not has_existing_analysis:
            analyze_clicked = st.button(
                "⚡ Run Full Business Analysis",
                key="analyze_main",
                type="primary",
                use_container_width=True,
            )
            st.markdown(
                '<div class="analyze-note">Runs all agents (cash flow, advisor, loan readiness, market, and summary).</div>',
                unsafe_allow_html=True,
            )
            # Auto-trigger when demo flow set the flag (judge-friendly one-click path).
            if not analyze_clicked and st.session_state.pop("auto_run_demo", False):
                analyze_clicked = True
                st.info("✨ Demo data loaded — running full analysis automatically…")

        # Allow downstream analyze branch to use the previously parsed sample
        # document when no fresh st.file_uploader event has occurred.
        if uploaded_analysis is None and st.session_state.get("parsed_document"):
            uploaded_analysis = st.session_state.get("parsed_document")

        if analyze_clicked:
            # Immediately scroll to the status panel so user sees agents working
            components.html(
                """
                <script>
                window.setTimeout(function() {
                    var el = window.parent.document.getElementById("duka-ai-analyze-anchor");
                    if (el) { el.scrollIntoView({ behavior: "smooth", block: "start" }); }
                }, 80);
                </script>
                """,
                height=0,
                width=0,
            )
            connected_analysis = get_connected_demo(st.session_state.connected_provider)
            manual_figures = {
                "revenue": float(st.session_state.manual_revenue),
                "expenses": float(st.session_state.manual_expenses),
                "debt": float(st.session_state.manual_debt),
                "staff": int(st.session_state.manual_staff),
            }
            has_manual_figures = any(float(value or 0) > 0 for value in manual_figures.values())
            business_profile = {
                "business_type": st.session_state.business_type,
                "location": st.session_state.location,
                "products_services": st.session_state.products_services,
                "main_question": st.session_state.main_question,
            }

            if (
                not any(value.strip() for value in business_profile.values())
                and not st.session_state.manual_notes.strip()
                and not has_manual_figures
                and uploaded_analysis is None
                and connected_analysis is None
            ):
                st.warning("Please provide business context, upload a document, connect demo transaction data, or combine them before running the analysis.")
            elif upload_error and not st.session_state.manual_notes.strip() and not has_manual_figures and connected_analysis is None:
                st.error("The uploaded file could not be analyzed, so there is not enough data to continue.")
            else:
                analysis_context, combined_parsed_data, basis_notes, source_labels, data_sources = combine_analysis_inputs(
                    business_profile,
                    st.session_state.manual_notes,
                    uploaded_analysis,
                    connected_analysis,
                    manual_figures=manual_figures,
                )

                with st.status("🧠 Duka AI — analyzing your business…", expanded=True) as status:
                    pre_steps = sum([
                        1 if any(value.strip() for value in business_profile.values()) else 0,
                        1 if st.session_state.manual_notes.strip() else 0,
                        1 if has_manual_figures else 0,
                        1 if uploaded_analysis else 0,
                        1 if connected_analysis else 0,
                    ])
                    total_steps = pre_steps + 5  # 5 downstream agents incl. executive summary
                    step_index = 0
                    progress = st.progress(0)

                    def _step(msg: str) -> None:
                        nonlocal step_index
                        step_index += 1
                        st.write(f"{step_index}/{total_steps} • {msg}")
                        progress.progress(min(step_index / total_steps, 1.0))

                    if any(value.strip() for value in business_profile.values()):
                        _step("📋 Reading your business context and profile…")
                    if st.session_state.manual_notes.strip():
                        _step("📝 Parsing your written business description…")
                    if has_manual_figures:
                        _step("📎 Locking in the numbers you entered…")
                    if uploaded_analysis:
                        _step(f"📄 Interpreting uploaded {uploaded_analysis.get('document_type', 'document').lower()}…")
                    if connected_analysis:
                        _step(f"🔗 Reading {connected_analysis.get('provider_selected', 'connected')} transaction data…")

                    completed_agents: set[str] = set()
                    def _on_agent_progress(agent_key: str) -> None:
                        rev = combined_parsed_data.get("revenue", 0) or 0
                        exp = combined_parsed_data.get("expenses", 0) or 0
                        profit = round(rev - exp, 0)
                        dynamic_msgs: dict[str, str] = {
                            "cashflow": (
                                f"📊 Cash Flow Agent — analyzing {format_currency(rev)} revenue "
                                f"and {format_currency(exp)} expenses."
                            ),
                            "advisor": (
                                f"💡 Business Advisor Agent — building recommendations "
                                f"from {format_currency(profit)} monthly profit."
                            ),
                            "loan": (
                                f"🏦 Loan Readiness Agent — calculating safe borrowing capacity "
                                f"from {format_currency(profit)} profit and {format_currency(combined_parsed_data.get('debt', 0) or 0)} existing debt."
                            ),
                            "market": (
                                "📈 Market Intelligence Agent — checking Zambian market trends "
                                f"for {combined_parsed_data.get('business_type') or 'your business type'}."
                            ),
                            "summary": (
                                "🔄 Executive Summary Agent — writing your personalized "
                                f"financial summary for {business_profile.get('location') or 'your business'}."
                            ),
                        }
                        msg = dynamic_msgs.get(agent_key)
                        if msg:
                            if agent_key not in completed_agents:
                                completed_agents.add(agent_key)
                                _step(msg)
                        else:
                            icon, name, action = AGENT_STATUS_MESSAGES.get(agent_key, ("⚙️", agent_key, "Running..."))
                            if agent_key not in completed_agents:
                                completed_agents.add(agent_key)
                                _step(f"{icon} {name} — {action}")

                    report = generate_business_report(
                        analysis_context,
                        parsed_data_override=combined_parsed_data,
                        analysis_basis=basis_notes,
                        source_labels=source_labels,
                        transaction_summary=connected_analysis,
                        business_profile=business_profile,
                        document_analysis=uploaded_analysis,
                        data_sources=data_sources,
                        progress_callback=_on_agent_progress,
                    )
                    progress.progress(1.0)
                    status.update(label="✅ Analysis complete! Scroll down to chat with your results.", state="complete", expanded=False)

                st.session_state.analysis_report = report
                st.session_state.pop("pending_demo_analysis_run", None)
                _baseline = get_baseline(st.session_state)
                st.session_state["baseline"] = _baseline
                st.session_state["metrics"] = compute_metrics(_baseline) if _baseline else None
                st.session_state.messages = [
                    {
                        "role": "assistant",
                        "content": report["final_summary"],
                        "type": "initial_analysis",
                        "report_key": True,
                    }
                ]
                st.session_state["scroll_to_results"] = True
                st.rerun()

    # ── Conversational Chat Interface ──────────────────────────────────────────
    report = st.session_state.get("analysis_report")
    messages: list[dict] = st.session_state.get("messages", [])
    scroll_requested = st.session_state.get("scroll_target") == "duka-ai-analyze-anchor"
    inject_ui_behavior(
        sample_cases=sample_cases,
        selected_example_prompt=st.session_state.get("selected_example_prompt", ""),
        selected_example_text=st.session_state.get("selected_example_prompt_text", ""),
        selected_numbers_mode=st.session_state.get("numbers_entry_mode", "upload"),
        scroll_requested=scroll_requested,
    )
    if scroll_requested:
        st.session_state.scroll_target = None

    if not messages:
        st.markdown(
            '<div class="chat-placeholder">Your chat workspace opens here after the first analysis. Fill in the form above, run the analysis, then ask follow-up questions about cash flow, borrowing, or growth.</div>',
            unsafe_allow_html=True,
        )
        return

    # ── Auto-scroll to results after analysis ──────────────────────────────────
    st.markdown('<div id="duka-ai-chat-anchor"></div>', unsafe_allow_html=True)
    if st.session_state.pop("scroll_to_results", False):
        components.html(
            """
            <script>
            window.setTimeout(function() {
                var el = window.parent.document.getElementById("duka-ai-chat-anchor");
                if (el) {
                    el.scrollIntoView({ behavior: "smooth", block: "start" });
                }
            }, 400);
            </script>
            """,
            height=0,
            width=0,
        )

    stage_col, reset_col = st.columns([4, 1], gap="medium")
    with stage_col:
        st.markdown(
            """
            <div class="chat-stage-header">
                <div class="chat-stage-title">Chat with your analysis</div>
                <div class="chat-stage-copy">Start with the summary below, tap a suggested question, or type your own follow-up when you're ready.</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with reset_col:
        if st.button("↺ Reset / Start New Analysis", key="reset_analysis", use_container_width=True, help="Clear this analysis and start fresh"):
            for _k in ["analysis_report", "baseline", "metrics"]:
                st.session_state[_k] = None
            st.session_state.messages = []
            st.session_state.manual_notes = ""
            st.session_state.manual_revenue = 0.0
            st.session_state.manual_expenses = 0.0
            st.session_state.manual_debt = 0.0
            st.session_state.manual_staff = 0
            st.session_state.business_type = ""
            st.session_state.location = ""
            st.session_state.products_services = ""
            st.session_state.main_question = ""
            st.session_state.selected_example_prompt = ""
            st.session_state.selected_example_prompt_text = ""
            st.session_state.hide_example_prompts = False
            st.session_state.show_welcome = True
            st.rerun()
    st.markdown('<div class="chat-block-gap"></div>', unsafe_allow_html=True)

    # Render all chat messages
    for msg in messages:
        if msg["role"] == "user":
            # Right-aligned bubble — no Streamlit chat component needed
            escaped = msg["content"].replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            st.markdown(
                f'<div class="user-bubble-wrap"><div class="user-bubble">{escaped}</div></div>',
                unsafe_allow_html=True,
            )
        else:
            agent_key = msg.get("agent", "")
            avatar_icon = AGENT_LABELS.get(agent_key, ("✨", ""))[0] if agent_key else "✨"
            with st.chat_message("assistant", avatar=avatar_icon):
                if msg.get("type") == "initial_analysis" and report:
                    render_initial_analysis_in_chat(report)
                else:
                    if agent_key and agent_key in AGENT_LABELS:
                        icon, label = AGENT_LABELS[agent_key]
                        st.markdown(
                            f'<div class="agent-label">{icon} {label}</div>',
                            unsafe_allow_html=True,
                        )
                    st.markdown(msg.get("content", ""), unsafe_allow_html=True)
                    if msg.get("visual"):
                        render_followup_visual(msg["visual"])
                    if msg.get("chart"):
                        render_followup_chart(msg["chart"])

    # ── Suggestion pills — above input ──────────────────────────────────
    if report and len(messages) == 1:
        st.markdown('<div class="chat-block-gap"></div>', unsafe_allow_html=True)
        sug_cols = st.columns(3, gap="small")
        for sug_i, (suggestion_label, suggestion_prompt) in enumerate(FOLLOWUP_SUGGESTION_PILLS):
            with sug_cols[sug_i % 3]:
                if st.button(suggestion_label, key=f"sug_{sug_i}", use_container_width=True):
                    agents_to_run = _resolve_agents(suggestion_prompt, messages)
                    primary_agent = agents_to_run[0]
                    st.session_state.messages.append({"role": "user", "content": suggestion_prompt})
                    with st.spinner(f"💬 {suggestion_label} …"):
                        result = answer_followup_question(suggestion_prompt, report, history=None)
                        visual = build_followup_visual(suggestion_prompt, report)
                        chart = build_chart_data(suggestion_prompt, report, st.session_state.messages)
                    followup_msg: dict[str, Any] = {
                        "role": "assistant",
                        "content": _normalize_assistant_response(result["answer"]),
                        "agent": primary_agent,
                    }
                    if visual:
                        followup_msg["visual"] = visual
                    if chart:
                        followup_msg["chart"] = chart
                    st.session_state.messages.append(followup_msg)
                    st.session_state.last_agent = primary_agent
                    st.session_state.scroll_to_chat_conversation = True
                    st.rerun()

    # ── Chat input with streaming ──────────────────────────────────────────────
    if report:
        st.markdown('<div class="chat-block-gap"></div>', unsafe_allow_html=True)
        _injected = st.session_state.pop("pending_chat_question", None)
        if prompt := (_injected or st.chat_input("Ask about your business...")):
            agents_to_run = _resolve_agents(prompt, messages)
            primary_agent = agents_to_run[0]
            icon, label = AGENT_LABELS.get(primary_agent, ("💡", "Business Advisor"))

            # Right-aligned user bubble (consistent with message loop)
            escaped_prompt = prompt.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            st.markdown(
                f'<div class="user-bubble-wrap"><div class="user-bubble">{escaped_prompt}</div></div>',
                unsafe_allow_html=True,
            )
            st.session_state.messages.append({"role": "user", "content": prompt})

            # ChatGPT-style anchoring: pin the user's question at the top of
            # the viewport so the streaming response fills the screen below it.
            # Hammered instant jumps in the first ~1.3s overpower Streamlit's
            # scroll-restore on rerun. After that, the script BOWS OUT and lets
            # the user scroll freely — never yanks them back up.
            components.html(
                """
                <script>
                (function() {
                    var win = window.parent;
                    var doc = win.document;
                    function pinToQuestion() {
                        var nodes = doc.querySelectorAll('.user-bubble-wrap');
                        if (!nodes.length) return;
                        var el = nodes[nodes.length - 1];
                        try {
                            el.scrollIntoView({
                                behavior: 'auto',
                                block: 'start',
                                inline: 'nearest'
                            });
                        } catch (e) {
                            try { el.scrollIntoView(true); } catch (e2) {}
                        }
                    }
                    // Cancel any further pins as soon as user shows scroll intent
                    // (wheel, touch, or arrow keys). This guarantees we never
                    // fight the user's manual scroll.
                    var cancelled = false;
                    function cancel() { cancelled = true; }
                    ['wheel', 'touchmove', 'keydown', 'mousedown'].forEach(function(ev) {
                        win.addEventListener(ev, cancel, { passive: true, once: true });
                    });
                    // Hammered instant pins to land us on the question quickly.
                    [0, 60, 150, 300, 550, 900, 1300].forEach(function(d) {
                        win.setTimeout(function() {
                            if (!cancelled) pinToQuestion();
                        }, d);
                    });
                })();
                </script>
                """,
                height=0,
                width=0,
            )

            with st.chat_message("assistant", avatar=icon):
                st.markdown(
                    f'<div class="agent-label">{icon} {label}</div>',
                    unsafe_allow_html=True,
                )

                if settings.llm_enabled:
                    # ChatGPT-style: stream tokens into the UI with typewriter effect + auto-follow scroll
                    stream_gen = stream_followup_for_agent(
                        primary_agent,
                        prompt,
                        report,
                        st.session_state.messages,
                    )
                    try:
                        full_response = st.write_stream(stream_gen, cursor="▌")
                    except TypeError:
                        # Older Streamlit versions don't support `cursor` kwarg
                        full_response = st.write_stream(stream_gen)
                    if isinstance(full_response, list):
                        full_response = "".join(str(x) for x in full_response)
                    else:
                        full_response = str(full_response or "")
                else:
                    result = answer_followup_question(prompt, report, history=None)
                    full_response = result["answer"]
                    st.markdown(full_response)

                visual = build_followup_visual(prompt, report)
                if visual:
                    render_followup_visual(visual)
                chart = build_chart_data(prompt, report, st.session_state.messages)
                if chart:
                    render_followup_chart(chart)

            followup_msg = {
                "role": "assistant",
                "content": _normalize_assistant_response(str(full_response)),
                "agent": primary_agent,
            }
            if visual:
                followup_msg["visual"] = visual
            if chart:
                followup_msg["chart"] = chart
            st.session_state.messages.append(followup_msg)
            st.session_state.last_agent = primary_agent
            st.session_state.scroll_to_chat_conversation = True

    # Anchor at bottom of chat thread (kept for compatibility with other flows
    # that may toggle scroll_to_chat_conversation).
    st.markdown('<div id="duka-ai-chat-scroll-target"></div>', unsafe_allow_html=True)
    if st.session_state.pop("scroll_to_chat_conversation", False):
        # Pin the latest user question to the top of viewport (ChatGPT-style)
        # so the response below it is visible. Bows out as soon as the user
        # shows scroll intent.
        components.html(
            """
            <script>
            (function() {
                var win = window.parent;
                var doc = win.document;
                function pinToLatestQuestion() {
                    var nodes = doc.querySelectorAll('.user-bubble-wrap');
                    if (!nodes.length) return false;
                    var el = nodes[nodes.length - 1];
                    try {
                        el.scrollIntoView({
                            behavior: 'auto',
                            block: 'start',
                            inline: 'nearest'
                        });
                    } catch (e) {
                        try { el.scrollIntoView(true); } catch (e2) {}
                    }
                    return true;
                }
                var cancelled = false;
                ['wheel', 'touchmove', 'keydown', 'mousedown'].forEach(function(ev) {
                    win.addEventListener(ev, function() { cancelled = true; }, { passive: true, once: true });
                });
                [0, 80, 220, 480, 800, 1300].forEach(function(d) {
                    win.setTimeout(function() {
                        if (!cancelled) pinToLatestQuestion();
                    }, d);
                });
            })();
            </script>
            """,
            height=0,
            width=0,
        )


if __name__ == "__main__":
    main()
