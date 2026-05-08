from __future__ import annotations

import io
import re
from typing import Any

import pandas as pd

from tools.financial_calculator import calculate_profit, calculate_profit_margin
from tools.text_parser import parse_business_input


DOCUMENT_TYPES = [
    "Income Statement",
    "Cash Flow Record",
    "Sales Record",
    "Expense Record",
    "Bank Statement",
    "Mobile Money Statement",
    "Receipts",
    "Inventory/Stock Record",
    "Loan or Debt Record",
    "General Business Notes",
    "Other",
]

# ── LLM system prompt for document parsing ─────────────────────────────────────

_DOC_PARSE_SYSTEM = """You are a financial document parser for African small business records.
Your job is to extract clean financial totals from raw document content — without double-counting.

CRITICAL RULES:
- Identify TOTALS vs LINE ITEMS vs SUBTOTALS. Only extract the final TOTAL revenue and TOTAL expenses.
- In multi-sheet Excel files, summary sheets often repeat totals already captured in detail sheets — do NOT add them together.
- If a document has both an Income Statement and a Cash Flow Statement, prefer the Income Statement for revenue and expenses.
- Negative numbers in cash flow statements represent outflows (expenses/payments) — treat their absolute value as expenses.
- For Income Statements: total_expenses = ALL costs combined (cost of goods + operating expenses + any other costs). Do NOT split them.
- Opening/closing balances, averages, and running totals are NOT revenue or expenses — ignore them.
- Do NOT sum all numbers blindly. Read labels carefully to understand what each number represents.
- If a value is unclear or cannot be reliably determined, return null — do NOT guess.
- Auto-detect the real document type regardless of what the user labeled it.
- Amounts are in Zambian Kwacha (K / ZMW) unless another currency is clearly stated.
- IMPORTANT: Every value in the JSON must be a plain pre-computed number. NEVER write arithmetic expressions like "18600 + 46980" — compute the sum yourself and write only the result.
- CRITICAL: Extract numbers ONLY from the DOCUMENT CONTENT provided by the user. NEVER reuse example numbers or labels from the schema/template. If the document does not contain a value, return null — do NOT invent or carry over numbers from prior conversations."""

_DOC_PARSE_USER_TEMPLATE = """Filename: {filename}
User-labeled type: {doc_type}

DOCUMENT CONTENT:
{content}

Extract financial data and return ONLY a JSON object. No markdown fences, no extra text.
All numeric fields must be plain pre-computed numbers (never expressions like "1000 + 2000").

CRITICAL: The schema below uses <PLACEHOLDER> tokens to show the SHAPE of the JSON only.
You MUST replace every <PLACEHOLDER> with a value computed from the actual DOCUMENT CONTENT above.
NEVER copy these placeholder strings or any example values into your output — extract from the document.

{{
  "detected_document_type": "<one of: Income Statement | Sales Record | Cash Flow Record | Expense Record | Bank Statement | Other>",
  "currency": "<currency code from document, e.g. K>",
  "period_description": "<human readable period from document, or null>",
  "period_start": "<YYYY-MM-DD or null>",
  "period_end": "<YYYY-MM-DD or null>",
  "total_revenue": <number computed from document, or null>,
  "total_expenses": <number computed from document, or null>,
  "profit": <number computed from document, or null>,
  "debt_balance": <number from document, or null>,
  "expense_breakdown": [
    {{"category": "<expense category name from document>", "amount": <number from document>}}
  ],
  "product_breakdown": [
    {{"product": "<product name from document>", "units_sold": <number>, "unit_price": <number>, "revenue": <number>}}
  ],
  "monthly_breakdown": [{{"month": "<month label from document>", "revenue": <number>, "expenses": <number>, "profit": <number>}}],
  "key_observations": ["<short factual observation grounded in the document>"],
  "confidence": "<high | medium | low>",
  "warnings": []
}}

Rules for detected_document_type:
- If the document has Product/Item columns with Units Sold and/or Unit Price → "Sales Record"
- If it has Revenue, Expenses, Profit rows → "Income Statement"
- If it has Cash In / Cash Out / Bank transactions → "Cash Flow Record" or "Bank Statement"
- Choose the most accurate type — do NOT default to "Income Statement".

Rules for expense_breakdown:
- List each distinct expense LINE ITEM (not subtotals). Return [] if none visible.
- Amounts must be totals across the full period (not per month).

Rules for product_breakdown:
- ONLY populate when detected_document_type is "Sales Record" or the document clearly lists products with quantities/prices.
- For each product row: "product" = name, "units_sold" = quantity, "unit_price" = price per unit, "revenue" = units × price.
- If revenue per product is stated directly, use that value.
- Return [] for all other document types.

Now extract the actual data from the document above and return your JSON:"""


def _normalize_currency(raw: Any) -> str:
    """Force any LLM-returned currency string down to a short symbol like 'K'.
    Prevents 'Zambian Kwacha (K)' + amount = 'Zambian Kwacha (K)18,600' double-prefix.
    """
    if not raw:
        return "K"
    s = str(raw).strip()
    upper = s.upper()
    if "ZMW" in upper or "KWACHA" in upper or upper == "K":
        return "K"
    if upper in {"USD", "US$", "$", "EUR", "€", "GBP", "£"}:
        return "K"
    # Strip parentheticals like "Zambian Kwacha (K)" → "K"
    m = re.search(r"\(([^)]{1,3})\)", s)
    if m:
        return m.group(1)
    return "K"


# ── Raw content extractors ─────────────────────────────────────────────────────

def _extract_excel_as_text(file_bytes: bytes) -> tuple[str, int]:
    """Read all sheets from an xlsx file and return as structured text + row count."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets_text: list[str] = []
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_lines = [f"=== SHEET: {sheet_name} ==="]

        for row in ws.iter_rows(values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            if row_text.strip(" |"):
                sheet_lines.append(row_text)
                total_rows += 1

        if len(sheet_lines) > 1:
            sheets_text.append("\n".join(sheet_lines))

    return "\n\n".join(sheets_text), total_rows


def _parse_excel_by_sheets(file_bytes: bytes, filename: str = "") -> dict[str, Any] | None:
    """
    Sheet-aware structured extraction for multi-sheet Excel workbooks.
    Reads Income Statement / Cash Flow / Debt sheets independently to prevent mixing.
    Returns a complete result dict on success, or None to fall through to the LLM path.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        return None

    names_lower = {s.lower().strip(): s for s in wb.sheetnames}

    def find_sheet(*keywords: str):
        for kw in keywords:
            for nl, original in names_lower.items():
                if kw in nl:
                    return wb[original]
        return None

    ws_income = find_sheet("income statement", "income", "profit & loss", "p&l", "pnl")
    ws_debt   = find_sheet("debt", "loan", "liabilit", "payable")
    ws_cf     = find_sheet("cash flow", "cashflow")

    # Single-sheet workbook: assume the only sheet is the income statement
    if ws_income is None and len(wb.sheetnames) == 1:
        ws_income = wb[wb.sheetnames[0]]

    if ws_income is None:
        return None  # No recognisable Income Statement → let LLM handle it

    # ── cell → float (handles parentheses negatives and K/ZMW prefixes) ─────────
    def to_num(v) -> float | None:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = re.sub(r"[KZMWkzmw\s,]", "", str(v))
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return float(s)
        except (ValueError, TypeError):
            return None

    def label_matches(label: str, keywords: list[str]) -> bool:
        nl = re.sub(r"[^a-z0-9/(). ]+", " ", label.lower()).strip()
        for kw in keywords:
            nk = re.sub(r"[^a-z0-9/(). ]+", " ", kw.lower()).strip()
            if nk == nl or nk in nl:
                return True
        return False

    def find_total_col_idx(ws) -> int | None:
        """
        Scan the first 5 rows (skipping column 0, which holds row labels) to find
        a column whose header contains 'total' or 'sum' but NOT 'average'.
        Returns the 0-based column index, or None if not found.
        """
        TOTAL_KWS = ("total", "sum", "full year", "6-month", "6 month", "period total")
        SKIP_KWS  = ("average", "avg", "mean")
        for row in ws.iter_rows(max_row=5, values_only=True):
            for idx, cell in enumerate(row):
                if idx == 0 or cell is None:
                    continue  # col 0 = row label column
                s = str(cell).strip().lower()
                if any(kw in s for kw in TOTAL_KWS) and not any(kw in s for kw in SKIP_KWS):
                    return idx
        return None

    def find_row_value(ws, keywords: list[str], total_col_idx: int | None = None) -> float | None:
        """
        Scan worksheet rows; return the period total of the best-matching row.

        Priority order:
        1. Cell in the explicitly identified Total column (most reliable).
        2. Sum all numeric cells in the row when the row has ≥ 4 numbers
           (handles multi-month statements that have no separate Total column).
        3. Last numeric cell in the row (single-value fallback).
        """
        candidates: list[tuple[float, bool]] = []
        for row in ws.iter_rows(values_only=True):
            label = next((str(c).strip() for c in row if c is not None and str(c).strip()), None)
            if not label or not label_matches(label, keywords):
                continue

            val: float | None = None

            # Strategy 1 — use the detected Total column
            if total_col_idx is not None and total_col_idx < len(row):
                val = to_num(row[total_col_idx])

            # Strategy 2 — derive total from numeric cells in the row
            if val is None:
                nums = [n for n in (to_num(c) for c in row) if n is not None]
                if nums:
                    # 4+ numbers → multi-month layout; sum gives the period total
                    # 1-3 numbers → single-value or 2-col row; take the last
                    val = sum(nums) if len(nums) >= 4 else nums[-1]

            if val is None:
                continue
            candidates.append((val, "total" in label.lower()))

        if not candidates:
            return None
        # Prefer rows whose label contains "total"
        for val, is_total in candidates:
            if is_total:
                return val
        return candidates[0][0]

    def classify_income_row(label: str) -> str | None:
        """Classify an income-statement line item as revenue/expense/ignore."""
        nl = re.sub(r"[^a-z0-9/(). ]+", " ", label.lower()).strip()
        if not nl:
            return None

        # Generic totals/subtotals often duplicate detail rows and should not be
        # included in derived sums.
        if "total" in nl or "subtotal" in nl or "grand total" in nl:
            return None

        ignore_tokens = (
            "line item",
            "key ratios",
            "analysis",
            "auditor",
            "notes",
            "margin",
            "ratio",
            "profit",
            "loss",
            "gross profit",
            "operating profit",
            "net profit",
            "total revenue",
            "total income",
            "total sales",
            "total expenses",
            "total operating expenses",
            "total other expenses",
            "total cost of goods sold",
            "total costs",
        )
        if any(tok in nl for tok in ignore_tokens):
            return None
        if nl in {"revenue", "operating expenses"}:
            return None

        revenue_tokens = ("sales", "revenue", "income", "turnover", "receipts")
        expense_tokens = (
            "expense",
            "expenses",
            "cost",
            "rent",
            "wage",
            "salary",
            "electricity",
            "water",
            "utility",
            "transport",
            "loan",
            "repayment",
            "airtime",
            "marketing",
            "maintenance",
            "repair",
            "interest",
            "fee",
            "freight",
            "delivery",
            "purchase",
            "supplier",
        )
        if any(tok in nl for tok in revenue_tokens):
            return "revenue"
        if any(tok in nl for tok in expense_tokens):
            return "expense"
        return None

    def row_period_value(row, total_col_idx: int | None = None) -> float | None:
        """Return one row's period value from total column or monthly cells."""
        value: float | None = None
        if total_col_idx is not None and total_col_idx < len(row):
            value = to_num(row[total_col_idx])
        if value is not None:
            return value

        nums = [n for n in (to_num(c) for c in row) if n is not None]
        if not nums:
            return None
        return sum(nums) if len(nums) >= 4 else nums[-1]

    def derive_totals_from_income_rows(
        ws, total_col_idx: int | None = None
    ) -> tuple[float | None, float | None, dict[str, float]]:
        """
        Derive revenue/expenses from detailed line items when explicit total rows
        are missing or formula cells are not available in data-only mode.
        """
        revenue_sum = 0.0
        expense_sum = 0.0
        seen_revenue = False
        seen_expense = False

        for row in ws.iter_rows(values_only=True):
            label = next((str(c).strip() for c in row if c is not None and str(c).strip()), None)
            if not label:
                continue
            kind = classify_income_row(label)
            if kind is None:
                continue

            val = row_period_value(row, total_col_idx)
            if val is None:
                continue
            val = abs(val)
            if kind == "revenue":
                revenue_sum += val
                seen_revenue = True
            elif kind == "expense":
                expense_sum += val
                seen_expense = True

        expense_breakdown: dict[str, float] = {}

        def normalize_expense_label(label: str) -> str:
            clean = re.sub(r"\s+", " ", label.strip())
            return clean.title() if clean.islower() else clean

        for row in ws.iter_rows(values_only=True):
            label = next((str(c).strip() for c in row if c is not None and str(c).strip()), None)
            if not label:
                continue
            if classify_income_row(label) != "expense":
                continue
            val = row_period_value(row, total_col_idx)
            if val is None:
                continue
            expense_breakdown[normalize_expense_label(label)] = round(abs(val), 2)

        return (
            revenue_sum if seen_revenue else None,
            expense_sum if seen_expense else None,
            expense_breakdown,
        )

    def infer_period_description(ws) -> str | None:
        """Infer period label like 'Nov 2025 – Apr 2026' from any leading metadata
        or column headers. Returns the chronological span (earliest – latest)."""
        month_pattern = re.compile(
            r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}\b",
            flags=re.IGNORECASE,
        )
        month_order = {
            m: i + 1
            for i, m in enumerate(
                ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
            )
        }

        def sort_key(token: str) -> tuple[int, int]:
            mtxt, ytxt = token.split()
            return int(ytxt), month_order.get(mtxt[:3].lower(), 0)

        tokens: list[str] = []
        for row in ws.iter_rows(max_row=10, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                text = str(cell)
                if not text.strip():
                    continue
                tokens.extend(m.group(0) for m in month_pattern.finditer(text))

        if not tokens:
            return None
        unique = sorted({t for t in tokens}, key=sort_key)
        if len(unique) == 1:
            return unique[0]
        return f"{unique[0]} – {unique[-1]}"

    def count_non_empty_rows(ws) -> int:
        count = 0
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                count += 1
        return count

    # ── Income Statement ──────────────────────────────────────────────────────────
    # Detect Total column once; all row lookups use it so we never pick up a
    # monthly slice or an average column by accident.
    income_total_col = find_total_col_idx(ws_income)

    revenue = find_row_value(ws_income, [
        "total revenue", "total sales", "total income", "gross revenue",
    ], income_total_col)
    expenses = find_row_value(ws_income, [
        "total expenses", "total costs", "total operating expenses",
        "total expenditure", "total cost of sales",
    ], income_total_col)
    profit = find_row_value(ws_income, [
        "net profit / (loss)", "net profit/(loss)", "net profit", "net loss", "net income",
    ], income_total_col)
    if profit is None:
        # Fall back to operating profit only when net profit is unavailable.
        profit = find_row_value(ws_income, [
            "profit / (loss)", "profit/(loss)", "operating profit",
        ], income_total_col)

    # If formula-based total rows are blank in data-only mode, derive from detail rows.
    derived_revenue, derived_expenses, derived_expense_breakdown = derive_totals_from_income_rows(
        ws_income, income_total_col
    )
    if revenue is None and derived_revenue is not None:
        revenue = derived_revenue
    if expenses is None and derived_expenses is not None:
        expenses = derived_expenses

    if revenue is None:
        return None  # Can't extract without at least a revenue figure

    revenue = abs(revenue)
    if expenses is not None:
        expenses = abs(expenses)
    elif profit is not None:
        expenses = revenue - profit
    else:
        expenses = 0.0

    if profit is None:
        profit = revenue - expenses
    # profit keeps its sign (negative = loss)

    # Some statements expose "total operating expenses" (excluding COGS/finance)
    # while profit is net. Force accounting identity to avoid mixed subtotals.
    mismatch = abs((revenue - expenses) - profit)
    tolerance = max(1.0, abs(revenue) * 0.01)
    if mismatch > tolerance:
        recomputed_expenses = revenue - profit
        if recomputed_expenses >= 0:
            expenses = recomputed_expenses

    # ── Debt sheet ───────────────────────────────────────────────────────────────
    debt = 0.0
    if ws_debt is not None:
        total_debt = find_row_value(ws_debt, [
            "total debt", "total loans", "total liabilities", "total balance",
            "total", "grand total",
        ])
        if total_debt is not None:
            debt = abs(total_debt)
        else:
            # Sum individual data rows, skipping obvious header cells
            HEADER_LABELS = {
                "date", "lender", "creditor", "description", "type",
                "debt type", "loan type", "amount", "balance", "notes",
            }
            for row in ws_debt.iter_rows(values_only=True):
                label = next((str(c).strip() for c in row if c is not None and str(c).strip()), None)
                if not label or label.lower().strip() in HEADER_LABELS:
                    continue
                val = row_last_num(row)
                if val is not None and val > 0:
                    debt += val

    # ── Cash Flow (inflow/outflow only — do NOT override IS revenue/expenses) ────
    cash_inflow, cash_outflow = revenue, expenses
    if ws_cf is not None:
        cf_in = find_row_value(ws_cf, [
            "total cash inflows", "total inflows", "total receipts",
        ])
        cf_out = find_row_value(ws_cf, [
            "total cash outflows", "total outflows", "total payments",
        ])
        if cf_in is not None:
            cash_inflow = abs(cf_in)
        if cf_out is not None:
            cash_outflow = abs(cf_out)

    margin = calculate_profit_margin(revenue, profit)
    if derived_expense_breakdown:
        biggest_cats = [
            c for c, _ in sorted(derived_expense_breakdown.items(), key=lambda x: x[1], reverse=True)[:3]
        ]
    else:
        biggest_cats = []
    sheets_used = {
        "income": ws_income.title,
        "debt":   ws_debt.title if ws_debt else None,
        "cashflow": ws_cf.title if ws_cf else None,
    }
    period_description = infer_period_description(ws_income)
    rows_analyzed = count_non_empty_rows(ws_income)

    result = _blank_result("xlsx", filename, "Income Statement")
    result.update({
        "rows_analyzed":  rows_analyzed,
        "period_description": period_description,
        "date_range": period_description or "Unknown",
        "revenue":        round(revenue, 2),
        "expenses":       round(expenses, 2),
        "profit":         round(profit, 2),
        "debt":           round(debt, 2),
        "cash_inflow":    round(cash_inflow, 2),
        "cash_outflow":   round(cash_outflow, 2),
        "profit_margin":  round(margin, 2),
        "confidence":     "high",
        "used_llm":       False,
        "expenses_breakdown": derived_expense_breakdown,
        "biggest_expense_categories": biggest_cats,
        "summary": (
            f"Structured multi-sheet extraction from '{ws_income.title}' "
            f"(Revenue: K{revenue:,.2f}, Expenses: K{expenses:,.2f}, "
            f"Profit: K{profit:,.2f}, Margin: {margin:.1f}%)."
        ),
        "warnings": [],
        "metadata": {"sheets_used": sheets_used},
    })
    return result


def _extract_csv_as_text(file_bytes: bytes) -> tuple[str, int]:
    frame = pd.read_csv(io.BytesIO(file_bytes))
    return frame.to_string(max_rows=300), len(frame)


def _extract_pdf_as_text(file_bytes: bytes) -> tuple[str, int]:
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(file_bytes))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages), len(pages)


def _truncate(text: str, max_chars: int = 10000) -> str:
    if len(text) <= max_chars:
        return text
    return text[:8000] + "\n\n... [content truncated] ...\n\n" + text[-2000:]


# ── LLM-based parser ───────────────────────────────────────────────────────────

def _sanitize_llm_json(text: str) -> str:
    """Replace arithmetic expressions like '18600 + 46980' with their computed value."""
    import re
    def _eval_expr(m: re.Match) -> str:
        try:
            val = eval(m.group(0), {"__builtins__": {}})  # noqa: S307 — safe: only digits/ops
            return str(val)
        except Exception:
            return m.group(0)
    # Match patterns like: 18600 + 46980  or  18600+46980  (digits, +, -, *, / only)
    return re.sub(r"\d[\d\s]*[\+\-\*\/][\d\s]+\d", _eval_expr, text)


def _llm_parse_document(raw_text: str, document_type: str, filename: str) -> dict[str, Any] | None:
    """Call the LLM to intelligently extract financial data from raw document text.

    Raises exceptions so callers can surface the real error rather than silently falling back.
    """
    from agents import get_chat_model, _extract_json_block
    from langchain_core.messages import HumanMessage, SystemMessage

    llm = get_chat_model()
    if llm is None:
        raise RuntimeError(
            "LLM not configured (set LLM_PROVIDER=amd with AMD_BASE_URL/AMD_MODEL "
            "or LLM_PROVIDER=openai with OPENAI_API_KEY, OPENAI_BASE_URL, MODEL_NAME)."
        )

    user_content = _DOC_PARSE_USER_TEMPLATE.format(
        filename=filename,
        doc_type=document_type,
        content=_truncate(raw_text),
    )

    response = llm.invoke(
        [SystemMessage(content=_DOC_PARSE_SYSTEM), HumanMessage(content=user_content)],
        temperature=0.1,
        max_tokens=1800,
    )
    raw_response = getattr(response, "content", None)
    if not isinstance(raw_response, str) or not raw_response.strip():
        raise ValueError("LLM returned an empty response.")

    sanitized = _sanitize_llm_json(raw_response)
    parsed = _extract_json_block(sanitized)
    if parsed is None:
        raise ValueError(f"Could not parse JSON from LLM response. First 300 chars: {sanitized[:300]!r}")
    return parsed


# ── Result schema ──────────────────────────────────────────────────────────────

def _blank_result(source_type: str, filename: str = "", document_type: str = "Other") -> dict[str, Any]:
    return {
        "document_type": document_type,
        "source_type": source_type,
        "file_name": filename,
        "rows_analyzed": 0,
        "pages_analyzed": 0,
        "characters_analyzed": 0,
        "revenue": 0.0,
        "expenses": 0.0,
        "profit": 0.0,
        "debt": 0.0,
        "cash_inflow": 0.0,
        "cash_outflow": 0.0,
        "profit_margin": 0.0,
        "date_range": "Unknown",
        "period_description": None,
        "detected_dates": [],
        "detected_columns": {},
        "summary": "",
        "warnings": [],
        "preview_rows": [],
        "expenses_breakdown": {},
        "biggest_expense_categories": [],
        "confidence": "low",
        "monthly_breakdown": None,
        "key_observations": [],
        "currency": "K",
        "used_llm": False,
        "metadata": {},
        # Sales-record specific fields (None when not a sales record)
        "product_breakdown": None,
        "top_product": None,
        "slowest_product": None,
    }


def _map_llm_result(
    llm: dict[str, Any],
    source_type: str,
    filename: str,
    rows_analyzed: int = 0,
) -> dict[str, Any]:
    """Map LLM JSON output to the standard document result schema."""
    revenue = float(llm.get("total_revenue") or 0)
    expenses = float(llm.get("total_expenses") or 0)
    profit_raw = llm.get("profit")
    profit = float(profit_raw) if profit_raw is not None else calculate_profit(revenue, expenses)
    debt = float(llm.get("debt_balance") or 0)

    detected_type = llm.get("detected_document_type") or "Other"
    period_desc = llm.get("period_description") or None
    currency = _normalize_currency(llm.get("currency"))

    start = llm.get("period_start")
    end = llm.get("period_end")
    if start and end:
        date_range = f"{start} to {end}"
    elif period_desc:
        date_range = period_desc
    else:
        date_range = "Unknown"

    margin = calculate_profit_margin(revenue, profit)
    confidence = llm.get("confidence", "low")
    warnings = list(llm.get("warnings") or [])
    observations = list(llm.get("key_observations") or [])

    if confidence == "low":
        warnings.insert(0, "Low confidence extraction — please verify figures manually before relying on them.")
    elif confidence == "medium":
        warnings.insert(0, "Partial extraction — some figures may need manual verification.")

    # Build expense breakdown dict (category → amount) from LLM list output
    raw_breakdown_list = llm.get("expense_breakdown") or []
    expenses_breakdown: dict[str, float] = {}
    for item in raw_breakdown_list:
        if isinstance(item, dict):
            cat = str(item.get("category") or "").strip()
            amt = item.get("amount")
            if cat and amt is not None:
                try:
                    expenses_breakdown[cat] = round(float(amt), 2)
                except (TypeError, ValueError):
                    pass

    # Compute largest expense category (Python-side, not LLM guess)
    if expenses_breakdown:
        biggest_cats = [c for c, _ in sorted(expenses_breakdown.items(), key=lambda x: x[1], reverse=True)[:3]]
    else:
        biggest_cats = []

    # Product breakdown (Sales Record documents only)
    raw_products = llm.get("product_breakdown") or []
    product_breakdown: list[dict] | None = None
    top_product: str | None = None
    slowest_product: str | None = None

    if raw_products and isinstance(raw_products, list):
        cleaned: list[dict] = []
        for item in raw_products:
            if not isinstance(item, dict):
                continue
            name = str(item.get("product") or "").strip()
            if not name:
                continue
            units = float(item.get("units_sold") or 0)
            price = float(item.get("unit_price") or 0)
            # Prefer LLM-stated revenue; fall back to units × price
            rev_stated = item.get("revenue")
            rev = float(rev_stated) if rev_stated is not None else round(units * price, 2)
            cleaned.append({"product": name, "units_sold": units, "unit_price": price, "revenue": rev})
        if cleaned:
            by_rev = sorted(cleaned, key=lambda x: x["revenue"], reverse=True)
            product_breakdown = by_rev
            top_product = by_rev[0]["product"]
            slowest_product = by_rev[-1]["product"] if len(by_rev) > 1 else None

    summary = (
        f"LLM-extracted {detected_type.lower()} covering {period_desc or 'an unknown period'}. "
        f"Revenue: {currency}{revenue:,.2f}, Expenses: {currency}{expenses:,.2f}, "
        f"Profit: {currency}{profit:,.2f} ({margin:.1f}% margin)."
    )

    result = _blank_result(source_type, filename, detected_type)
    result.update({
        "rows_analyzed": rows_analyzed,
        "revenue": round(revenue, 2),
        "expenses": round(expenses, 2),
        "profit": round(profit, 2),
        "debt": round(debt, 2),
        "cash_inflow": round(revenue, 2),
        "cash_outflow": round(expenses, 2),
        "profit_margin": round(margin, 2),
        "date_range": date_range,
        "period_description": period_desc,
        "summary": summary,
        "warnings": warnings,
        "expenses_breakdown": expenses_breakdown,
        "biggest_expense_categories": biggest_cats,
        "confidence": confidence,
        "monthly_breakdown": llm.get("monthly_breakdown"),
        "key_observations": observations,
        "currency": currency,
        "used_llm": True,
        "product_breakdown": product_breakdown,
        "top_product": top_product,
        "slowest_product": slowest_product,
    })
    return result


# ── Keyword-based fallback (kept for when LLM is unavailable) ─────────────────

def _normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def _find_matching_column(columns: list[str], keywords: list[str]) -> str | None:
    normalized_map = {column: _normalize_name(column) for column in columns}
    for keyword in keywords:
        nk = _normalize_name(keyword)
        for original, normalized in normalized_map.items():
            if normalized == nk:
                return original
    for keyword in keywords:
        nk = _normalize_name(keyword)
        for original, normalized in normalized_map.items():
            if nk in normalized:
                return original
    return None


def _to_numeric(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace(",", "", regex=False)
        .str.replace(r"[^0-9.\-]", "", regex=True)
        .replace("", pd.NA)
    )
    return pd.to_numeric(cleaned, errors="coerce")


def _extract_date_range(frame: pd.DataFrame, date_column: str | None) -> tuple[str, list[str]]:
    if not date_column:
        return "Unknown", []
    dates = pd.to_datetime(frame[date_column], errors="coerce").dropna()
    if dates.empty:
        return "Unknown", []
    start = dates.min().date().isoformat()
    end = dates.max().date().isoformat()
    sample_dates = sorted({d.date().isoformat() for d in dates.head(5)})
    return f"{start} to {end}", sample_dates


def _document_type_columns(frame: pd.DataFrame, document_type: str) -> dict[str, str | None]:
    columns = frame.columns.tolist()
    return {
        "date": _find_matching_column(columns, ["date", "period", "day", "month"]),
        "description": _find_matching_column(columns, ["description", "details", "notes"]),
        "product": _find_matching_column(columns, ["product", "item", "product name", "item name", "service", "goods"]),
        "category": _find_matching_column(columns, ["category", "type", "transaction type", "class"]),
        "amount": _find_matching_column(columns, ["amount", "value", "total", "price"]),
        "revenue": _find_matching_column(columns, ["revenue", "sales", "income", "total sales"]),
        "expense": _find_matching_column(columns, ["expense", "expenses", "cost", "operating expense", "cash out"]),
        "profit": _find_matching_column(columns, ["profit", "net profit", "gross profit"]),
        "debt": _find_matching_column(columns, ["debt", "loan", "balance", "repayment"]),
        "cash_in": _find_matching_column(columns, ["cash in", "money in", "inflow", "credit"]),
        "cash_out": _find_matching_column(columns, ["cash out", "money out", "outflow", "debit"]),
        "quantity": _find_matching_column(columns, ["quantity", "qty", "units", "units sold", "qty sold"]),
        "unit_price": _find_matching_column(columns, ["unit price", "price per unit", "rate", "price"]),
    }


def _detect_is_sales_record(frame: pd.DataFrame) -> bool:
    """Return True when the DataFrame looks like a product sales breakdown.

    Signals: has a product/item name column AND at least one of units-sold or
    unit-price columns — the hallmark of a sales-by-product table rather than a
    P&L or cash-flow statement.
    """
    cols_lower = [c.lower().strip() for c in frame.columns]
    product_hit = any(
        kw in col
        for col in cols_lower
        for kw in ("product", "item name", "item", "goods", "service")
    )
    unit_hit = any(
        kw in col
        for col in cols_lower
        for kw in ("units sold", "qty sold", "quantity sold", "units", "qty")
    )
    price_hit = any(
        kw in col
        for col in cols_lower
        for kw in ("unit price", "price per unit", "price", "rate")
    )
    revenue_expense_hit = any(
        kw in col
        for col in cols_lower
        for kw in ("revenue", "expenses", "profit", "net profit")
    )
    # Only flag as sales record if there are product signals but NOT
    # a full income-statement layout (revenue + expenses + profit columns).
    return product_hit and (unit_hit or price_hit) and not revenue_expense_hit


REVENUE_KEYWORDS = ["revenue", "sales", "sale", "income", "money in", "inflow", "customer payment", "credit", "deposit"]
EXPENSE_KEYWORDS = ["expense", "expenses", "cost", "costs", "purchase", "purchases", "rent", "stock",
                    "transport", "supplier", "money out", "outflow", "debit", "fees", "utilities", "wages"]
DEBT_KEYWORDS = ["debt", "loan", "borrowed", "payable", "supplier debt", "owed", "repayment", "balance due"]


def _classify_from_text(text: str) -> str:
    lowered = text.lower()
    if any(k in lowered for k in DEBT_KEYWORDS):
        return "debt"
    if any(k in lowered for k in REVENUE_KEYWORDS):
        return "revenue"
    if any(k in lowered for k in EXPENSE_KEYWORDS):
        return "expense"
    return "unknown"


def classify_transaction_row(row: dict[str, Any], document_type: str) -> str:
    text = " ".join(str(v) for v in row.values() if v is not None and str(v).strip())
    base = _classify_from_text(text)
    document_type = document_type or "Other"
    if document_type == "Sales Record":
        return "revenue"
    if document_type == "Expense Record":
        return "expense"
    if document_type == "Receipts":
        return "expense" if "sales receipt" not in text.lower() else "revenue"
    if document_type == "Loan or Debt Record":
        return "debt"
    if document_type == "Inventory/Stock Record":
        return "expense"
    if document_type == "Cash Flow Record":
        return base
    if document_type in {"Bank Statement", "Mobile Money Statement"} and base == "unknown":
        tx_type = str(row.get("Transaction Type", row.get("Type", ""))).lower()
        if "credit" in tx_type or "deposit" in tx_type:
            return "revenue"
        if "debit" in tx_type or "withdraw" in tx_type or "cash-out" in tx_type:
            return "expense"
    return base


def _keyword_parse_frame(data: pd.DataFrame, document_type: str) -> dict[str, Any]:
    """Original keyword-based parser — used as fallback when LLM is unavailable."""
    result = _blank_result("table", document_type=document_type)
    if data.empty:
        result["summary"] = "The uploaded file did not contain any rows to analyze."
        result["warnings"].append("No rows were found in the uploaded document.")
        return result

    frame = data.copy()
    frame.columns = [str(c).strip() for c in frame.columns]
    detected = _document_type_columns(frame, document_type)
    result["detected_columns"] = {k: v for k, v in detected.items() if v}
    result["rows_analyzed"] = int(len(frame))
    result["preview_rows"] = frame.head(5).fillna("").to_dict(orient="records")

    date_range, detected_dates = _extract_date_range(frame, detected["date"])
    result["date_range"] = date_range
    result["detected_dates"] = detected_dates

    revenue = expenses = debt = cash_inflow = cash_outflow = 0.0
    warnings: list[str] = []

    if detected["revenue"] and detected["revenue"] != detected["cash_in"]:
        revenue += float(_to_numeric(frame[detected["revenue"]]).fillna(0).sum())
    if detected["expense"] and detected["expense"] != detected["cash_out"]:
        expenses += float(_to_numeric(frame[detected["expense"]]).fillna(0).sum())
    if detected["debt"]:
        debt += float(_to_numeric(frame[detected["debt"]]).fillna(0).sum())
    if detected["cash_in"]:
        cash_inflow += float(_to_numeric(frame[detected["cash_in"]]).fillna(0).sum())
        revenue += cash_inflow if document_type in {"Cash Flow Record", "Bank Statement", "Mobile Money Statement"} else 0
    if detected["cash_out"]:
        cash_outflow += float(_to_numeric(frame[detected["cash_out"]]).fillna(0).sum())
        expenses += cash_outflow if document_type in {"Cash Flow Record", "Bank Statement", "Mobile Money Statement"} else 0

    amount_column = detected["amount"]
    if document_type == "Expense Record" and amount_column and not detected["expense"]:
        expenses += float(_to_numeric(frame[amount_column]).fillna(0).sum())
    elif document_type == "Loan or Debt Record" and amount_column and not detected["debt"]:
        debt += float(_to_numeric(frame[amount_column]).fillna(0).sum())

    unclassified = 0
    expense_breakdown: dict[str, float] = {}
    if amount_column:
        amounts = _to_numeric(frame[amount_column]).fillna(0)
        for index, amount in amounts.items():
            if amount == 0:
                continue
            row_dict = frame.loc[index].to_dict()
            classification = classify_transaction_row(row_dict, document_type)
            if document_type in {"Income Statement", "Other"} and classification == "revenue" and not detected["revenue"]:
                revenue += abs(float(amount))
            elif document_type in {"Income Statement", "Other"} and classification == "expense" and not detected["expense"]:
                expenses += abs(float(amount))
            elif classification == "debt" and not detected["debt"]:
                debt += abs(float(amount))
            elif document_type in {"Bank Statement", "Mobile Money Statement", "Cash Flow Record"}:
                if classification == "revenue" and not detected["cash_in"]:
                    cash_inflow += abs(float(amount))
                    revenue += abs(float(amount))
                elif classification == "expense" and not detected["cash_out"]:
                    cash_outflow += abs(float(amount))
                    expenses += abs(float(amount))
                elif classification == "debt" and not detected["debt"]:
                    debt += abs(float(amount))
            elif document_type == "Sales Record" and not detected["revenue"]:
                revenue += abs(float(amount))
            elif document_type in {"Expense Record", "Receipts", "Inventory/Stock Record"} and not detected["expense"]:
                expenses += abs(float(amount))
            else:
                unclassified += 1

            category_column = detected["category"]
            if classification == "expense" and category_column and pd.notna(frame.at[index, category_column]):
                cat = str(frame.at[index, category_column]).strip()
                expense_breakdown[cat] = expense_breakdown.get(cat, 0.0) + abs(float(amount))

    if unclassified:
        warnings.append(f"{unclassified} row(s) could not be classified for this document type.")

    if document_type == "Sales Record" and revenue == 0 and detected["quantity"] and detected["unit_price"]:
        revenue = float((_to_numeric(frame[detected["quantity"]]) * _to_numeric(frame[detected["unit_price"]])).fillna(0).sum())

    if detected["profit"]:
        profit = float(_to_numeric(frame[detected["profit"]]).fillna(0).sum())
        warnings.append("Profit was read directly from the document.")
    else:
        profit = calculate_profit(revenue, expenses)

    margin = calculate_profit_margin(revenue, profit)
    biggest = [n for n, _ in sorted(expense_breakdown.items(), key=lambda x: x[1], reverse=True)[:3]]

    # ── Product-level extraction for Sales Record documents ───────────────────
    product_breakdown: list[dict] | None = None
    top_product: str | None = None
    slowest_product: str | None = None

    if document_type == "Sales Record" and detected.get("product"):
        prod_col  = detected["product"]
        units_col = detected.get("quantity")
        price_col = detected.get("unit_price")
        amt_col   = detected.get("amount")

        rows: list[dict] = []
        for _, row in frame.iterrows():
            name = str(row.get(prod_col, "")).strip()
            if not name or name.lower() in ("nan", "none", "total", "grand total"):
                continue
            units = (
                float(_to_numeric(pd.Series([row.get(units_col)])).fillna(0).iloc[0])
                if units_col else 0.0
            )
            price = (
                float(_to_numeric(pd.Series([row.get(price_col)])).fillna(0).iloc[0])
                if price_col else 0.0
            )
            rev = units * price if price > 0 and units > 0 else (
                abs(float(_to_numeric(pd.Series([row.get(amt_col)])).fillna(0).iloc[0]))
                if amt_col else 0.0
            )
            if rev > 0 or units > 0:
                rows.append({"product": name, "units_sold": units, "unit_price": price, "revenue": rev})

        if rows:
            by_rev = sorted(rows, key=lambda x: x["revenue"], reverse=True)
            product_breakdown = by_rev
            top_product = by_rev[0]["product"]
            slowest_product = by_rev[-1]["product"] if len(by_rev) > 1 else None

            # Derive total revenue from product rows if not yet captured
            if revenue == 0:
                revenue = round(sum(r["revenue"] for r in rows), 2)
                profit  = calculate_profit(revenue, expenses)
                margin  = calculate_profit_margin(revenue, profit)

    result.update({
        "revenue": round(revenue, 2),
        "expenses": round(expenses, 2),
        "profit": round(profit, 2),
        "debt": round(debt, 2),
        "cash_inflow": round(cash_inflow or revenue, 2),
        "cash_outflow": round(cash_outflow or expenses, 2),
        "profit_margin": round(margin, 2),
        "warnings": warnings,
        "biggest_expense_categories": biggest,
        "confidence": "low",
        "used_llm": False,
        "product_breakdown": product_breakdown,
        "top_product": top_product,
        "slowest_product": slowest_product,
    })
    result["summary"] = f"Keyword-based fallback parser used for this {document_type.lower()}. LLM unavailable."
    return result


# ── Public parse functions ─────────────────────────────────────────────────────

def _classify_llm_error(exc: Exception) -> str:
    """Return a user-friendly message for known LLM failure modes."""
    msg = str(exc)
    if "429" in msg or "rate_limit" in msg.lower() or "Rate limit" in msg:
        import re
        wait = re.search(r"try again in (\S+)", msg)
        wait_str = f" Please try again in {wait.group(1)}." if wait else " Please try again later."
        return (
            f"AI daily usage limit reached.{wait_str} "
            "Document parsing will use the keyword-based fallback method."
        )
    if "401" in msg or "auth" in msg.lower() or "api_key" in msg.lower():
        return "AI service key invalid or missing. Check your API key configuration."
    if "timeout" in msg.lower() or "connect" in msg.lower():
        return "LLM request timed out. Check your internet connection and try again."
    return f"LLM error: {msg[:200]}"


def parse_excel_file(uploaded_file: Any, document_type: str = "Other") -> dict[str, Any]:
    file_bytes = uploaded_file.getvalue()
    filename = uploaded_file.name
    extraction_error: str | None = None
    llm_error: str | None = None

    # Stage 0: sheet-aware structured extraction for multi-sheet workbooks
    try:
        structured = _parse_excel_by_sheets(file_bytes, filename)
        if structured is not None:
            return structured
    except Exception:
        pass  # Non-critical — fall through to LLM path

    # Stage 1: extract all sheets as structured text
    raw_text: str = ""
    rows: int = 0
    try:
        raw_text, rows = _extract_excel_as_text(file_bytes)
    except Exception as exc:
        extraction_error = str(exc)

    # Auto-correct document_type for product-level spreadsheets
    try:
        _xl_frame = pd.read_excel(io.BytesIO(file_bytes), nrows=5)
        if document_type in ("Income Statement", "Other") and _detect_is_sales_record(_xl_frame):
            document_type = "Sales Record"
    except Exception:
        pass

    # Stage 2: LLM parse — errors are surfaced, not swallowed
    if raw_text.strip():
        print("=== RAW EXCEL CONTENT BEING SENT TO LLM ===")
        print(raw_text[:2000])
        print("=== END RAW CONTENT ===")
        try:
            llm = _llm_parse_document(raw_text, document_type, filename)
            if llm:
                return _map_llm_result(llm, "xlsx", filename, rows)
        except Exception as exc:
            llm_error = _classify_llm_error(exc)

    # Fallback: read first sheet only with pandas keyword parser
    try:
        frame = pd.read_excel(io.BytesIO(file_bytes))
        if document_type in ("Income Statement", "Other") and _detect_is_sales_record(frame):
            document_type = "Sales Record"
        result = _keyword_parse_frame(frame, document_type)
        result["source_type"] = "xlsx"
        result["file_name"] = filename
        msgs: list[str] = []
        if llm_error:
            msgs.append(llm_error)
        if extraction_error:
            msgs.append(f"Sheet extraction error: {extraction_error}")
        if not llm_error and not extraction_error:
            msgs.append("Keyword-based fallback used — numbers may not be accurate.")
        result["warnings"] = msgs + result.get("warnings", [])
        return result
    except Exception as exc:
        result = _blank_result("xlsx", filename, document_type)
        if llm_error:
            result["warnings"].append(llm_error)
        if extraction_error:
            result["warnings"].append(f"Sheet extraction error: {extraction_error}")
        result["warnings"].append(f"Could not read Excel file: {exc}")
        return result


def parse_csv_file(uploaded_file: Any, document_type: str = "Other") -> dict[str, Any]:
    file_bytes = uploaded_file.getvalue()
    filename = uploaded_file.name

    try:
        frame = pd.read_csv(io.BytesIO(file_bytes))
        # Auto-correct to Sales Record when columns signal a product breakdown
        if document_type in ("Income Statement", "Other") and _detect_is_sales_record(frame):
            document_type = "Sales Record"
        raw_text = frame.to_string(max_rows=300)
        llm = _llm_parse_document(raw_text, document_type, filename)
        if llm:
            return _map_llm_result(llm, "csv", filename, len(frame))
    except Exception:
        pass

    try:
        frame = pd.read_csv(io.BytesIO(file_bytes))
        if document_type in ("Income Statement", "Other") and _detect_is_sales_record(frame):
            document_type = "Sales Record"
        result = _keyword_parse_frame(frame, document_type)
        result["source_type"] = "csv"
        result["file_name"] = filename
        result["warnings"].insert(0, "LLM parsing unavailable — used keyword-based fallback.")
        return result
    except Exception as exc:
        result = _blank_result("csv", filename, document_type)
        result["warnings"].append(f"Could not read CSV file: {exc}")
        return result


def parse_text_file(uploaded_file: Any, document_type: str = "Other") -> dict[str, Any]:
    content = uploaded_file.getvalue().decode("utf-8", errors="ignore")
    filename = uploaded_file.name

    try:
        llm = _llm_parse_document(content, document_type, filename)
        if llm:
            return _map_llm_result(llm, "txt", filename, content.count("\n"))
    except Exception:
        pass

    parsed = parse_business_input(content)
    result = _blank_result("txt", filename, document_type)
    result.update({
        "characters_analyzed": len(content),
        "revenue": parsed["revenue"],
        "expenses": parsed["expenses"],
        "profit": round(parsed["revenue"] - parsed["expenses"], 2),
        "debt": parsed["debt"],
        "cash_inflow": parsed["revenue"],
        "cash_outflow": parsed["expenses"],
        "summary": f"Text file interpreted as {document_type.lower()} using keyword extraction.",
        "metadata": {
            "location": parsed.get("location"),
            "business_type": parsed.get("business_type"),
        },
    })
    return result


def parse_pdf_file(uploaded_file: Any, document_type: str = "Other") -> dict[str, Any]:
    result = _blank_result("pdf", uploaded_file.name, document_type)
    result["warnings"].append("PDF support is experimental. For best results, upload CSV or Excel.")
    try:
        from pypdf import PdfReader
    except Exception:
        result["summary"] = "PDF parsing unavailable — pypdf not installed."
        result["warnings"].append("Install pypdf to enable PDF text extraction.")
        return result

    try:
        raw_text, page_count = _extract_pdf_as_text(uploaded_file.getvalue())
        result["pages_analyzed"] = page_count
        result["characters_analyzed"] = len(raw_text)

        llm = _llm_parse_document(raw_text, document_type, uploaded_file.name)
        if llm:
            mapped = _map_llm_result(llm, "pdf", uploaded_file.name, page_count)
            mapped["pages_analyzed"] = page_count
            mapped["characters_analyzed"] = len(raw_text)
            mapped["warnings"] = result["warnings"] + mapped["warnings"]
            return mapped

        # Fallback: keyword parse on extracted text
        parsed = parse_business_input(raw_text)
        result.update({
            "revenue": parsed["revenue"],
            "expenses": parsed["expenses"],
            "profit": round(parsed["revenue"] - parsed["expenses"], 2),
            "debt": parsed["debt"],
            "cash_inflow": parsed["revenue"],
            "cash_outflow": parsed["expenses"],
            "summary": f"PDF read as text and interpreted as {document_type.lower()}.",
            "metadata": {
                "location": parsed.get("location"),
                "business_type": parsed.get("business_type"),
            },
        })
    except Exception as exc:
        result["warnings"].append(f"PDF extraction failed: {exc}")
    return result


def parse_uploaded_file(uploaded_file: Any, document_type: str = "Other") -> dict[str, Any]:
    suffix = uploaded_file.name.lower().rsplit(".", maxsplit=1)[-1] if "." in uploaded_file.name else ""
    if suffix == "csv":
        return parse_csv_file(uploaded_file, document_type)
    if suffix in {"xlsx", "xlsm"}:
        return parse_excel_file(uploaded_file, document_type)
    if suffix == "txt":
        return parse_text_file(uploaded_file, document_type)
    if suffix == "pdf":
        return parse_pdf_file(uploaded_file, document_type)

    result = _blank_result(suffix or "unknown", uploaded_file.name, document_type)
    result["summary"] = "The uploaded file format is not supported for document analysis."
    result["warnings"].append("Please upload a CSV, Excel (.xlsx), TXT, or PDF file.")
    return result


def summarize_document_records(data: dict[str, Any], document_type: str | None = None) -> str:
    active_type = document_type or data.get("document_type", "Other")
    period = data.get("period_description") or data.get("date_range") or "unknown period"
    obs = (" Observations: " + "; ".join(data.get("key_observations", [])[:2])) if data.get("key_observations") else ""
    return (
        f"{data.get('summary', '').strip()} "
        f"Document type: {active_type}. Period: {period}. "
        f"Revenue: {data.get('revenue', 0):,.2f}. "
        f"Expenses: {data.get('expenses', 0):,.2f}. "
        f"Profit: {data.get('profit', 0):,.2f}. "
        f"Debt: {data.get('debt', 0):,.2f}."
        f"{obs}"
    ).strip()
